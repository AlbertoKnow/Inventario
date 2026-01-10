#!/usr/bin/env python
"""
Script para generar un archivo Excel de prueba para la importación masiva.
Genera datos de ejemplo con algunas filas válidas y algunas con errores para testing.
"""

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from datetime import datetime, timedelta

# Crear workbook
wb = Workbook()
ws = wb.active
ws.title = "Datos de Prueba"

# Headers
headers = [
    'serie', 'nombre', 'area', 'tipo_item', 'precio', 'fecha_adquisicion',
    'descripcion', 'ambiente_codigo', 'estado', 'garantia_hasta',
    'observaciones', 'lote_codigo', 'es_leasing', 'leasing_empresa',
    'leasing_contrato', 'leasing_vencimiento',
    'marca', 'modelo', 'procesador', 'generacion_procesador',
    'ram_total_gb', 'ram_configuracion', 'ram_tipo',
    'almacenamiento_gb', 'almacenamiento_tipo', 'sistema_operativo'
]

# Agregar headers
for idx, header in enumerate(headers, 1):
    cell = ws.cell(row=1, column=idx, value=header)
    cell.font = Font(bold=True)
    cell.fill = PatternFill(start_color="C8102E", end_color="C8102E", fill_type="solid")
    cell.font = Font(bold=True, color="FFFFFF")

# Datos de prueba - SISTEMAS (válidos)
items_sistemas = [
    [
        'SN-SYS-001', 'Laptop HP EliteBook 840 G9', 'sistemas', 'Laptop', '4200.00', '2026-01-10',
        'Laptop corporativa para desarrollo', '', 'nuevo', '2028-01-10',
        'En buen estado', '', 'NO', '', '', '',
        'HP', 'EliteBook 840 G9', 'Intel Core i7-1265U', '12th Gen',
        '16', '2x8GB', 'DDR4', '512', 'NVMe', 'Windows 11 Pro'
    ],
    [
        'SN-SYS-002', 'Laptop Dell Latitude 5430', 'sistemas', 'Laptop', '3800.00', '2026-01-10',
        'Laptop para diseño gráfico', '', 'nuevo', '2028-01-10',
        '', '', 'NO', '', '', '',
        'Dell', 'Latitude 5430', 'Intel Core i5-1245U', '12th Gen',
        '16', '1x16GB', 'DDR4', '256', 'SSD', 'Windows 11 Pro'
    ],
    [
        'SN-SYS-003', 'Monitor Dell UltraSharp 27"', 'sistemas', 'Monitor', '650.00', '2026-01-10',
        'Monitor para estación de trabajo', '', 'nuevo', '2027-01-10',
        '', '', 'NO', '', '', '',
        'Dell', 'U2722D', '', '',
        '', '', '', '', '', ''
    ],
]

# Datos de prueba - OPERACIONES (válidos)
items_operaciones = [
    [
        'SN-OPE-001', 'Silla ergonómica Herman Miller', 'operaciones', 'Silla', '850.00', '2026-01-10',
        'Silla con soporte lumbar ajustable', '', 'nuevo', '',
        '', '', 'NO', '', '', '',
        '', '', '', '', '', '', '', '', '', ''
    ],
    [
        'SN-OPE-002', 'Escritorio regulable en altura', 'operaciones', 'Escritorio', '1200.00', '2026-01-10',
        'Escritorio eléctrico regulable', '', 'nuevo', '2028-01-10',
        '', '', 'NO', '', '', '',
        '', '', '', '', '', '', '', '', '', ''
    ],
]

# Datos de prueba - LABORATORIO (válidos)
items_laboratorio = [
    [
        'SN-LAB-001', 'Osciloscopio Digital Tektronix', 'laboratorio', 'Osciloscopio', '8500.00', '2026-01-10',
        'Osciloscopio de 4 canales 200MHz', '', 'nuevo', '2029-01-10',
        '', '', 'NO', '', '', '',
        'Tektronix', 'TBS2204B', '', '',
        '', '', '', '', '', ''
    ],
]

# Datos con ERRORES para testing
items_con_errores = [
    # Serie duplicada (asumiendo que SN-SYS-001 ya existe en BD)
    # Comentar esta línea si es la primera vez que importas
    # ['SN-SYS-001', 'Item duplicado', 'sistemas', 'Laptop', '1000.00', '2026-01-10', '', '', 'nuevo', '', '', '', 'NO', '', '', '', '', '', '', '', '', '', '', '', '', ''],

    # Área inválida
    ['SN-ERR-001', 'Item con área inválida', 'inventario', 'Laptop', '1000.00', '2026-01-10', '', '', 'nuevo', '', '', '', 'NO', '', '', '', '', '', '', '', '', '', '', '', '', ''],

    # Precio inválido
    ['SN-ERR-002', 'Item con precio inválido', 'sistemas', 'Laptop', 'ABC', '2026-01-10', '', '', 'nuevo', '', '', '', 'NO', '', '', '', '', '', '', '', '', '', '', '', '', ''],

    # Fecha inválida
    ['SN-ERR-003', 'Item con fecha inválida', 'sistemas', 'Laptop', '1000.00', '32/13/2026', '', '', 'nuevo', '', '', '', 'NO', '', '', '', '', '', '', '', '', '', '', '', '', ''],
]

# Agregar todos los datos
row = 2
for item in items_sistemas + items_operaciones + items_laboratorio + items_con_errores:
    for col, value in enumerate(item, 1):
        ws.cell(row=row, column=col, value=value)
    row += 1

# Ajustar ancho de columnas
for idx in range(1, len(headers) + 1):
    ws.column_dimensions[ws.cell(row=1, column=idx).column_letter].width = 18

# Guardar archivo
filename = 'items_prueba_importacion.xlsx'
wb.save(filename)
print(f"[OK] Archivo '{filename}' generado exitosamente.")
print(f"Total de items: {len(items_sistemas) + len(items_operaciones) + len(items_laboratorio) + len(items_con_errores)}")
print(f"   - Sistemas: {len(items_sistemas)}")
print(f"   - Operaciones: {len(items_operaciones)}")
print(f"   - Laboratorio: {len(items_laboratorio)}")
print(f"   - Con errores: {len(items_con_errores)}")
print("\n[INFO] Puedes usar este archivo para probar la importacion masiva.")
