from django.shortcuts import render, redirect, get_object_or_404
from django.views.generic import (
    ListView, DetailView, CreateView, UpdateView, DeleteView, TemplateView, View
)
from django.contrib.auth.mixins import LoginRequiredMixin, UserPassesTestMixin
from django.contrib.auth.models import User
from django.urls import reverse_lazy, reverse
from django.db.models import Q, Sum, Count
from django.contrib import messages
from django.http import JsonResponse, HttpResponse, Http404
from django.utils import timezone
from datetime import timedelta, date

from .models import (
    Area, Campus, Sede, Pabellon, Ambiente, TipoItem, Item, EspecificacionesSistemas,
    Movimiento, MovimientoItem, HistorialCambio, Notificacion, PerfilUsuario,
    Proveedor, Contrato, AnexoContrato, Lote, Mantenimiento,
    MarcaEquipo, ModeloEquipo, ProcesadorEquipo, Colaborador
)
from .forms import (
    ItemForm, ItemSistemasForm, MovimientoForm, TipoItemForm, AmbienteForm,
    CampusForm, SedeForm, PabellonForm, MantenimientoForm, MantenimientoFinalizarForm,
    MantenimientoLoteForm
)
from .signals import set_current_user
from .ratelimit import RateLimitMixin


# ============================================================================
# MIXINS DE PERMISOS
# ============================================================================

class PerfilRequeridoMixin(LoginRequiredMixin):
    """Mixin que verifica que el usuario tenga perfil."""
    
    def dispatch(self, request, *args, **kwargs):
        if request.user.is_authenticated:
            # Establecer usuario actual para signals
            set_current_user(request.user)
            
            # Crear perfil si no existe
            if not hasattr(request.user, 'perfil'):
                # Superusuarios obtienen rol admin automáticamente
                rol = 'admin' if request.user.is_superuser else 'operador'
                PerfilUsuario.objects.create(usuario=request.user, rol=rol)
        return super().dispatch(request, *args, **kwargs)
    
    def get_user_area(self):
        """Obtiene el área del usuario actual."""
        if hasattr(self.request.user, 'perfil'):
            return self.request.user.perfil.area
        return None
    
    def get_user_rol(self):
        """Obtiene el rol del usuario actual."""
        if hasattr(self.request.user, 'perfil'):
            return self.request.user.perfil.rol
        return 'operador'
    
    def es_admin(self):
        """Verifica si el usuario es admin."""
        return self.get_user_rol() == 'admin'
    
    def es_supervisor(self):
        """Verifica si el usuario es supervisor."""
        return self.get_user_rol() == 'supervisor'


class SupervisorRequeridoMixin(PerfilRequeridoMixin, UserPassesTestMixin):
    """Solo supervisores y admins pueden acceder."""
    
    def test_func(self):
        return self.get_user_rol() in ['admin', 'supervisor']


class AdminRequeridoMixin(PerfilRequeridoMixin, UserPassesTestMixin):
    """Solo admins pueden acceder."""

    def test_func(self):
        return self.get_user_rol() == 'admin'


class AlmacenRequeridoMixin(PerfilRequeridoMixin, UserPassesTestMixin):
    """Solo admin, gerente y almacén pueden crear/editar/eliminar items."""

    def test_func(self):
        return self.get_user_rol() in ['admin', 'gerente', 'almacen']


class CampusFilterMixin:
    """
    Mixin para filtrar querysets según los campus permitidos del usuario.

    Uso:
    - Admin: ve todo (sin filtro)
    - Supervisor: ve solo los campus que tiene asignados
    - Operador: ve solo su campus asignado
    """

    def get_campus_permitidos(self):
        """Retorna los campus que el usuario puede ver."""
        if hasattr(self.request.user, 'perfil'):
            return self.request.user.perfil.get_campus_permitidos()
        return Campus.objects.none()

    def filtrar_por_campus(self, queryset, campo_campus='ambiente__pabellon__sede__campus'):
        """
        Filtra un queryset según los campus permitidos del usuario.

        Args:
            queryset: El queryset a filtrar
            campo_campus: El campo que relaciona con campus (ej: 'ambiente__pabellon__sede__campus')

        Returns:
            Queryset filtrado
        """
        if not hasattr(self.request.user, 'perfil'):
            return queryset.none()

        perfil = self.request.user.perfil

        # Admin ve todo
        if perfil.rol == 'admin':
            return queryset

        # Obtener IDs de campus permitidos
        campus_ids = list(self.get_campus_permitidos().values_list('id', flat=True))

        if not campus_ids:
            return queryset.none()

        # Construir filtro dinámico
        filtro = {f'{campo_campus}__in': campus_ids}
        return queryset.filter(**filtro)


# ============================================================================
# VISTA DE INICIO
# ============================================================================

class HomeView(TemplateView):
    """Página de inicio - accesible a todos."""
    template_name = 'index.html'
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        
        if self.request.user.is_authenticated:
            user = self.request.user
            perfil = getattr(user, 'perfil', None)
            
            # Filtrar por área si no es admin
            items = Item.objects.all()
            if perfil and perfil.rol != 'admin' and perfil.area:
                items = items.filter(area=perfil.area)
            
            context['total_items'] = items.count()
            context['items_nuevos'] = items.filter(estado='nuevo').count()
            context['items_instalados'] = items.filter(estado='instalado').count()
            context['items_dañados'] = items.filter(estado='dañado').count()
            context['items_sin_asignar'] = items.filter(colaborador_asignado__isnull=True).count()
            
            # Notificaciones no leídas
            context['notificaciones_count'] = Notificacion.objects.filter(
                usuario=user, leida=False
            ).count()
            
            # Movimientos pendientes (para supervisores)
            if perfil and perfil.rol in ['admin', 'supervisor']:
                pendientes = Movimiento.objects.filter(estado='pendiente')
                if perfil.rol == 'supervisor' and perfil.area:
                    pendientes = pendientes.filter(item__area=perfil.area)
                context['movimientos_pendientes'] = pendientes.count()
        
        return context


# ============================================================================
# DASHBOARD
# ============================================================================

class DashboardView(PerfilRequeridoMixin, CampusFilterMixin, TemplateView):
    """Dashboard principal del inventario."""
    template_name = 'productos/dashboard.html'

    def get_context_data(self, **kwargs):
        from django.db.models import Case, When, IntegerField, Value
        from django.core.cache import cache
        from datetime import timedelta

        context = super().get_context_data(**kwargs)

        user = self.request.user
        perfil = getattr(user, 'perfil', None)
        hoy = timezone.now().date()

        # Base queryset - filtrado por campus permitidos
        items = Item.objects.all()
        items = self.filtrar_por_campus(items, 'ambiente__pabellon__sede__campus')

        # Filtrar por área si no es admin
        if perfil and perfil.rol != 'admin' and perfil.area:
            items = items.filter(area=perfil.area)

        # OPTIMIZACIÓN: Obtener todos los conteos en una sola query usando agregación
        fecha_limite_garantia = hoy + timedelta(days=30)
        stats = items.aggregate(
            total=Count('id'),
            estado_backup=Count('id', filter=Q(estado='backup')),
            estado_custodia=Count('id', filter=Q(estado='custodia')),
            estado_instalado=Count('id', filter=Q(estado='instalado')),
            estado_mantenimiento=Count('id', filter=Q(estado='mantenimiento')),
            estado_garantia=Count('id', filter=Q(estado='garantia')),
            estado_baja=Count('id', filter=Q(estado='baja')),
            sin_asignar=Count('id', filter=Q(colaborador_asignado__isnull=True)),
            sin_codigo_utp=Count('id', filter=Q(codigo_utp='PENDIENTE')),
            garantias_proximas=Count('id', filter=Q(
                garantia_hasta__lte=fecha_limite_garantia,
                garantia_hasta__gte=hoy
            )),
        )

        context['total_items'] = stats['total']
        context['items_por_estado'] = {
            'backup': stats['estado_backup'],
            'custodia': stats['estado_custodia'],
            'instalado': stats['estado_instalado'],
            'mantenimiento': stats['estado_mantenimiento'],
            'garantia': stats['estado_garantia'],
            'baja': stats['estado_baja'],
        }
        context['items_sin_asignar'] = stats['sin_asignar']
        context['items_sin_codigo_utp'] = stats['sin_codigo_utp']
        context['garantias_proximas'] = stats['garantias_proximas']

        # Items por área (filtrado por campus) - una sola query
        campus_ids = list(self.get_campus_permitidos().values_list('id', flat=True))
        if perfil and perfil.rol == 'admin':
            context['items_por_area'] = Area.objects.annotate(
                total=Count('items')
            ).values('nombre', 'total')
        else:
            context['items_por_area'] = Area.objects.annotate(
                total=Count('items', filter=Q(items__ambiente__pabellon__sede__campus_id__in=campus_ids))
            ).values('nombre', 'total')

        # Últimos movimientos - limitado a últimos 30 días y optimizado
        fecha_limite_movimientos = timezone.now() - timedelta(days=30)
        movimientos = Movimiento.objects.filter(
            fecha_solicitud__gte=fecha_limite_movimientos
        ).select_related(
            'item', 'item__tipo_item', 'solicitado_por'
        )
        movimientos = self.filtrar_por_campus(movimientos, 'item__ambiente__pabellon__sede__campus')
        if perfil and perfil.rol != 'admin' and perfil.area:
            movimientos = movimientos.filter(item__area=perfil.area)
        context['ultimos_movimientos'] = movimientos[:10]

        # Movimientos pendientes de aprobar
        if perfil and perfil.rol in ['admin', 'supervisor', 'gerente']:
            pendientes = Movimiento.objects.filter(estado='pendiente').select_related(
                'item', 'item__tipo_item', 'solicitado_por'
            )
            pendientes = self.filtrar_por_campus(pendientes, 'item__ambiente__pabellon__sede__campus')
            if perfil.rol == 'supervisor' and perfil.area:
                pendientes = pendientes.filter(item__area=perfil.area)
            context['movimientos_pendientes'] = pendientes[:5]

        # Notificaciones no leídas
        context['notificaciones'] = Notificacion.objects.filter(
            usuario=user, leida=False
        ).only('id', 'titulo', 'tipo', 'fecha', 'url')[:5]

        # Mantenimientos - optimizado con una sola agregación
        mantenimientos_base = Mantenimiento.objects.all()
        mantenimientos_base = self.filtrar_por_campus(mantenimientos_base, 'item__ambiente__pabellon__sede__campus')
        if perfil and perfil.rol != 'admin' and perfil.area:
            mantenimientos_base = mantenimientos_base.filter(item__area=perfil.area)

        mant_stats = mantenimientos_base.aggregate(
            pendientes=Count('id', filter=Q(estado='pendiente')),
            vencidos=Count('id', filter=Q(estado='pendiente', fecha_programada__lt=hoy)),
        )
        context['mantenimientos_pendientes'] = mant_stats['pendientes']
        context['mantenimientos_vencidos'] = mant_stats['vencidos']

        context['ultimos_mantenimientos'] = mantenimientos_base.select_related(
            'item', 'item__tipo_item'
        ).order_by('-fecha_programada')[:5]

        # Campus del usuario para mostrar en dashboard
        context['campus_usuario'] = self.get_campus_permitidos()

        return context


# ============================================================================
# VISTAS DE ITEMS
# ============================================================================

class ItemListView(PerfilRequeridoMixin, CampusFilterMixin, ListView):
    """Lista de ítems del inventario."""
    model = Item
    template_name = 'productos/item_list.html'
    context_object_name = 'items'
    paginate_by = 20

    def get_queryset(self):
        queryset = Item.objects.select_related(
            'area', 'tipo_item', 'ambiente', 'colaborador_asignado',
            'ambiente__pabellon__sede__campus'
        )

        perfil = getattr(self.request.user, 'perfil', None)

        # Filtrar por campus según permisos del usuario
        queryset = self.filtrar_por_campus(queryset, 'ambiente__pabellon__sede__campus')

        # Filtrar por área si no es admin - SIEMPRE aplicar esta restricción
        if perfil and perfil.rol != 'admin' and perfil.area:
            # Operadores y supervisores SOLO ven su área, no pueden filtrar otras
            queryset = queryset.filter(area=perfil.area)
        else:
            # Solo admin puede filtrar por cualquier área
            area = self.request.GET.get('area')
            if area:
                queryset = queryset.filter(area__codigo=area)

        # ===== FILTROS AVANZADOS =====

        # Búsqueda general (código interno, código UTP, serie, nombre)
        search = self.request.GET.get('q')
        if search:
            queryset = queryset.filter(
                Q(codigo_interno__icontains=search) |
                Q(codigo_utp__icontains=search) |
                Q(serie__icontains=search) |
                Q(nombre__icontains=search) |
                Q(descripcion__icontains=search)
            )

        # Filtro por estado (puede ser múltiple)
        estados = self.request.GET.getlist('estado')
        if estados:
            queryset = queryset.filter(estado__in=estados)

        # Filtro por tipo de ítem
        tipo_item = self.request.GET.get('tipo_item')
        if tipo_item:
            queryset = queryset.filter(tipo_item_id=tipo_item)

        # Filtro por código UTP pendiente
        utp_pendiente = self.request.GET.get('utp_pendiente')
        if utp_pendiente == '1':
            queryset = queryset.filter(codigo_utp='PENDIENTE')
        elif utp_pendiente == '0':
            queryset = queryset.exclude(codigo_utp='PENDIENTE')

        # Filtro por colaborador asignado
        colaborador_asignado = self.request.GET.get('colaborador_asignado')
        if colaborador_asignado == 'sin_asignar':
            queryset = queryset.filter(colaborador_asignado__isnull=True)
        elif colaborador_asignado == 'asignado':
            queryset = queryset.filter(colaborador_asignado__isnull=False)
        elif colaborador_asignado:
            queryset = queryset.filter(colaborador_asignado_id=colaborador_asignado)

        # Filtro por ambiente
        ambiente = self.request.GET.get('ambiente')
        if ambiente == 'sin_ubicacion':
            queryset = queryset.filter(ambiente__isnull=True)
        elif ambiente:
            queryset = queryset.filter(ambiente_id=ambiente)

        # Filtro por campus
        campus = self.request.GET.get('campus')
        if campus:
            queryset = queryset.filter(ambiente__pabellon__sede__campus_id=campus)

        # Filtro por rango de precios
        precio_min = self.request.GET.get('precio_min')
        precio_max = self.request.GET.get('precio_max')
        if precio_min:
            queryset = queryset.filter(precio__gte=precio_min)
        if precio_max:
            queryset = queryset.filter(precio__lte=precio_max)

        # Filtro por rango de fechas de adquisición
        fecha_adq_desde = self.request.GET.get('fecha_adq_desde')
        fecha_adq_hasta = self.request.GET.get('fecha_adq_hasta')
        if fecha_adq_desde:
            queryset = queryset.filter(fecha_adquisicion__gte=fecha_adq_desde)
        if fecha_adq_hasta:
            queryset = queryset.filter(fecha_adquisicion__lte=fecha_adq_hasta)

        # Filtro por garantía
        garantia = self.request.GET.get('garantia')
        if garantia == 'vigente':
            queryset = queryset.filter(garantia_hasta__gte=timezone.now().date())
        elif garantia == 'vencida':
            queryset = queryset.filter(garantia_hasta__lt=timezone.now().date())
        elif garantia == 'sin_garantia':
            queryset = queryset.filter(garantia_hasta__isnull=True)
        elif garantia == 'proxima_vencer':  # 30 días
            fecha_limite = timezone.now().date() + timedelta(days=30)
            queryset = queryset.filter(
                garantia_hasta__lte=fecha_limite,
                garantia_hasta__gte=timezone.now().date()
            )

        # Filtro por leasing
        es_leasing = self.request.GET.get('es_leasing')
        if es_leasing == '1':
            queryset = queryset.filter(es_leasing=True)
        elif es_leasing == '0':
            queryset = queryset.filter(es_leasing=False)

        # Filtro por lote
        lote = self.request.GET.get('lote')
        if lote == 'sin_lote':
            queryset = queryset.filter(lote__isnull=True)
        elif lote:
            queryset = queryset.filter(lote_id=lote)

        # Ordenamiento
        order_by = self.request.GET.get('order_by', '-creado_en')
        if order_by:
            queryset = queryset.order_by(order_by)

        return queryset
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)

        # Opciones para filtros
        context['areas'] = Area.objects.filter(activo=True)
        context['campus_list'] = Campus.objects.filter(activo=True)
        context['ambientes'] = Ambiente.objects.filter(activo=True).select_related('pabellon__sede__campus')
        context['estados'] = Item.ESTADOS
        context['tipos_item'] = TipoItem.objects.filter(activo=True).select_related('area')
        context['colaboradores'] = Colaborador.objects.filter(activo=True).order_by('nombre_completo')
        context['lotes'] = Lote.objects.filter(activo=True).order_by('-creado_en')[:50]

        # Filtros activos (para mostrar chips)
        context['filtros_activos'] = {
            'q': self.request.GET.get('q', ''),
            'area': self.request.GET.get('area', ''),
            'estado': self.request.GET.getlist('estado'),
            'tipo_item': self.request.GET.get('tipo_item', ''),
            'utp_pendiente': self.request.GET.get('utp_pendiente', ''),
            'colaborador_asignado': self.request.GET.get('colaborador_asignado', ''),
            'ambiente': self.request.GET.get('ambiente', ''),
            'campus': self.request.GET.get('campus', ''),
            'precio_min': self.request.GET.get('precio_min', ''),
            'precio_max': self.request.GET.get('precio_max', ''),
            'fecha_adq_desde': self.request.GET.get('fecha_adq_desde', ''),
            'fecha_adq_hasta': self.request.GET.get('fecha_adq_hasta', ''),
            'garantia': self.request.GET.get('garantia', ''),
            'es_leasing': self.request.GET.get('es_leasing', ''),
            'lote': self.request.GET.get('lote', ''),
            'order_by': self.request.GET.get('order_by', '-creado_en'),
        }

        # Contar filtros activos (para badge)
        filtros_count = sum(1 for k, v in context['filtros_activos'].items()
                           if v and k != 'order_by' and v != [])
        context['filtros_count'] = filtros_count

        # Contar solo filtros avanzados (excluyendo búsqueda simple 'q')
        filtros_avanzados = sum(1 for k, v in context['filtros_activos'].items()
                               if v and k not in ['order_by', 'q'] and v != [])
        context['filtros_avanzados_activos'] = filtros_avanzados > 0

        return context


class ItemDetailView(PerfilRequeridoMixin, DetailView):
    """Detalle de un ítem."""
    model = Item
    template_name = 'productos/item_detail.html'
    context_object_name = 'item'
    slug_url_kwarg = 'codigo'

    def get_queryset(self):
        """Restringir acceso por área si no es admin."""
        queryset = super().get_queryset()
        perfil = getattr(self.request.user, 'perfil', None)

        if perfil and perfil.rol != 'admin' and perfil.area:
            queryset = queryset.filter(area=perfil.area)

        return queryset

    def get_object(self, queryset=None):
        """Buscar item por codigo_interno o codigo_utp."""
        if queryset is None:
            queryset = self.get_queryset()

        codigo = self.kwargs.get(self.slug_url_kwarg)

        # Primero buscar por codigo_interno
        obj = queryset.filter(codigo_interno=codigo).first()

        # Si no existe, buscar por codigo_utp
        if not obj:
            obj = queryset.filter(codigo_utp=codigo).first()

        if not obj:
            raise Http404("Item no encontrado")

        return obj
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        item = self.object
        
        # Especificaciones de sistemas si aplica
        if item.area.codigo == 'sistemas':
            context['especificaciones'] = getattr(item, 'especificaciones_sistemas', None)
        
        # Historial de movimientos
        context['movimientos'] = item.movimientos.select_related(
            'solicitado_por', 'aprobado_por'
        )[:10]
        
        # Historial de cambios
        context['historial'] = item.historial_cambios.select_related('usuario')[:10]
        
        return context


class ItemCreateView(AlmacenRequeridoMixin, CreateView):
    """Crear un nuevo ítem. Solo admin, gerente y almacén."""
    model = Item
    form_class = ItemSistemasForm  # Usar formulario con especificaciones
    template_name = 'productos/item_form.html'

    def get_form_kwargs(self):
        kwargs = super().get_form_kwargs()
        kwargs['user'] = self.request.user
        return kwargs

    def get_initial(self):
        """Pre-seleccionar el área del usuario si no es admin."""
        initial = super().get_initial()
        perfil = getattr(self.request.user, 'perfil', None)

        if perfil and perfil.rol != 'admin' and perfil.area:
            initial['area'] = perfil.area

        return initial

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['titulo'] = 'Nuevo Ítem'
        context['es_sistemas'] = self.request.GET.get('area') == 'sistemas'
        return context

    def form_valid(self, form):
        item = form.save(commit=False)
        item.creado_por = self.request.user
        item.modificado_por = self.request.user

        # Si no es admin, forzar el área del usuario
        perfil = getattr(self.request.user, 'perfil', None)
        if perfil and perfil.rol != 'admin' and perfil.area:
            item.area = perfil.area

        # Auto-generar código UTP
        if not item.codigo_utp:
            item.codigo_utp = Item.generar_codigo_utp(item.area.codigo)

        item.save()

        # Guardar especificaciones si es área de Sistemas
        if item.area.codigo == 'sistemas':
            form.save()  # Esto guarda las especificaciones

        messages.success(self.request, f'Ítem {item.codigo_interno} creado correctamente.')
        return redirect('productos:item-detail', codigo=item.codigo_interno)


class ItemUpdateView(AlmacenRequeridoMixin, UpdateView):
    """Editar un ítem existente. Solo admin, gerente y almacén."""
    model = Item
    form_class = ItemSistemasForm  # Usar formulario con especificaciones
    template_name = 'productos/item_form.html'
    slug_field = 'codigo_interno'
    slug_url_kwarg = 'codigo'

    def get_queryset(self):
        """Restringir acceso por área si no es admin."""
        queryset = super().get_queryset()
        perfil = getattr(self.request.user, 'perfil', None)

        if perfil and perfil.rol != 'admin' and perfil.area:
            queryset = queryset.filter(area=perfil.area)

        return queryset

    def get_form_kwargs(self):
        kwargs = super().get_form_kwargs()
        kwargs['user'] = self.request.user
        return kwargs

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['titulo'] = f'Editar {self.object.codigo_utp}'
        context['es_sistemas'] = self.object.area.codigo == 'sistemas'
        return context

    def form_valid(self, form):
        item = form.save(commit=False)
        item.modificado_por = self.request.user
        item.save()

        # Guardar especificaciones si es área de Sistemas
        if item.area.codigo == 'sistemas':
            form.save()  # Esto guarda las especificaciones

        messages.success(self.request, f'Ítem {item.codigo_interno} actualizado correctamente.')
        return redirect('productos:item-detail', codigo=item.codigo_interno)


class ItemDeleteView(AlmacenRequeridoMixin, DeleteView):
    """Eliminar un ítem. Solo admin, gerente y almacén."""
    model = Item
    template_name = 'productos/item_confirm_delete.html'
    success_url = reverse_lazy('productos:item-list')
    slug_field = 'codigo_interno'
    slug_url_kwarg = 'codigo'

    def delete(self, request, *args, **kwargs):
        item = self.get_object()
        messages.success(request, f'Ítem {item.codigo_interno} eliminado.')
        return super().delete(request, *args, **kwargs)


# ============================================================================
# VISTAS DE MOVIMIENTOS
# ============================================================================

class MovimientoListView(PerfilRequeridoMixin, ListView):
    """Lista de movimientos."""
    model = Movimiento
    template_name = 'productos/movimiento_list.html'
    context_object_name = 'movimientos'
    paginate_by = 20
    
    def get_queryset(self):
        queryset = Movimiento.objects.select_related(
            'item', 'item__area', 'item__tipo_item',
            'solicitado_por', 'aprobado_por', 'ejecutado_por',
            'ambiente_origen__pabellon__sede',
            'ambiente_destino__pabellon__sede',
            'colaborador_nuevo', 'colaborador_anterior'
        ).prefetch_related(
            'items_movimiento__item'  # Para movimientos con múltiples ítems
        )

        perfil = getattr(self.request.user, 'perfil', None)

        # Filtrar por área si no es admin
        if perfil and perfil.rol != 'admin' and perfil.area:
            queryset = queryset.filter(item__area=perfil.area)

        # Filtro por estado
        estado = self.request.GET.get('estado')
        if estado:
            queryset = queryset.filter(estado=estado)

        return queryset
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['estados'] = Movimiento.ESTADOS_MOVIMIENTO
        return context


class MovimientoPendientesView(SupervisorRequeridoMixin, ListView):
    """Movimientos pendientes de aprobar."""
    model = Movimiento
    template_name = 'productos/movimiento_pendientes.html'
    context_object_name = 'movimientos'
    
    def get_queryset(self):
        queryset = Movimiento.objects.filter(estado='pendiente').select_related(
            'item', 'item__area', 'item__tipo_item',
            'solicitado_por',
            'ambiente_origen__pabellon__sede',
            'ambiente_destino__pabellon__sede',
            'colaborador_nuevo'
        ).prefetch_related(
            'items_movimiento__item'
        )

        perfil = getattr(self.request.user, 'perfil', None)

        # Supervisor solo ve de su área
        if perfil and perfil.rol == 'supervisor' and perfil.area:
            queryset = queryset.filter(item__area=perfil.area)

        return queryset


class MovimientoCreateView(PerfilRequeridoMixin, CreateView):
    """Crear un nuevo movimiento."""
    model = Movimiento
    form_class = MovimientoForm
    template_name = 'productos/movimiento_form.html'
    
    def get_form_kwargs(self):
        kwargs = super().get_form_kwargs()
        kwargs['user'] = self.request.user
        
        # Pre-cargar ítem si viene en la URL
        item_pk = self.request.GET.get('item')
        if item_pk:
            kwargs['item'] = get_object_or_404(Item, pk=item_pk)
        
        return kwargs
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['titulo'] = 'Nueva Solicitud de Movimiento'
        
        # Datos para los filtros de búsqueda de ítems
        context['campus_list'] = Campus.objects.filter(activo=True)
        context['areas_list'] = Area.objects.filter(activo=True)
        
        # Ítem preseleccionado si viene en URL
        item_pk = self.request.GET.get('item')
        if item_pk:
            context['item'] = get_object_or_404(Item, pk=item_pk)
        
        return context
    
    def form_valid(self, form):
        movimiento = form.save(commit=False)
        movimiento.solicitado_por = self.request.user
        movimiento.estado = 'pendiente'

        # Obtener los ítems seleccionados
        items = form.cleaned_data.get('items')

        if items:
            # Usar el primer ítem para el campo item del modelo (compatibilidad)
            primer_item = items[0]
            movimiento.item = primer_item
            movimiento.ambiente_origen = primer_item.ambiente
            movimiento.colaborador_anterior = primer_item.colaborador_asignado

            movimiento.save()

            # Crear registros MovimientoItem para todos los ítems
            for item in items:
                MovimientoItem.objects.create(
                    movimiento=movimiento,
                    item=item
                )

            cantidad = len(items)
            if cantidad == 1:
                msg = f'Solicitud de {movimiento.get_tipo_display()} creada para 1 ítem. Pendiente de aprobación.'
            else:
                msg = f'Solicitud de {movimiento.get_tipo_display()} creada para {cantidad} ítems. Pendiente de aprobación.'

            messages.success(self.request, msg)
        else:
            messages.error(self.request, 'Debe seleccionar al menos un ítem.')
            return self.form_invalid(form)

        return redirect('productos:movimiento-list')


class MovimientoDetailView(PerfilRequeridoMixin, DetailView):
    """Detalle de un movimiento."""
    model = Movimiento
    template_name = 'productos/movimiento_detail.html'
    context_object_name = 'movimiento'
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        # Pasar el perfil del usuario al contexto
        context['perfil'] = getattr(self.request.user, 'perfil', None)
        # Detectar si viene de pendientes para el breadcrumb
        context['from_pendientes'] = 'pendientes' in self.request.META.get('HTTP_REFERER', '')
        return context


class MovimientoAprobarView(SupervisorRequeridoMixin, View):
    """Aprobar un movimiento - Solo cambia estado a 'aprobado'."""

    def post(self, request, pk):
        movimiento = get_object_or_404(Movimiento, pk=pk)

        # Verificar que puede aprobar
        perfil = getattr(request.user, 'perfil', None)
        if perfil and perfil.rol == 'supervisor':
            if perfil.area != movimiento.item.area:
                messages.error(request, 'No puedes aprobar movimientos de otra área.')
                return redirect('productos:movimiento-detail', pk=pk)

        if movimiento.estado != 'pendiente':
            messages.error(request, 'Este movimiento no puede ser aprobado.')
            return redirect('productos:movimiento-detail', pk=pk)

        # Solo aprobar - el auxiliar debe marcar en ejecución después
        movimiento.aprobar(request.user)

        messages.success(
            request,
            'Movimiento aprobado. El auxiliar debe marcar "En Ejecución" cuando retire el equipo.'
        )
        return redirect('productos:movimiento-detail', pk=pk)


class MovimientoRechazarView(SupervisorRequeridoMixin, View):
    """Rechazar un movimiento."""
    
    def post(self, request, pk):
        movimiento = get_object_or_404(Movimiento, pk=pk)
        motivo = request.POST.get('motivo_rechazo', 'Sin motivo especificado')
        
        # Verificar que puede rechazar
        perfil = getattr(request.user, 'perfil', None)
        if perfil and perfil.rol == 'supervisor':
            if perfil.area != movimiento.item.area:
                messages.error(request, 'No puedes rechazar movimientos de otra área.')
                return redirect('productos:movimiento-detail', pk=pk)
        
        if movimiento.estado != 'pendiente':
            messages.error(request, 'Este movimiento no puede ser rechazado.')
            return redirect('productos:movimiento-detail', pk=pk)

        movimiento.rechazar(request.user, motivo)
        
        messages.success(request, 'Movimiento rechazado.')
        return redirect('productos:movimiento-pendientes')


class MovimientoEnEjecucionView(PerfilRequeridoMixin, View):
    """
    Marcar movimiento como 'En Ejecución'.
    El auxiliar de origen usa esto cuando retira físicamente el equipo.
    """

    def post(self, request, pk):
        movimiento = get_object_or_404(Movimiento, pk=pk)

        # Verificar permisos - debe ser auxiliar del campus de origen
        perfil = getattr(request.user, 'perfil', None)
        if perfil and perfil.rol == 'auxiliar':
            if movimiento.ambiente_origen:
                campus_origen = movimiento.ambiente_origen.campus
                if not perfil.puede_ver_campus(campus_origen):
                    messages.error(request, 'No tienes permiso para este movimiento.')
                    return redirect('productos:movimiento-detail', pk=pk)

        if movimiento.estado != 'aprobado':
            messages.error(request, 'Este movimiento no está en estado "Aprobado".')
            return redirect('productos:movimiento-detail', pk=pk)

        if movimiento.marcar_en_ejecucion(request.user):
            messages.success(request, 'Movimiento marcado como "En Ejecución". Equipo retirado.')

            # Si es traslado entre campus, indicar siguiente paso
            if movimiento.es_entre_campus:
                messages.info(
                    request,
                    'Este es un traslado entre campus. Marca "En Tránsito" cuando el equipo salga.'
                )
            else:
                messages.info(
                    request,
                    'Puedes marcar "Ejecutado" cuando el equipo esté instalado/entregado.'
                )
        else:
            messages.error(request, 'No se pudo marcar el movimiento como "En Ejecución".')

        return redirect('productos:movimiento-detail', pk=pk)


class MovimientoEnTransitoView(PerfilRequeridoMixin, View):
    """
    Marcar movimiento como 'En Tránsito'.
    Solo aplica para traslados entre campus diferentes.
    El auxiliar de origen usa esto cuando el equipo sale físicamente.
    """

    def post(self, request, pk):
        movimiento = get_object_or_404(Movimiento, pk=pk)

        # Verificar permisos - debe ser auxiliar del campus de origen
        perfil = getattr(request.user, 'perfil', None)
        if perfil and perfil.rol == 'auxiliar':
            if movimiento.ambiente_origen:
                campus_origen = movimiento.ambiente_origen.campus
                if not perfil.puede_ver_campus(campus_origen):
                    messages.error(request, 'No tienes permiso para este movimiento.')
                    return redirect('productos:movimiento-detail', pk=pk)

        if movimiento.estado != 'en_ejecucion':
            messages.error(request, 'Este movimiento debe estar "En Ejecución" primero.')
            return redirect('productos:movimiento-detail', pk=pk)

        if not movimiento.es_entre_campus:
            messages.error(request, 'Solo los traslados entre campus requieren estado "En Tránsito".')
            return redirect('productos:movimiento-detail', pk=pk)

        if movimiento.marcar_en_transito(request.user):
            campus_destino = movimiento.campus_destino
            messages.success(
                request,
                f'Equipo marcado "En Tránsito" hacia {campus_destino.nombre if campus_destino else "destino"}.'
            )
            messages.info(
                request,
                'El auxiliar del campus destino debe confirmar la recepción.'
            )
        else:
            messages.error(request, 'No se pudo marcar el movimiento como "En Tránsito".')

        return redirect('productos:movimiento-detail', pk=pk)


class MovimientoEjecutarView(PerfilRequeridoMixin, View):
    """
    Marcar movimiento como 'Ejecutado'.
    Para asignaciones y préstamos, redirige a crear Acta de Entrega.
    El auxiliar (de destino si es traslado entre campus) confirma la recepción/instalación.
    """

    # Tipos de movimiento que requieren Acta de Entrega obligatoria
    TIPOS_REQUIEREN_ACTA = ['asignacion', 'prestamo']

    def post(self, request, pk):
        movimiento = get_object_or_404(Movimiento, pk=pk)

        # Verificar permisos
        perfil = getattr(request.user, 'perfil', None)
        if perfil and perfil.rol == 'auxiliar':
            # Si es traslado entre campus, debe ser auxiliar del destino
            if movimiento.es_entre_campus and movimiento.ambiente_destino:
                campus_destino = movimiento.ambiente_destino.campus
                if not perfil.puede_ver_campus(campus_destino):
                    messages.error(
                        request,
                        'Solo el auxiliar del campus destino puede confirmar la recepción.'
                    )
                    return redirect('productos:movimiento-detail', pk=pk)

        # Validar estado previo
        estados_validos = ['aprobado', 'en_ejecucion', 'en_transito']
        if movimiento.estado not in estados_validos:
            messages.error(request, f'Este movimiento no puede ejecutarse (estado: {movimiento.get_estado_display()}).')
            return redirect('productos:movimiento-detail', pk=pk)

        # Si es traslado entre campus y no está en tránsito, advertir
        if movimiento.es_entre_campus and movimiento.estado != 'en_transito':
            messages.warning(
                request,
                'Este es un traslado entre campus. Se recomienda marcar "En Tránsito" antes de ejecutar.'
            )

        # Para asignaciones y préstamos: redirigir a crear Acta de Entrega
        if movimiento.tipo in self.TIPOS_REQUIEREN_ACTA:
            # Verificar si ya tiene acta
            if hasattr(movimiento, 'acta_entrega') and movimiento.acta_entrega:
                messages.info(request, 'Este movimiento ya tiene un acta de entrega asociada.')
                return redirect('productos:acta-detail', pk=movimiento.acta_entrega.pk)

            # Redirigir a crear acta con el movimiento precargado
            messages.info(
                request,
                'Para completar esta asignación, debe generar el Acta de Entrega con las firmas correspondientes.'
            )
            from django.urls import reverse
            url = reverse('productos:acta-create')
            return redirect(f"{url}?movimiento={pk}")

        # Para otros tipos de movimiento, ejecutar directamente
        if movimiento.ejecutar(request.user):
            messages.success(request, 'Movimiento ejecutado correctamente. Inventario actualizado.')

            # Si requiere formato de traslado, ofrecer descarga
            if movimiento.requiere_formato_traslado:
                from django.urls import reverse
                url = reverse('productos:movimiento-detail', kwargs={'pk': pk})
                return redirect(f"{url}?descargar_formato=1")
        else:
            messages.error(request, 'No se pudo ejecutar el movimiento.')

        return redirect('productos:movimiento-detail', pk=pk)


# ============================================================================
# VISTAS DE NOTIFICACIONES
# ============================================================================

class NotificacionListView(PerfilRequeridoMixin, ListView):
    """Lista de notificaciones del usuario."""
    model = Notificacion
    template_name = 'productos/notificacion_list.html'
    context_object_name = 'notificaciones'
    paginate_by = 20
    
    def get_queryset(self):
        return Notificacion.objects.filter(usuario=self.request.user)


class NotificacionMarcarLeidaView(PerfilRequeridoMixin, View):
    """Marcar notificación como leída."""
    
    def post(self, request, pk):
        notificacion = get_object_or_404(Notificacion, pk=pk, usuario=request.user)
        notificacion.marcar_leida()
        
        # Redirigir a la URL de la notificación si existe
        if notificacion.url:
            return redirect(notificacion.url)
        return redirect('productos:notificacion-list')


class NotificacionMarcarTodasLeidasView(PerfilRequeridoMixin, View):
    """Marcar todas las notificaciones como leídas."""
    
    def post(self, request):
        Notificacion.objects.filter(usuario=request.user, leida=False).update(leida=True)
        messages.success(request, 'Todas las notificaciones marcadas como leídas.')
        return redirect('productos:notificacion-list')


# ============================================================================
# API ENDPOINTS (JSON)
# ============================================================================

class TiposItemPorAreaView(PerfilRequeridoMixin, View):
    """Obtiene los tipos de ítem para un área (para formularios dinámicos)."""
    
    def get(self, request):
        area_id = request.GET.get('area')
        tipos = []
        
        if area_id:
            tipos = list(TipoItem.objects.filter(
                area_id=area_id, activo=True
            ).values('id', 'nombre'))
        
        return JsonResponse({'tipos': tipos})


class SupervisoresPorAreaView(PerfilRequeridoMixin, View):
    """Obtiene los supervisores de un área (para seleccionar autorizador)."""
    
    def get(self, request):
        area_id = request.GET.get('area')
        supervisores = []
        
        if area_id:
            from django.contrib.auth.models import User
            supervisores = list(
                User.objects.filter(
                    perfil__area_id=area_id,
                    perfil__rol='supervisor',
                    perfil__activo=True,
                    is_active=True
                ).values('id', 'first_name', 'last_name', 'username')
            )
        
        # Agregar admins siempre
        admins = list(
            User.objects.filter(
                perfil__rol='admin',
                perfil__activo=True,
                is_active=True
            ).values('id', 'first_name', 'last_name', 'username')
        )
        
        return JsonResponse({'supervisores': supervisores + admins})


# ============================================================================
# VISTAS PARA CREAR TIPO ITEM (accesible a operadores)
# ============================================================================

class TipoItemCreateView(PerfilRequeridoMixin, CreateView):
    """Vista para crear tipos de ítem (accesible a todos los roles excepto externos)."""
    model = TipoItem
    form_class = TipoItemForm
    template_name = 'productos/tipoitem_form.html'
    
    def get_form_kwargs(self):
        kwargs = super().get_form_kwargs()
        kwargs['user'] = self.request.user
        return kwargs
    
    def form_valid(self, form):
        # Si el usuario tiene área asignada y no es admin, usar esa área
        perfil = getattr(self.request.user, 'perfil', None)
        if perfil and perfil.area and perfil.rol != 'admin':
            form.instance.area = perfil.area
        
        messages.success(
            self.request, 
            f'Tipo de ítem "{form.instance.nombre}" creado exitosamente.'
        )
        return super().form_valid(form)
    
    def get_success_url(self):
        # Redirigir a donde vino o al listado de ítems
        next_url = self.request.GET.get('next')
        if next_url:
            return next_url
        return reverse_lazy('productos:item-create')
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['titulo'] = 'Crear Tipo de Ítem'
        
        # Verificar si hay advertencia de tipos similares
        if hasattr(self.get_form(), 'advertencia_similares'):
            context['advertencia_similares'] = self.get_form().advertencia_similares
        
        return context


class TipoItemListView(PerfilRequeridoMixin, ListView):
    """Lista de tipos de ítem."""
    model = TipoItem
    template_name = 'productos/tipoitem_list.html'
    context_object_name = 'tipos'
    paginate_by = 20
    
    def get_queryset(self):
        queryset = TipoItem.objects.select_related('area').filter(activo=True)
        
        # Filtrar por área si el usuario no es admin
        perfil = getattr(self.request.user, 'perfil', None)
        if perfil and perfil.rol != 'admin' and perfil.area:
            queryset = queryset.filter(area=perfil.area)
        
        # Filtro por área desde request
        area_id = self.request.GET.get('area')
        if area_id:
            queryset = queryset.filter(area_id=area_id)
        
        # Búsqueda
        search = self.request.GET.get('q')
        if search:
            queryset = queryset.filter(nombre__icontains=search)
        
        return queryset.order_by('area', 'nombre')
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['areas'] = Area.objects.filter(activo=True)
        return context


# ============================================================================
# VISTAS PARA UBICACIONES
# ============================================================================

class UbicacionListView(PerfilRequeridoMixin, CampusFilterMixin, ListView):
    """Lista de ambientes (ubicaciones) según permisos de campus."""
    model = Ambiente
    template_name = 'productos/ubicacion_list.html'
    context_object_name = 'ambientes'
    paginate_by = 20

    def get_queryset(self):
        queryset = Ambiente.objects.filter(activo=True).select_related(
            'pabellon', 'pabellon__sede', 'pabellon__sede__campus'
        )

        # Filtrar por campus permitidos del usuario
        queryset = self.filtrar_por_campus(queryset, 'pabellon__sede__campus')

        # Filtros adicionales
        campus = self.request.GET.get('campus')
        sede = self.request.GET.get('sede')
        pabellon = self.request.GET.get('pabellon')
        tipo = self.request.GET.get('tipo')
        search = self.request.GET.get('q')

        if campus:
            queryset = queryset.filter(pabellon__sede__campus_id=campus)
        if sede:
            queryset = queryset.filter(pabellon__sede_id=sede)
        if pabellon:
            queryset = queryset.filter(pabellon_id=pabellon)
        if tipo:
            queryset = queryset.filter(tipo=tipo)
        if search:
            queryset = queryset.filter(
                Q(nombre__icontains=search) |
                Q(codigo__icontains=search)
            )

        return queryset.order_by('pabellon__sede__campus__nombre', 'pabellon__sede__nombre', 'pabellon__nombre', 'piso')

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        # Solo mostrar campus/sedes/pabellones que el usuario puede ver
        campus_permitidos = self.get_campus_permitidos()
        campus_ids = list(campus_permitidos.values_list('id', flat=True))

        context['campus_list'] = campus_permitidos
        context['sedes_list'] = Sede.objects.filter(activo=True, campus_id__in=campus_ids).select_related('campus')
        context['pabellones_list'] = Pabellon.objects.filter(activo=True, sede__campus_id__in=campus_ids).select_related('sede')
        context['tipos_ambiente'] = Ambiente.TIPOS_AMBIENTE
        return context


class UbicacionDetailView(PerfilRequeridoMixin, DetailView):
    """Detalle de ambiente (ubicación) con ítems."""
    model = Ambiente
    template_name = 'productos/ubicacion_detail.html'
    context_object_name = 'ambiente'
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        # Ítems en este ambiente
        context['items'] = Item.objects.filter(ambiente=self.object).select_related('area', 'tipo_item')
        return context


class UbicacionCreateView(SupervisorRequeridoMixin, CreateView):
    """Crear ambiente (solo supervisores y admins)."""
    model = Ambiente
    form_class = AmbienteForm
    template_name = 'productos/ubicacion_form.html'
    success_url = reverse_lazy('productos:ubicacion-list')
    
    def form_valid(self, form):
        messages.success(self.request, 'Ambiente creado exitosamente.')
        return super().form_valid(form)
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['titulo'] = 'Crear Ambiente'
        return context


class UbicacionUpdateView(SupervisorRequeridoMixin, UpdateView):
    """Editar ambiente (solo supervisores y admins)."""
    model = Ambiente
    form_class = AmbienteForm
    template_name = 'productos/ubicacion_form.html'
    
    def get_success_url(self):
        return reverse_lazy('productos:ubicacion-detail', kwargs={'pk': self.object.pk})
    
    def form_valid(self, form):
        messages.success(self.request, 'Ambiente actualizado exitosamente.')
        return super().form_valid(form)
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['titulo'] = 'Editar Ambiente'
        context['editando'] = True
        return context


# ============================================================================
# APIs AJAX PARA DROPDOWNS EN CASCADA
# ============================================================================

class SedesPorCampusView(LoginRequiredMixin, View):
    """API para obtener sedes de un campus."""

    def get(self, request):
        campus_id = request.GET.get('campus_id')
        if campus_id:
            # Validar que el usuario tiene acceso al campus
            perfil = getattr(request.user, 'perfil', None)
            if perfil:
                campus_permitidos = perfil.get_campus_permitidos()
                if not campus_permitidos.filter(pk=campus_id).exists():
                    return JsonResponse([], safe=False)
            sedes = Sede.objects.filter(campus_id=campus_id, activo=True).values('id', 'nombre')
            return JsonResponse(list(sedes), safe=False)
        return JsonResponse([], safe=False)


class PabellonesPorSedeView(LoginRequiredMixin, View):
    """API para obtener pabellones de una sede."""

    def get(self, request):
        sede_id = request.GET.get('sede_id')
        if sede_id:
            # Validar que el usuario tiene acceso al campus de la sede
            perfil = getattr(request.user, 'perfil', None)
            if perfil:
                try:
                    sede = Sede.objects.select_related('campus').get(pk=sede_id)
                    campus_permitidos = perfil.get_campus_permitidos()
                    if not campus_permitidos.filter(pk=sede.campus_id).exists():
                        return JsonResponse([], safe=False)
                except Sede.DoesNotExist:
                    return JsonResponse([], safe=False)
            pabellones = Pabellon.objects.filter(sede_id=sede_id, activo=True).values('id', 'nombre')
            return JsonResponse(list(pabellones), safe=False)
        return JsonResponse([], safe=False)


class AmbientesPorPabellonView(LoginRequiredMixin, View):
    """API para obtener ambientes de un pabellón."""

    def get(self, request):
        pabellon_id = request.GET.get('pabellon_id')
        if pabellon_id:
            # Validar que el usuario tiene acceso al campus del pabellón
            perfil = getattr(request.user, 'perfil', None)
            if perfil:
                try:
                    pabellon = Pabellon.objects.select_related('sede__campus').get(pk=pabellon_id)
                    campus_permitidos = perfil.get_campus_permitidos()
                    if not campus_permitidos.filter(pk=pabellon.sede.campus_id).exists():
                        return JsonResponse([], safe=False)
                except Pabellon.DoesNotExist:
                    return JsonResponse([], safe=False)
            ambientes = Ambiente.objects.filter(pabellon_id=pabellon_id, activo=True).values('id', 'nombre', 'piso', 'tipo')
            return JsonResponse(list(ambientes), safe=False)
        return JsonResponse([], safe=False)


class BuscarItemsView(RateLimitMixin, LoginRequiredMixin, View):
    """API para buscar ítems con autocompletado."""
    ratelimit_key = 'search'

    def get(self, request):
        query = request.GET.get('q', '').strip()
        area_id = request.GET.get('area')
        tipo_id = request.GET.get('tipo')
        campus_id = request.GET.get('campus')
        sede_id = request.GET.get('sede')
        pabellon_id = request.GET.get('pabellon')
        ambiente_id = request.GET.get('ambiente')
        tipo_movimiento = request.GET.get('tipo_movimiento', '')

        items = Item.objects.select_related(
            'area', 'tipo_item', 'ambiente', 'ambiente__pabellon',
            'ambiente__pabellon__sede', 'ambiente__pabellon__sede__campus',
            'usuario_asignado', 'colaborador_asignado'
        )

        # Filtrar por área del usuario si no es admin
        perfil = getattr(request.user, 'perfil', None)
        if perfil and perfil.rol != 'admin' and perfil.area:
            items = items.filter(area=perfil.area)

        # Filtrar según tipo de movimiento
        if tipo_movimiento == 'asignacion':
            # Solo ítems disponibles para asignar: backup o custodia, sin colaborador
            items = items.filter(
                estado__in=['backup', 'custodia'],
                colaborador_asignado__isnull=True
            )
        elif tipo_movimiento == 'traslado':
            # Cualquier ítem excepto baja o en tránsito
            items = items.exclude(estado__in=['baja', 'transito'])
        elif tipo_movimiento == 'mantenimiento':
            # Ítems que no estén ya en mantenimiento o garantía
            items = items.exclude(estado__in=['mantenimiento', 'garantia', 'baja', 'transito'])
        elif tipo_movimiento == 'garantia':
            # Ítems instalados o en custodia que necesiten garantía
            items = items.filter(estado__in=['instalado', 'custodia', 'backup'])
        elif tipo_movimiento == 'reemplazo':
            # Ítems disponibles para usar como reemplazo
            items = items.filter(
                estado__in=['backup', 'custodia'],
                colaborador_asignado__isnull=True
            )
        elif tipo_movimiento == 'prestamo':
            # Ítems disponibles para préstamo
            items = items.filter(
                estado__in=['backup', 'custodia'],
                colaborador_asignado__isnull=True
            )
        elif tipo_movimiento == 'leasing':
            # Ítems de leasing que necesitan devolverse
            items = items.filter(es_leasing=True).exclude(estado='baja')
        
        # Búsqueda por texto
        if query:
            items = items.filter(
                Q(codigo_interno__icontains=query) |
                Q(codigo_utp__icontains=query) |
                Q(serie__icontains=query) |
                Q(nombre__icontains=query)
            )
        
        # Filtros adicionales
        if area_id:
            items = items.filter(area_id=area_id)
        if tipo_id:
            items = items.filter(tipo_item_id=tipo_id)
        if campus_id:
            items = items.filter(ambiente__pabellon__sede__campus_id=campus_id)
        if sede_id:
            items = items.filter(ambiente__pabellon__sede_id=sede_id)
        if pabellon_id:
            items = items.filter(ambiente__pabellon_id=pabellon_id)
        if ambiente_id:
            items = items.filter(ambiente_id=ambiente_id)
        
        # Limitar resultados
        items = items[:20]
        
        # Formatear respuesta
        resultados = []
        for item in items:
            ubicacion = ""
            if item.ambiente:
                amb = item.ambiente
                if amb.pabellon and amb.pabellon.sede and amb.pabellon.sede.campus:
                    ubicacion = f"{amb.pabellon.sede.campus.codigo} > {amb.pabellon.sede.nombre} > Pab. {amb.pabellon.letra} - {amb.nombre}"
                else:
                    ubicacion = amb.nombre
            
            # Obtener marca y modelo si tiene especificaciones
            marca = ''
            modelo = ''
            if hasattr(item, 'especificaciones_sistemas'):
                try:
                    specs = item.especificaciones_sistemas
                    if specs and specs.marca_equipo:
                        marca = specs.marca_equipo.nombre
                    if specs and specs.modelo_equipo:
                        modelo = specs.modelo_equipo.nombre
                except AttributeError:
                    pass

            resultados.append({
                'id': item.id,
                'codigo_interno': item.codigo_interno,
                'codigo_utp': item.codigo_utp,
                'serie': item.serie,
                'nombre': item.nombre,
                'area': item.area.nombre if item.area else '',
                'tipo': item.tipo_item.nombre if item.tipo_item else '',
                'estado': item.estado,
                'estado_display': item.get_estado_display(),
                'ubicacion': ubicacion,
                'ambiente_id': item.ambiente_id,
                'usuario_asignado': item.usuario_asignado.get_full_name() if item.usuario_asignado else None,
                'marca': marca,
                'modelo': modelo,
                'texto': f"{item.codigo_utp or item.codigo_interno} - {item.nombre}"
            })
        
        return JsonResponse({'items': resultados})


class BuscarItemsActaView(RateLimitMixin, LoginRequiredMixin, View):
    """API para buscar ítems disponibles para actas de entrega/devolución."""
    ratelimit_key = 'search'

    def get(self, request):
        query = request.GET.get('q', '').strip()
        tipo_acta = request.GET.get('tipo_acta', 'entrega')
        colaborador_id = request.GET.get('colaborador_id')
        tipo_item_id = request.GET.get('tipo_item')

        if not query or len(query) < 2:
            return JsonResponse({'items': [], 'total': 0})

        if tipo_acta == 'entrega':
            items = Item.objects.filter(
                colaborador_asignado__isnull=True,
                estado__in=['nuevo', 'instalado']
            )
        else:
            # Devolución: items asignados al colaborador
            if not colaborador_id:
                return JsonResponse({'items': [], 'total': 0})
            items = Item.objects.filter(colaborador_asignado_id=colaborador_id)

        # Filtrar por área del usuario si no es admin
        perfil = getattr(request.user, 'perfil', None)
        if perfil and perfil.rol != 'admin' and perfil.area:
            items = items.filter(area=perfil.area)

        # Filtrar por tipo de item
        if tipo_item_id:
            items = items.filter(tipo_item_id=tipo_item_id)

        # Búsqueda por texto
        items = items.filter(
            Q(codigo_interno__icontains=query) |
            Q(codigo_utp__icontains=query) |
            Q(serie__icontains=query) |
            Q(nombre__icontains=query)
        ).select_related('tipo_item', 'area')

        total = items.count()
        items = items[:50]

        resultados = []
        for item in items:
            resultados.append({
                'id': item.id,
                'codigo': item.codigo_utp if not item.codigo_utp_pendiente else item.codigo_interno,
                'nombre': item.nombre,
                'serie': item.serie or '',
                'tipo': item.tipo_item.nombre if item.tipo_item else '',
                'estado': item.get_estado_display(),
                'estado_key': item.estado,
            })

        return JsonResponse({'items': resultados, 'total': total})


class ObtenerItemDetalleView(LoginRequiredMixin, View):
    """API para obtener detalles de un ítem específico."""
    
    def get(self, request):
        item_id = request.GET.get('id')
        if not item_id:
            return JsonResponse({'error': 'ID requerido'}, status=400)
        
        try:
            item = Item.objects.select_related(
                'area', 'tipo_item', 'ambiente', 'ambiente__pabellon',
                'ambiente__pabellon__sede', 'ambiente__pabellon__sede__campus',
                'usuario_asignado'
            ).get(pk=item_id)
        except Item.DoesNotExist:
            return JsonResponse({'error': 'Ítem no encontrado'}, status=404)
        
        ubicacion = ""
        ubicacion_completa = {}
        if item.ambiente:
            amb = item.ambiente
            if amb.pabellon and amb.pabellon.sede and amb.pabellon.sede.campus:
                ubicacion = f"{amb.pabellon.sede.campus.nombre} > {amb.pabellon.sede.nombre} > Pab. {amb.pabellon.letra} - {amb.nombre}"
                ubicacion_completa = {
                    'campus_id': amb.pabellon.sede.campus_id,
                    'campus': amb.pabellon.sede.campus.nombre,
                    'sede_id': amb.pabellon.sede_id,
                    'sede': amb.pabellon.sede.nombre,
                    'pabellon_id': amb.pabellon_id,
                    'pabellon': amb.pabellon.nombre,
                    'ambiente_id': amb.id,
                    'ambiente': amb.nombre,
                    'piso': amb.piso
                }
            else:
                ubicacion = amb.nombre
                ubicacion_completa = {
                    'ambiente_id': amb.id,
                    'ambiente': amb.nombre,
                    'piso': amb.piso
                }

        return JsonResponse({
            'id': item.id,
            'codigo_utp': item.codigo_utp,
            'serie': item.serie,
            'nombre': item.nombre,
            'descripcion': item.descripcion,
            'area': item.area.nombre if item.area else '',
            'area_id': item.area_id,
            'tipo': item.tipo_item.nombre if item.tipo_item else '',
            'estado': item.estado,
            'estado_display': item.get_estado_display(),
            'ubicacion': ubicacion,
            'ubicacion_completa': ubicacion_completa,
            'usuario_asignado': item.usuario_asignado.get_full_name() if item.usuario_asignado else None,
            'usuario_asignado_id': item.usuario_asignado_id,
        })


# ============================================================================
# VISTAS CRUD PARA CAMPUS
# ============================================================================

class CampusListView(SupervisorRequeridoMixin, CampusFilterMixin, ListView):
    """Listar campus según permisos del usuario."""
    model = Campus
    template_name = 'productos/campus_list.html'
    context_object_name = 'campus_list'

    def test_func(self):
        # Admin, supervisor y gerente pueden acceder
        return self.get_user_rol() in ['admin', 'supervisor', 'gerente']

    def get_queryset(self):
        queryset = Campus.objects.annotate(
            total_sedes=Count('sedes')
        ).order_by('nombre')

        # Filtrar por campus permitidos según rol
        queryset = self.filtrar_por_campus(queryset, 'pk')

        # Búsqueda
        q = self.request.GET.get('q')
        if q:
            queryset = queryset.filter(
                Q(nombre__icontains=q) | Q(codigo__icontains=q)
            )

        # Filtro por estado
        activo = self.request.GET.get('activo')
        if activo == '1':
            queryset = queryset.filter(activo=True)
        elif activo == '0':
            queryset = queryset.filter(activo=False)

        return queryset
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['titulo'] = 'Gestión de Campus'
        return context


class CampusCreateView(AdminRequeridoMixin, CreateView):
    """Crear campus (solo admins)."""
    model = Campus
    form_class = CampusForm
    template_name = 'productos/campus_form.html'
    success_url = reverse_lazy('productos:campus-list')
    
    def form_valid(self, form):
        messages.success(self.request, 'Campus creado exitosamente.')
        return super().form_valid(form)
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['titulo'] = 'Crear Campus'
        return context


class CampusUpdateView(AdminRequeridoMixin, UpdateView):
    """Editar campus (solo admins)."""
    model = Campus
    form_class = CampusForm
    template_name = 'productos/campus_form.html'
    success_url = reverse_lazy('productos:campus-list')
    
    def form_valid(self, form):
        messages.success(self.request, 'Campus actualizado exitosamente.')
        return super().form_valid(form)
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['titulo'] = 'Editar Campus'
        context['editando'] = True
        return context


class CampusDeleteView(AdminRequeridoMixin, DeleteView):
    """Eliminar campus (solo admins)."""
    model = Campus
    template_name = 'productos/campus_confirm_delete.html'
    success_url = reverse_lazy('productos:campus-list')
    
    def form_valid(self, form):
        messages.success(self.request, 'Campus eliminado exitosamente.')
        return super().form_valid(form)


# ============================================================================
# VISTAS CRUD PARA SEDES
# ============================================================================

class SedeListView(SupervisorRequeridoMixin, CampusFilterMixin, ListView):
    """Listar sedes según permisos de campus del usuario."""
    model = Sede
    template_name = 'productos/sede_list.html'
    context_object_name = 'sedes'

    def test_func(self):
        # Admin, supervisor y gerente pueden acceder
        return self.get_user_rol() in ['admin', 'supervisor', 'gerente']

    def get_queryset(self):
        queryset = Sede.objects.select_related('campus').annotate(
            total_pabellones=Count('pabellones')
        ).order_by('campus__nombre', 'nombre')

        # Filtrar por campus permitidos del usuario
        queryset = self.filtrar_por_campus(queryset, 'campus')

        # Búsqueda
        q = self.request.GET.get('q')
        if q:
            queryset = queryset.filter(
                Q(nombre__icontains=q) | Q(codigo__icontains=q) | Q(campus__nombre__icontains=q)
            )

        # Filtro por campus (solo si tiene acceso a múltiples)
        campus_id = self.request.GET.get('campus')
        if campus_id:
            queryset = queryset.filter(campus_id=campus_id)

        # Filtro por estado
        activo = self.request.GET.get('activo')
        if activo == '1':
            queryset = queryset.filter(activo=True)
        elif activo == '0':
            queryset = queryset.filter(activo=False)

        return queryset

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['titulo'] = 'Gestión de Sedes'
        # Solo mostrar campus que el usuario puede ver
        context['campus_list'] = self.get_campus_permitidos()
        return context


class SedeCreateView(AdminRequeridoMixin, CreateView):
    """Crear sede (solo admins)."""
    model = Sede
    form_class = SedeForm
    template_name = 'productos/sede_form.html'
    success_url = reverse_lazy('productos:sede-list')
    
    def form_valid(self, form):
        messages.success(self.request, 'Sede creada exitosamente.')
        return super().form_valid(form)
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['titulo'] = 'Crear Sede'
        return context


class SedeUpdateView(AdminRequeridoMixin, UpdateView):
    """Editar sede (solo admins)."""
    model = Sede
    form_class = SedeForm
    template_name = 'productos/sede_form.html'
    success_url = reverse_lazy('productos:sede-list')
    
    def form_valid(self, form):
        messages.success(self.request, 'Sede actualizada exitosamente.')
        return super().form_valid(form)
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['titulo'] = 'Editar Sede'
        context['editando'] = True
        return context


class SedeDeleteView(AdminRequeridoMixin, DeleteView):
    """Eliminar sede (solo admins)."""
    model = Sede
    template_name = 'productos/sede_confirm_delete.html'
    success_url = reverse_lazy('productos:sede-list')
    
    def form_valid(self, form):
        messages.success(self.request, 'Sede eliminada exitosamente.')
        return super().form_valid(form)


# ============================================================================
# VISTAS CRUD PARA PABELLONES
# ============================================================================

class PabellonListView(SupervisorRequeridoMixin, CampusFilterMixin, ListView):
    """Listar pabellones según permisos de campus del usuario."""
    model = Pabellon
    template_name = 'productos/pabellon_list.html'
    context_object_name = 'pabellones'

    def test_func(self):
        # Admin, supervisor y gerente pueden acceder
        return self.get_user_rol() in ['admin', 'supervisor', 'gerente']

    def get_queryset(self):
        queryset = Pabellon.objects.select_related('sede', 'sede__campus').annotate(
            total_ambientes=Count('ambientes')
        ).order_by('sede__campus__nombre', 'sede__nombre', 'nombre')

        # Filtrar por campus permitidos del usuario
        queryset = self.filtrar_por_campus(queryset, 'sede__campus')

        # Búsqueda
        q = self.request.GET.get('q')
        if q:
            queryset = queryset.filter(
                Q(nombre__icontains=q) | Q(sede__nombre__icontains=q) | Q(sede__campus__nombre__icontains=q)
            )

        # Filtro por campus
        campus_id = self.request.GET.get('campus')
        if campus_id:
            queryset = queryset.filter(sede__campus_id=campus_id)

        # Filtro por sede
        sede_id = self.request.GET.get('sede')
        if sede_id:
            queryset = queryset.filter(sede_id=sede_id)

        # Filtro por estado
        activo = self.request.GET.get('activo')
        if activo == '1':
            queryset = queryset.filter(activo=True)
        elif activo == '0':
            queryset = queryset.filter(activo=False)

        return queryset

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['titulo'] = 'Gestión de Pabellones'
        # Solo campus permitidos
        context['campus_list'] = self.get_campus_permitidos()
        # Sedes filtradas por campus permitidos
        campus_ids = list(self.get_campus_permitidos().values_list('id', flat=True))
        context['sedes'] = Sede.objects.filter(activo=True, campus_id__in=campus_ids).select_related('campus')
        return context


class PabellonCreateView(AdminRequeridoMixin, CreateView):
    """Crear pabellón (solo admins)."""
    model = Pabellon
    form_class = PabellonForm
    template_name = 'productos/pabellon_form.html'
    success_url = reverse_lazy('productos:pabellon-list')
    
    def form_valid(self, form):
        messages.success(self.request, 'Pabellón creado exitosamente.')
        return super().form_valid(form)
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['titulo'] = 'Crear Pabellón'
        return context


class PabellonUpdateView(AdminRequeridoMixin, UpdateView):
    """Editar pabellón (solo admins)."""
    model = Pabellon
    form_class = PabellonForm
    template_name = 'productos/pabellon_form.html'
    success_url = reverse_lazy('productos:pabellon-list')
    
    def form_valid(self, form):
        messages.success(self.request, 'Pabellón actualizado exitosamente.')
        return super().form_valid(form)
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['titulo'] = 'Editar Pabellón'
        context['editando'] = True
        return context


class PabellonDeleteView(AdminRequeridoMixin, DeleteView):
    """Eliminar pabellón (solo admins)."""
    model = Pabellon
    template_name = 'productos/pabellon_confirm_delete.html'
    success_url = reverse_lazy('productos:pabellon-list')
    
    def form_valid(self, form):
        messages.success(self.request, 'Pabellón eliminado exitosamente.')
        return super().form_valid(form)


# ============================================================================
# PROVEEDORES (Solo supervisor/admin)
# ============================================================================

class ProveedorListView(SupervisorRequeridoMixin, ListView):
    """Lista de proveedores."""
    model = Proveedor
    template_name = 'productos/proveedor_list.html'
    context_object_name = 'proveedores'
    paginate_by = 20
    
    def get_queryset(self):
        queryset = Proveedor.objects.all()
        q = self.request.GET.get('q')
        if q:
            queryset = queryset.filter(
                Q(ruc__icontains=q) |
                Q(razon_social__icontains=q) |
                Q(nombre_comercial__icontains=q)
            )
        return queryset


class ProveedorCreateView(SupervisorRequeridoMixin, CreateView):
    """Crear proveedor."""
    model = Proveedor
    template_name = 'productos/proveedor_form.html'
    fields = ['ruc', 'razon_social', 'nombre_comercial', 'direccion', 'telefono', 'email', 'contacto']
    success_url = reverse_lazy('productos:proveedor-list')
    
    def form_valid(self, form):
        messages.success(self.request, 'Proveedor creado exitosamente.')
        return super().form_valid(form)


class ProveedorDetailView(SupervisorRequeridoMixin, DetailView):
    """Detalle de proveedor."""
    model = Proveedor
    template_name = 'productos/proveedor_detail.html'
    context_object_name = 'proveedor'
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['contratos'] = self.object.contratos.all()[:10]
        return context


class ProveedorUpdateView(SupervisorRequeridoMixin, UpdateView):
    """Editar proveedor."""
    model = Proveedor
    template_name = 'productos/proveedor_form.html'
    fields = ['ruc', 'razon_social', 'nombre_comercial', 'direccion', 'telefono', 'email', 'contacto', 'activo']
    
    def get_success_url(self):
        return reverse('productos:proveedor-detail', kwargs={'pk': self.object.pk})
    
    def form_valid(self, form):
        messages.success(self.request, 'Proveedor actualizado exitosamente.')
        return super().form_valid(form)


# ============================================================================
# CONTRATOS (Solo supervisor/admin)
# ============================================================================

class ContratoListView(SupervisorRequeridoMixin, ListView):
    """Lista de contratos."""
    model = Contrato
    template_name = 'productos/contrato_list.html'
    context_object_name = 'contratos'
    paginate_by = 20
    
    def get_queryset(self):
        queryset = Contrato.objects.select_related('proveedor')
        q = self.request.GET.get('q')
        if q:
            queryset = queryset.filter(
                Q(numero_contrato__icontains=q) |
                Q(proveedor__razon_social__icontains=q)
            )
        estado = self.request.GET.get('estado')
        if estado:
            queryset = queryset.filter(estado=estado)
        return queryset
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['estados'] = Contrato.ESTADOS
        return context


class ContratoCreateView(SupervisorRequeridoMixin, CreateView):
    """Crear contrato."""
    model = Contrato
    template_name = 'productos/contrato_form.html'
    fields = ['numero_contrato', 'proveedor', 'descripcion', 'fecha_inicio', 'fecha_fin', 'monto_total', 'estado', 'observaciones']
    success_url = reverse_lazy('productos:contrato-list')
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['proveedores'] = Proveedor.objects.filter(activo=True)
        return context
    
    def form_valid(self, form):
        form.instance.creado_por = self.request.user
        messages.success(self.request, 'Contrato creado exitosamente.')
        return super().form_valid(form)


class ContratoDetailView(SupervisorRequeridoMixin, DetailView):
    """Detalle de contrato."""
    model = Contrato
    template_name = 'productos/contrato_detail.html'
    context_object_name = 'contrato'
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['anexos'] = self.object.anexos.all()
        context['lotes'] = self.object.lotes.all()
        context['items_directos'] = self.object.items_directos.all()[:20]
        return context


class ContratoUpdateView(SupervisorRequeridoMixin, UpdateView):
    """Editar contrato."""
    model = Contrato
    template_name = 'productos/contrato_form.html'
    fields = ['numero_contrato', 'proveedor', 'descripcion', 'fecha_inicio', 'fecha_fin', 'monto_total', 'estado', 'observaciones']
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['proveedores'] = Proveedor.objects.filter(activo=True)
        context['editando'] = True
        return context
    
    def get_success_url(self):
        return reverse('productos:contrato-detail', kwargs={'pk': self.object.pk})
    
    def form_valid(self, form):
        messages.success(self.request, 'Contrato actualizado exitosamente.')
        return super().form_valid(form)


class AnexoContratoCreateView(SupervisorRequeridoMixin, CreateView):
    """Crear anexo de contrato."""
    model = AnexoContrato
    template_name = 'productos/anexo_form.html'
    fields = ['numero_anexo', 'fecha', 'descripcion', 'monto_modificacion']
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['contrato'] = get_object_or_404(Contrato, pk=self.kwargs['pk'])
        return context
    
    def form_valid(self, form):
        form.instance.contrato = get_object_or_404(Contrato, pk=self.kwargs['pk'])
        messages.success(self.request, 'Anexo creado exitosamente.')
        return super().form_valid(form)
    
    def get_success_url(self):
        return reverse('productos:contrato-detail', kwargs={'pk': self.kwargs['pk']})


# ============================================================================
# LOTES
# ============================================================================

class LoteListView(PerfilRequeridoMixin, ListView):
    """Lista de lotes."""
    model = Lote
    template_name = 'productos/lote_list.html'
    context_object_name = 'lotes'
    paginate_by = 20
    
    def get_queryset(self):
        queryset = Lote.objects.select_related('contrato__proveedor').annotate(
            num_items=Count('items')
        )
        q = self.request.GET.get('q')
        if q:
            queryset = queryset.filter(
                Q(codigo_lote__icontains=q) |
                Q(codigo_interno__icontains=q) |
                Q(descripcion__icontains=q)
            )
        return queryset
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        # Ocultar info de contrato para operadores
        context['mostrar_contrato'] = self.get_user_rol() in ['admin', 'supervisor']
        return context


class LoteCreateView(SupervisorRequeridoMixin, CreateView):
    """Crear lote - solo supervisores y admins."""
    model = Lote
    template_name = 'productos/lote_form.html'
    fields = ['codigo_lote', 'contrato', 'descripcion', 'fecha_adquisicion', 'observaciones']
    success_url = reverse_lazy('productos:lote-list')
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['contratos'] = Contrato.objects.filter(estado='vigente')
        context['mostrar_contrato'] = True
        return context
    
    def form_valid(self, form):
        form.instance.creado_por = self.request.user
        messages.success(self.request, 'Lote creado exitosamente.')
        return super().form_valid(form)


class LoteDetailView(PerfilRequeridoMixin, DetailView):
    """Detalle de lote."""
    model = Lote
    template_name = 'productos/lote_detail.html'
    context_object_name = 'lote'
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['items'] = self.object.items.select_related('area', 'tipo_item', 'ambiente')[:50]
        context['mostrar_contrato'] = self.get_user_rol() in ['admin', 'supervisor']
        
        # Alertas de garantía agrupadas
        context['alertas_garantia'] = self.object.items_por_garantia
        return context


class LoteUpdateView(SupervisorRequeridoMixin, UpdateView):
    """Editar lote - solo supervisores y admins."""
    model = Lote
    template_name = 'productos/lote_form.html'
    fields = ['codigo_lote', 'contrato', 'descripcion', 'fecha_adquisicion', 'observaciones', 'activo']

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['editando'] = True
        context['contratos'] = Contrato.objects.filter(estado='vigente')
        context['mostrar_contrato'] = True
        return context

    def get_success_url(self):
        return reverse('productos:lote-detail', kwargs={'pk': self.object.pk})

    def form_valid(self, form):
        messages.success(self.request, 'Lote actualizado exitosamente.')
        return super().form_valid(form)


# ============================================================================
# IMPORTACIÓN MASIVA DESDE EXCEL
# ============================================================================

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils.exceptions import InvalidFileException
from django.http import HttpResponse
from django.db import transaction
from datetime import datetime
import io
import logging

logger = logging.getLogger(__name__)


class ItemImportarPlantillaView(PerfilRequeridoMixin, View):
    """Descargar plantilla Excel para importación masiva."""

    def get(self, request):
        # Crear libro de Excel
        wb = Workbook()

        # Hoja 1: Plantilla de datos
        ws = wb.active
        ws.title = "Plantilla Items"

        # Definir encabezados
        headers_obligatorios = [
            'serie', 'nombre', 'area', 'tipo_item', 'precio', 'fecha_adquisicion'
        ]
        headers_opcionales = [
            'codigo_utp', 'descripcion', 'ambiente_codigo', 'estado', 'garantia_hasta',
            'observaciones', 'lote_codigo', 'es_leasing', 'leasing_empresa',
            'leasing_contrato', 'leasing_vencimiento'
        ]
        headers_sistemas = [
            'marca', 'modelo', 'procesador', 'generacion_procesador',
            'ram_total_gb', 'ram_configuracion', 'ram_tipo',
            'almacenamiento_gb', 'almacenamiento_tipo', 'sistema_operativo'
        ]

        headers = headers_obligatorios + headers_opcionales + headers_sistemas

        # Estilo para encabezados obligatorios
        fill_obligatorio = PatternFill(start_color="C8102E", end_color="C8102E", fill_type="solid")
        font_obligatorio = Font(bold=True, color="FFFFFF")

        # Estilo para encabezados opcionales
        fill_opcional = PatternFill(start_color="4A4A4A", end_color="4A4A4A", fill_type="solid")
        font_opcional = Font(bold=True, color="FFFFFF")

        # Estilo para encabezados de sistemas
        fill_sistemas = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
        font_sistemas = Font(bold=True, color="FFFFFF")

        # Escribir encabezados con estilos
        for idx, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=idx, value=header)
            cell.alignment = Alignment(horizontal='center', vertical='center')

            if header in headers_obligatorios:
                cell.fill = fill_obligatorio
                cell.font = font_obligatorio
            elif header in headers_sistemas:
                cell.fill = fill_sistemas
                cell.font = font_sistemas
            else:
                cell.fill = fill_opcional
                cell.font = font_opcional

        # Ajustar ancho de columnas
        for idx in range(1, len(headers) + 1):
            ws.column_dimensions[ws.cell(row=1, column=idx).column_letter].width = 18

        # Agregar filas de ejemplo
        ejemplo1 = [
            'SN123456789',  # serie
            'Laptop Dell Latitude 5430',  # nombre
            'sistemas',  # area
            'Laptop',  # tipo_item
            '3500.00',  # precio
            '2026-01-10',  # fecha_adquisicion
            'UTP296375',  # codigo_utp (opcional, puede ser PENDIENTE o vacío)
            'Laptop corporativa i7',  # descripcion
            '',  # ambiente_codigo
            'nuevo',  # estado
            '2028-01-10',  # garantia_hasta
            '',  # observaciones
            '',  # lote_codigo
            'NO',  # es_leasing
            '',  # leasing_empresa
            '',  # leasing_contrato
            '',  # leasing_vencimiento
            'Dell',  # marca
            'Latitude 5430',  # modelo
            'Intel Core i7-1365U',  # procesador
            '13th Gen',  # generacion_procesador
            '16',  # ram_total_gb
            '2x8GB',  # ram_configuracion
            'DDR4',  # ram_tipo
            '512',  # almacenamiento_gb
            'NVMe',  # almacenamiento_tipo
            'Windows 11 Pro'  # sistema_operativo
        ]

        ejemplo2 = [
            'SN987654321',  # serie
            'Silla ergonómica',  # nombre
            'operaciones',  # area
            'Silla',  # tipo_item
            '450.00',  # precio
            '2026-01-10',  # fecha_adquisicion
            'PENDIENTE',  # codigo_utp (aún sin etiqueta de logística)
            'Silla de oficina con soporte lumbar',  # descripcion
            '',  # ambiente_codigo
            'nuevo',  # estado
            '',  # garantia_hasta
            '',  # observaciones
            '',  # lote_codigo
            'NO',  # es_leasing
        ]
        # Rellenar con vacíos para completar las columnas
        ejemplo2.extend([''] * (len(headers) - len(ejemplo2)))

        ws.append(ejemplo1)
        ws.append(ejemplo2)

        # Hoja 2: Instrucciones
        ws_instrucciones = wb.create_sheet("Instrucciones")
        instrucciones = [
            ["INSTRUCCIONES PARA IMPORTACIÓN MASIVA DE ÍTEMS", ""],
            ["", ""],
            ["1. COLUMNAS OBLIGATORIAS (Rojo):", ""],
            ["   - serie: Número de serie único del fabricante", ""],
            ["   - nombre: Nombre descriptivo del ítem", ""],
            ["   - area: sistemas, operaciones o laboratorio", ""],
            ["   - tipo_item: Debe existir en el sistema para el área", ""],
            ["   - precio: Precio en formato numérico (ej: 1500.00)", ""],
            ["   - fecha_adquisicion: Formato YYYY-MM-DD (ej: 2026-01-10)", ""],
            ["", ""],
            ["2. COLUMNAS OPCIONALES (Gris):", ""],
            ["   - codigo_utp: Código de etiqueta física (ej: UTP296375)", ""],
            ["     Dejar vacío o PENDIENTE si aún no tiene etiqueta de logística", ""],
            ["     Formato: UTP seguido de números", ""],
            ["   - descripcion: Descripción adicional", ""],
            ["   - ambiente_codigo: Código del ambiente (ej: CLN-SP-A-P1-LC-001)", ""],
            ["   - estado: nuevo, instalado, dañado u obsoleto (default: nuevo)", ""],
            ["   - garantia_hasta: Fecha en formato YYYY-MM-DD", ""],
            ["   - lote_codigo: Código del lote existente (ej: LOT-2026-0001)", ""],
            ["   - es_leasing: SI o NO", ""],
            ["", ""],
            ["3. COLUMNAS PARA SISTEMAS (Azul):", ""],
            ["   - Solo usar si area = sistemas", ""],
            ["   - Especificaciones técnicas del equipo", ""],
            ["", ""],
            ["4. NOTAS IMPORTANTES:", ""],
            ["   - El código interno se genera automáticamente (ej: SIS-2026-0001)", ""],
            ["   - El código UTP es la etiqueta física de logística (opcional)", ""],
            ["   - La serie debe ser única en el sistema", ""],
            ["   - Máximo 1000 ítems por archivo", ""],
            ["   - Formato de archivo: .xlsx", ""],
            ["", ""],
            ["5. VALIDACIONES:", ""],
            ["   - Serie única y no vacía", ""],
            ["   - Área válida y coincide con tu perfil", ""],
            ["   - Tipo de ítem existe para el área", ""],
            ["   - Precio numérico positivo", ""],
            ["   - Fechas en formato correcto", ""],
        ]

        for row_data in instrucciones:
            ws_instrucciones.append(row_data)

        # Ajustar ancho de columnas en instrucciones
        ws_instrucciones.column_dimensions['A'].width = 60
        ws_instrucciones.column_dimensions['B'].width = 20

        # Estilo para el título
        ws_instrucciones['A1'].font = Font(bold=True, size=14, color="C8102E")

        # Preparar respuesta HTTP
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = 'attachment; filename=plantilla_items_inventario.xlsx'

        wb.save(response)
        return response


class ItemImportarView(RateLimitMixin, PerfilRequeridoMixin, TemplateView):
    """Vista para subir archivo Excel y mostrar preview con validaciones."""
    ratelimit_key = 'import'
    ratelimit_method = 'POST'

    template_name = 'productos/item_importar.html'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['areas'] = Area.objects.filter(activo=True)
        context['lotes'] = Lote.objects.filter(activo=True)
        return context

    def post(self, request, *args, **kwargs):
        archivo = request.FILES.get('archivo_excel')
        crear_lote = request.POST.get('crear_lote') == 'on'
        lote_descripcion = request.POST.get('lote_descripcion', '')
        lote_existente_id = request.POST.get('lote_existente')

        if not archivo:
            messages.error(request, 'Debe seleccionar un archivo Excel.')
            return redirect('productos:item-importar')

        # CRÍTICO: Validar tamaño de archivo antes de procesarlo
        MAX_FILE_SIZE = 10 * 1024 * 1024  # 10MB
        if archivo.size > MAX_FILE_SIZE:
            messages.error(request, f'El archivo es demasiado grande. Tamaño máximo: 10MB')
            logger.warning(f'Usuario {request.user.username} intentó subir archivo de {archivo.size} bytes')
            return redirect('productos:item-importar')

        if not archivo.name.endswith('.xlsx'):
            messages.error(request, 'El archivo debe ser formato .xlsx')
            return redirect('productos:item-importar')

        try:
            # Cargar archivo Excel
            wb = load_workbook(archivo, data_only=True)
            ws = wb.active

            # Leer encabezados
            headers = [cell.value for cell in ws[1]]

            # Validar que existan las columnas obligatorias
            required_headers = ['serie', 'nombre', 'area', 'tipo_item', 'precio', 'fecha_adquisicion']
            for header in required_headers:
                if header not in headers:
                    messages.error(request, f'Falta la columna obligatoria: {header}')
                    return redirect('productos:item-importar')

            # Procesar filas
            items_preview = []
            errores_globales = []
            series_en_archivo = set()  # CRÍTICO: Detectar series duplicadas dentro del Excel

            # Obtener perfil del usuario
            perfil = request.user.perfil
            area_usuario = perfil.area if perfil.rol != 'admin' else None

            for idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
                if not any(row):  # Fila vacía
                    continue

                if len(items_preview) >= 1000:
                    errores_globales.append('Se alcanzó el límite de 1000 ítems. Las filas restantes no se procesaron.')
                    break

                # Crear diccionario de datos
                item_data = dict(zip(headers, row))

                # Función auxiliar para obtener valor como string
                def get_str(value):
                    if value is None:
                        return ''
                    return str(value).strip()

                # Validaciones
                errores = []
                advertencias = []

                # Validar serie
                serie = get_str(item_data.get('serie'))
                if not serie:
                    errores.append('Serie vacía')
                elif serie in series_en_archivo:
                    # CRÍTICO: Detectar duplicados dentro del mismo archivo
                    errores.append(f'Serie {serie} duplicada dentro del archivo')
                elif Item.objects.filter(serie=serie).exists():
                    errores.append(f'Serie {serie} ya existe en el sistema')
                else:
                    # Agregar a lista de series procesadas
                    series_en_archivo.add(serie)

                # Validar código UTP (opcional)
                codigo_utp = get_str(item_data.get('codigo_utp', 'PENDIENTE')).upper()
                if not codigo_utp:
                    codigo_utp = 'PENDIENTE'
                    advertencias.append('Código UTP pendiente - se asignará etiqueta de logística posteriormente')
                elif codigo_utp != 'PENDIENTE':
                    # Validar formato UTP + números
                    import re
                    if not re.match(r'^UTP\d+$', codigo_utp):
                        errores.append(f'Código UTP "{codigo_utp}" inválido - debe ser UTP seguido de números (ej: UTP296375)')
                    elif Item.objects.filter(codigo_utp=codigo_utp).exists():
                        errores.append(f'Código UTP {codigo_utp} ya existe en el sistema')
                else:
                    # Es PENDIENTE
                    advertencias.append('Código UTP pendiente - se asignará etiqueta de logística posteriormente')

                # Validar área
                area_codigo = get_str(item_data.get('area')).lower()
                if area_codigo not in ['sistemas', 'operaciones', 'laboratorio']:
                    errores.append(f'Área inválida: {area_codigo}')
                elif area_usuario and area_usuario.codigo != area_codigo:
                    errores.append(f'No tienes permiso para crear ítems en el área {area_codigo}')

                # Validar tipo_item
                tipo_item_nombre = get_str(item_data.get('tipo_item'))
                tipo_item = None
                if tipo_item_nombre and area_codigo in ['sistemas', 'operaciones', 'laboratorio']:
                    try:
                        area_obj = Area.objects.get(codigo=area_codigo)
                        tipo_item = TipoItem.objects.get(nombre__iexact=tipo_item_nombre, area=area_obj)
                    except TipoItem.DoesNotExist:
                        errores.append(f'Tipo de ítem "{tipo_item_nombre}" no existe para el área {area_codigo}')
                    except Area.DoesNotExist:
                        pass

                # Validar precio
                precio = item_data.get('precio', '')
                try:
                    precio_decimal = float(str(precio).replace(',', '')) if precio else 0
                    if precio_decimal <= 0:
                        errores.append('Precio debe ser mayor a 0')
                except (ValueError, TypeError):
                    errores.append(f'Precio inválido: {precio}')

                # Validar fecha_adquisicion
                fecha_adq = item_data.get('fecha_adquisicion', '')
                fecha_adq_obj = None
                if fecha_adq:
                    try:
                        if isinstance(fecha_adq, datetime):
                            fecha_adq_obj = fecha_adq.date()
                        else:
                            fecha_adq_str = str(fecha_adq).strip()
                            # Intentar parsear varios formatos
                            for fmt in ['%Y-%m-%d', '%d/%m/%Y', '%d-%m-%Y']:
                                try:
                                    fecha_adq_obj = datetime.strptime(fecha_adq_str, fmt).date()
                                    break
                                except ValueError:
                                    continue
                            if not fecha_adq_obj:
                                errores.append(f'Fecha de adquisición inválida: {fecha_adq}')
                    except:
                        errores.append(f'Fecha de adquisición inválida: {fecha_adq}')
                else:
                    errores.append('Fecha de adquisición vacía')

                # Validar ambiente (opcional)
                ambiente_codigo = get_str(item_data.get('ambiente_codigo'))
                ambiente = None
                if ambiente_codigo:
                    try:
                        ambiente = Ambiente.objects.get(codigo=ambiente_codigo, activo=True)
                    except Ambiente.DoesNotExist:
                        errores.append(f'Ambiente {ambiente_codigo} no existe')
                else:
                    advertencias.append('Sin ubicación asignada')

                # Validar lote (opcional)
                lote_codigo = get_str(item_data.get('lote_codigo'))
                lote = None
                if lote_codigo:
                    try:
                        lote = Lote.objects.get(codigo_interno=lote_codigo, activo=True)
                    except Lote.DoesNotExist:
                        errores.append(f'Lote {lote_codigo} no existe')

                # Validar estado
                estado = get_str(item_data.get('estado', 'nuevo')).lower()
                if estado not in ['nuevo', 'instalado', 'dañado', 'obsoleto']:
                    advertencias.append(f'Estado "{estado}" inválido, se usará "nuevo"')
                    estado = 'nuevo'

                # Validar garantía
                garantia_hasta = item_data.get('garantia_hasta', '')
                garantia_obj = None
                if garantia_hasta:
                    try:
                        if isinstance(garantia_hasta, datetime):
                            garantia_obj = garantia_hasta.date()
                        else:
                            garantia_str = str(garantia_hasta).strip()
                            for fmt in ['%Y-%m-%d', '%d/%m/%Y', '%d-%m-%Y']:
                                try:
                                    garantia_obj = datetime.strptime(garantia_str, fmt).date()
                                    break
                                except ValueError:
                                    continue
                            if not garantia_obj:
                                advertencias.append(f'Fecha de garantía inválida: {garantia_hasta}')
                    except:
                        advertencias.append(f'Fecha de garantía inválida: {garantia_hasta}')
                else:
                    advertencias.append('Sin fecha de garantía')

                # Determinar estado de la fila
                if errores:
                    estado_fila = 'error'
                elif advertencias:
                    estado_fila = 'advertencia'
                else:
                    estado_fila = 'ok'

                items_preview.append({
                    'fila': idx,
                    'data': item_data,
                    'errores': errores,
                    'advertencias': advertencias,
                    'estado': estado_fila
                })

            # Contar estados
            total = len(items_preview)
            con_errores = sum(1 for item in items_preview if item['estado'] == 'error')
            con_advertencias = sum(1 for item in items_preview if item['estado'] == 'advertencia')
            validos = sum(1 for item in items_preview if item['estado'] == 'ok')

            # Guardar en sesión para confirmar después
            request.session['items_preview'] = items_preview
            request.session['crear_lote'] = crear_lote
            request.session['lote_descripcion'] = lote_descripcion
            request.session['lote_existente_id'] = lote_existente_id

            context = self.get_context_data()
            context['items_preview'] = items_preview
            context['total'] = total
            context['con_errores'] = con_errores
            context['con_advertencias'] = con_advertencias
            context['validos'] = validos
            context['puede_importar'] = con_errores == 0 and total > 0
            context['errores_globales'] = errores_globales
            context['crear_lote'] = crear_lote
            context['lote_descripcion'] = lote_descripcion

            return self.render_to_response(context)

        except InvalidFileException:
            messages.error(request, 'El archivo Excel está corrupto o no es válido.')
            logger.error(f'Archivo Excel inválido subido por {request.user.username}')
            return redirect('productos:item-importar')
        except MemoryError:
            messages.error(request, 'El archivo es demasiado grande para procesar.')
            logger.error(f'MemoryError al procesar archivo de {request.user.username}')
            return redirect('productos:item-importar')
        except Exception as e:
            # Log detallado para debugging, mensaje genérico para usuario
            logger.exception(f'Error inesperado en importación de {request.user.username}: {e}')
            messages.error(request, 'Ocurrió un error al procesar el archivo. Por favor, contacte al administrador.')
            return redirect('productos:item-importar')


class ItemImportarConfirmarView(PerfilRequeridoMixin, View):
    """Confirmar y ejecutar la importación masiva."""

    def post(self, request):
        # Función auxiliar para obtener valor como string
        def get_str(value):
            if value is None:
                return ''
            return str(value).strip()

        items_preview = request.session.get('items_preview', [])
        crear_lote = request.session.get('crear_lote', False)
        lote_descripcion = request.session.get('lote_descripcion', '')
        lote_existente_id = request.session.get('lote_existente_id')

        if not items_preview:
            messages.error(request, 'No hay datos para importar. Por favor sube el archivo nuevamente.')
            return redirect('productos:item-importar')

        # Validar que no haya errores
        items_validos = [item for item in items_preview if item['estado'] != 'error']

        if not items_validos:
            messages.error(request, 'No hay ítems válidos para importar.')
            return redirect('productos:item-importar')

        try:
            with transaction.atomic():
                # Crear lote si es necesario
                lote = None
                if crear_lote and lote_descripcion:
                    lote = Lote.objects.create(
                        descripcion=lote_descripcion,
                        fecha_adquisicion=timezone.now().date(),
                        creado_por=request.user
                    )
                elif lote_existente_id:
                    lote = Lote.objects.get(pk=lote_existente_id)

                items_creados = []

                for item_info in items_validos:
                    data = item_info['data']

                    # Obtener área
                    area = Area.objects.get(codigo=get_str(data.get('area')).lower())

                    # Obtener tipo_item
                    tipo_item = TipoItem.objects.get(
                        nombre__iexact=get_str(data.get('tipo_item')),
                        area=area
                    )

                    # Obtener código UTP del Excel (opcional, default PENDIENTE)
                    codigo_utp = get_str(data.get('codigo_utp', 'PENDIENTE')).upper()
                    if not codigo_utp:
                        codigo_utp = 'PENDIENTE'

                    # Parsear fecha de adquisición
                    fecha_adq = data.get('fecha_adquisicion', '')
                    if isinstance(fecha_adq, datetime):
                        fecha_adq_obj = fecha_adq.date()
                    else:
                        fecha_adq_str = str(fecha_adq).strip()
                        for fmt in ['%Y-%m-%d', '%d/%m/%Y', '%d-%m-%Y']:
                            try:
                                fecha_adq_obj = datetime.strptime(fecha_adq_str, fmt).date()
                                break
                            except ValueError:
                                continue

                    # Parsear garantía
                    garantia_hasta = data.get('garantia_hasta', '')
                    garantia_obj = None
                    if garantia_hasta:
                        try:
                            if isinstance(garantia_hasta, datetime):
                                garantia_obj = garantia_hasta.date()
                            else:
                                garantia_str = str(garantia_hasta).strip()
                                for fmt in ['%Y-%m-%d', '%d/%m/%Y', '%d-%m-%Y']:
                                    try:
                                        garantia_obj = datetime.strptime(garantia_str, fmt).date()
                                        break
                                    except ValueError:
                                        continue
                        except:
                            pass

                    # Obtener ambiente
                    ambiente = None
                    ambiente_codigo = get_str(data.get('ambiente_codigo'))
                    if ambiente_codigo:
                        try:
                            ambiente = Ambiente.objects.get(codigo=ambiente_codigo)
                        except Ambiente.DoesNotExist:
                            pass

                    # Parsear precio
                    precio = float(str(data.get('precio', 0)).replace(',', ''))

                    # Estado
                    estado = get_str(data.get('estado', 'nuevo')).lower()
                    if estado not in ['nuevo', 'instalado', 'dañado', 'obsoleto']:
                        estado = 'nuevo'

                    # Leasing
                    es_leasing_str = get_str(data.get('es_leasing', 'NO')).upper()
                    es_leasing = es_leasing_str in ['SI', 'SÍ', 'YES', 'S', 'Y', '1', 'TRUE']

                    # Crear ítem
                    item = Item.objects.create(
                        codigo_utp=codigo_utp,
                        serie=get_str(data.get('serie')),
                        nombre=get_str(data.get('nombre')),
                        descripcion=get_str(data.get('descripcion')),
                        area=area,
                        tipo_item=tipo_item,
                        ambiente=ambiente,
                        estado=estado,
                        precio=precio,
                        fecha_adquisicion=fecha_adq_obj,
                        garantia_hasta=garantia_obj,
                        observaciones=get_str(data.get('observaciones')),
                        lote=lote,
                        es_leasing=es_leasing,
                        leasing_empresa=get_str(data.get('leasing_empresa')) if es_leasing else '',
                        leasing_contrato=get_str(data.get('leasing_contrato')) if es_leasing else '',
                        creado_por=request.user
                    )

                    # Si es área de sistemas, crear especificaciones
                    if area.codigo == 'sistemas':
                        specs_data = {
                            'marca': get_str(data.get('marca')),
                            'modelo': get_str(data.get('modelo')),
                            'procesador': get_str(data.get('procesador')),
                            'generacion_procesador': get_str(data.get('generacion_procesador')),
                            'ram_total_gb': data.get('ram_total_gb', None),
                            'ram_configuracion': get_str(data.get('ram_configuracion')),
                            'ram_tipo': get_str(data.get('ram_tipo')),
                            'almacenamiento_gb': data.get('almacenamiento_gb', None),
                            'almacenamiento_tipo': get_str(data.get('almacenamiento_tipo')),
                            'sistema_operativo': get_str(data.get('sistema_operativo')),
                        }

                        # Limpiar valores None y vacíos
                        specs_data = {k: v for k, v in specs_data.items() if v}

                        if specs_data:
                            EspecificacionesSistemas.objects.create(
                                item=item,
                                **specs_data
                            )

                    items_creados.append(item)

                # Limpiar sesión
                del request.session['items_preview']
                if 'crear_lote' in request.session:
                    del request.session['crear_lote']
                if 'lote_descripcion' in request.session:
                    del request.session['lote_descripcion']
                if 'lote_existente_id' in request.session:
                    del request.session['lote_existente_id']

                # Mensaje de éxito
                mensaje = f'Se importaron exitosamente {len(items_creados)} ítems.'
                if lote:
                    mensaje += f' Asociados al lote {lote.codigo_interno}.'

                messages.success(request, mensaje)

                # AUDITORÍA: Registrar importación exitosa
                logger.info(
                    f'IMPORT_SUCCESS: User={request.user.username}, '
                    f'Items={len(items_creados)}, Area={perfil.area.codigo if perfil.area else "todas"}, '
                    f'IP={request.META.get("REMOTE_ADDR")}, Lote={lote.codigo_interno if lote else "N/A"}'
                )

                # Redirigir a la lista de ítems
                return redirect('productos:item-list')

        except Exception as e:
            # Log detallado para debugging
            logger.exception(f'Error al confirmar importación de {request.user.username}: {e}')
            messages.error(request, 'Ocurrió un error al importar los ítems. Por favor, contacte al administrador.')
            return redirect('productos:item-importar')





# ==============================================================================
# VISTAS DE REPORTES Y EXPORTACIÓN
# ==============================================================================

from .utils.export_utils import ExcelExporter, PDFExporter, format_currency, format_date, format_boolean


class ReportesView(LoginRequiredMixin, TemplateView):
    """Vista principal para seleccionar y generar reportes"""
    template_name = 'productos/reportes.html'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        perfil = self.request.user.perfil

        if perfil.area:
            items = Item.objects.filter(area=perfil.area)
        else:
            items = Item.objects.all()

        context['total_items'] = items.count()
        context['items_sin_asignar'] = items.filter(colaborador_asignado__isnull=True).count()
        context['items_en_custodia'] = items.filter(estado='custodia').count()
        context['items_mantenimiento'] = items.filter(estado='en_mantenimiento').count()

        if perfil.area:
            context['areas'] = [perfil.area]
        else:
            context['areas'] = Area.objects.filter(activo=True)

        context['tipos_item'] = TipoItem.objects.filter(activo=True)

        # Datos para filtro de Leasing
        from datetime import datetime
        anio_actual = datetime.now().year
        context['anio_actual'] = anio_actual
        context['anios_leasing'] = list(range(anio_actual - 2, anio_actual + 5))

        # Datos para filtro de Especificaciones Técnicas (solo Sistemas)
        from .models import EspecificacionesSistemas
        specs = EspecificacionesSistemas.objects.all()

        # Obtener valores únicos para cada filtro
        context['procesadores'] = list(specs.exclude(procesador__isnull=True).exclude(procesador='').values_list('procesador', flat=True).distinct().order_by('procesador'))
        context['rams'] = list(specs.exclude(ram_total_gb__isnull=True).values_list('ram_total_gb', flat=True).distinct().order_by('ram_total_gb'))
        context['almacenamientos'] = list(specs.exclude(almacenamiento_gb__isnull=True).values_list('almacenamiento_gb', flat=True).distinct().order_by('almacenamiento_gb'))
        context['tipos_disco'] = list(specs.exclude(almacenamiento_tipo__isnull=True).exclude(almacenamiento_tipo='').values_list('almacenamiento_tipo', flat=True).distinct().order_by('almacenamiento_tipo'))
        context['marcas'] = list(specs.exclude(marca__isnull=True).exclude(marca='').values_list('marca', flat=True).distinct().order_by('marca'))
        context['modelos'] = list(specs.exclude(modelo__isnull=True).exclude(modelo='').values_list('modelo', flat=True).distinct().order_by('modelo'))

        return context


class ExportarInventarioExcelView(RateLimitMixin, LoginRequiredMixin, View):
    """Exporta el inventario completo a Excel"""
    ratelimit_key = 'export'

    def get(self, request, *args, **kwargs):
        perfil = request.user.perfil
        area_id = request.GET.get('area')
        tipo_id = request.GET.get('tipo')
        estado = request.GET.get('estado')

        if perfil.area:
            items = Item.objects.filter(area=perfil.area)
        else:
            items = Item.objects.all()

        if area_id:
            items = items.filter(area_id=area_id)
        if tipo_id:
            items = items.filter(tipo_item_id=tipo_id)
        if estado:
            items = items.filter(estado=estado)

        exporter = ExcelExporter(title="Inventario")
        titulo = "INVENTARIO COMPLETO"
        subtitulo = f"Total de ítems: {items.count()}"
        if perfil.area:
            subtitulo += f" | Área: {perfil.area.nombre}"

        exporter.add_title(titulo, subtitulo)

        headers = ['Código Interno', 'Código UTP', 'Serie', 'Nombre', 'Área', 'Tipo', 'Estado',
                   'Ubicación', 'Usuario Asignado', 'Precio', 'Fecha Adquisición', 'Garantía Hasta', 'Leasing']
        exporter.add_headers(headers)

        for idx, item in enumerate(items.select_related('area', 'tipo_item', 'ambiente', 'ambiente__pabellon', 'usuario_asignado')):
            ubicacion = item.ambiente.ubicacion_completa if item.ambiente else 'Sin asignar'
            usuario = item.usuario_asignado.get_full_name() if item.usuario_asignado else 'Sin asignar'
            codigo_mostrar = item.codigo_utp if not item.codigo_utp_pendiente else item.codigo_interno
            row = [codigo_mostrar, item.codigo_utp, item.serie, item.nombre,
                   item.area.nombre if item.area else '',
                   item.tipo_item.nombre if item.tipo_item else '', item.get_estado_display(), ubicacion, usuario,
                   format_currency(item.precio), format_date(item.fecha_adquisicion),
                   format_date(item.garantia_hasta), format_boolean(item.es_leasing)]
            exporter.add_row(row, alternate=(idx % 2 == 0))

        valor_total = items.aggregate(Sum('precio'))['precio__sum'] or 0
        summary = {'Total de ítems': items.count(), 'Valor total': format_currency(valor_total),
                   'Ítems operativos': items.filter(estado='operativo').count(),
                   'Ítems en mantenimiento': items.filter(estado='en_mantenimiento').count(),
                   'Ítems con código UTP pendiente': items.filter(codigo_utp='PENDIENTE').count()}
        exporter.add_summary(summary)

        fecha = timezone.now().strftime('%Y%m%d_%H%M%S')
        return exporter.get_response(f"inventario_{fecha}.xlsx")


class ExportarReportePorAreaExcelView(RateLimitMixin, LoginRequiredMixin, View):
    """Exporta reporte de ítems agrupados por área"""
    ratelimit_key = 'export'

    def get(self, request, *args, **kwargs):
        perfil = request.user.perfil
        if perfil.area:
            areas = Area.objects.filter(id=perfil.area.id, activo=True)
        else:
            areas = Area.objects.filter(activo=True)

        exporter = ExcelExporter(title="Reporte por Área")
        exporter.add_title("REPORTE DE INVENTARIO POR ÁREA", "Distribución de ítems y valores")

        headers = ['Área', 'Cantidad de Ítems', 'Valor Total', '% del Total', 'Operativos', 'En Mantenimiento', 'Dañados']
        exporter.add_headers(headers)

        total_items = Item.objects.count()
        total_valor = Item.objects.aggregate(Sum('precio'))['precio__sum'] or 0

        for idx, area in enumerate(areas):
            items_area = Item.objects.filter(area=area)
            cantidad = items_area.count()
            valor = items_area.aggregate(Sum('precio'))['precio__sum'] or 0
            porcentaje = (valor / total_valor * 100) if total_valor > 0 else 0
            row = [area.nombre, cantidad, format_currency(valor), f"{porcentaje:.2f}%",
                   items_area.filter(estado='operativo').count(),
                   items_area.filter(estado='en_mantenimiento').count(),
                   items_area.filter(estado='danado').count()]
            exporter.add_row(row, alternate=(idx % 2 == 0))

        summary = {'Total general de ítems': total_items, 'Valor total general': format_currency(total_valor)}
        exporter.add_summary(summary)

        fecha = timezone.now().strftime('%Y%m%d_%H%M%S')
        return exporter.get_response(f"reporte_por_area_{fecha}.xlsx")


class ExportarGarantiasVencenExcelView(RateLimitMixin, LoginRequiredMixin, View):
    """Exporta reporte de garantías próximas a vencer"""
    ratelimit_key = 'export'

    def get(self, request, *args, **kwargs):
        perfil = request.user.perfil
        dias = int(request.GET.get('dias', 30))
        hoy = date.today()
        fecha_limite = hoy + timedelta(days=dias)

        if perfil.area:
            items = Item.objects.filter(area=perfil.area, garantia_hasta__gte=hoy, garantia_hasta__lte=fecha_limite)
        else:
            items = Item.objects.filter(garantia_hasta__gte=hoy, garantia_hasta__lte=fecha_limite)

        exporter = ExcelExporter(title="Garantías")
        exporter.add_title(f"GARANTÍAS QUE VENCEN EN {dias} DÍAS",
                          f"Del {hoy.strftime('%d/%m/%Y')} al {fecha_limite.strftime('%d/%m/%Y')}")

        headers = ['Código Interno', 'Serie', 'Nombre', 'Área', 'Fecha Adquisición',
                   'Garantía Hasta', 'Días Restantes', 'Precio', 'Proveedor/Lote']
        exporter.add_headers(headers)

        for idx, item in enumerate(items.select_related('area', 'lote')):
            dias_restantes = (item.garantia_hasta - hoy).days
            proveedor = item.lote.contrato.proveedor.nombre if (item.lote and item.lote.contrato and item.lote.contrato.proveedor) else 'N/A'
            codigo_mostrar = item.codigo_utp if not item.codigo_utp_pendiente else item.codigo_interno
            row = [codigo_mostrar, item.serie, item.nombre, item.area.nombre if item.area else '',
                   format_date(item.fecha_adquisicion), format_date(item.garantia_hasta),
                   f"{dias_restantes} días", format_currency(item.precio), proveedor]
            exporter.add_row(row, alternate=(idx % 2 == 0))

        valor_total = items.aggregate(Sum('precio'))['precio__sum'] or 0
        summary = {'Total de ítems': items.count(), 'Valor total en riesgo': format_currency(valor_total)}
        exporter.add_summary(summary)

        fecha = timezone.now().strftime('%Y%m%d_%H%M%S')
        return exporter.get_response(f"garantias_vencen_{dias}dias_{fecha}.xlsx")


class ExportarInventarioPDFView(RateLimitMixin, LoginRequiredMixin, View):
    """Exporta el inventario completo a PDF"""
    ratelimit_key = 'export'

    def get(self, request, *args, **kwargs):
        perfil = request.user.perfil
        area_id = request.GET.get('area')
        tipo_id = request.GET.get('tipo')
        estado = request.GET.get('estado')

        if perfil.area:
            items = Item.objects.filter(area=perfil.area)
        else:
            items = Item.objects.all()

        if area_id:
            items = items.filter(area_id=area_id)
        if tipo_id:
            items = items.filter(tipo_item_id=tipo_id)
        if estado:
            items = items.filter(estado=estado)

        exporter = PDFExporter(title="Inventario", orientation="landscape")
        titulo = "INVENTARIO COMPLETO"
        subtitulo = f"Total de ítems: {items.count()}"
        if perfil.area:
            subtitulo += f" | Área: {perfil.area.nombre}"

        exporter.add_title(titulo, subtitulo)

        headers = ['Código', 'Serie', 'Nombre', 'Área', 'Tipo', 'Estado', 'Precio']
        data = []
        for item in items.select_related('area', 'tipo_item')[:100]:
            codigo_mostrar = item.codigo_utp if not item.codigo_utp_pendiente else item.codigo_interno
            data.append([codigo_mostrar, item.serie[:15], item.nombre[:25], item.area.codigo,
                        item.tipo_item.nombre[:15], item.get_estado_display()[:10], format_currency(item.precio)])

        exporter.add_table(headers, data)

        valor_total = items.aggregate(Sum('precio'))['precio__sum'] or 0
        summary = {'Total de ítems': items.count(), 'Valor total': format_currency(valor_total),
                   'Ítems operativos': items.filter(estado='operativo').count(),
                   'Ítems en mantenimiento': items.filter(estado='en_mantenimiento').count()}
        exporter.add_summary_section(summary)

        fecha = timezone.now().strftime('%Y%m%d_%H%M%S')
        return exporter.get_response(f"inventario_{fecha}.pdf")


class ExportarReportePorAreaPDFView(RateLimitMixin, LoginRequiredMixin, View):
    """Exporta reporte de ítems agrupados por área a PDF"""
    ratelimit_key = 'export'

    def get(self, request, *args, **kwargs):
        perfil = request.user.perfil
        if perfil.area:
            areas = Area.objects.filter(id=perfil.area.id, activo=True)
        else:
            areas = Area.objects.filter(activo=True)

        exporter = PDFExporter(title="Reporte por Área")
        exporter.add_title("REPORTE DE INVENTARIO POR ÁREA", "Distribución de ítems y valores")

        total_items = Item.objects.count()
        total_valor = Item.objects.aggregate(Sum('precio'))['precio__sum'] or 0

        headers = ['Área', 'Cantidad', 'Valor Total', '% Total', 'Operativos', 'Mant.', 'Dañados']
        data = []
        for area in areas:
            items_area = Item.objects.filter(area=area)
            cantidad = items_area.count()
            valor = items_area.aggregate(Sum('precio'))['precio__sum'] or 0
            porcentaje = (valor / total_valor * 100) if total_valor > 0 else 0
            data.append([area.nombre[:20], str(cantidad), format_currency(valor), f"{porcentaje:.1f}%",
                        str(items_area.filter(estado='operativo').count()),
                        str(items_area.filter(estado='en_mantenimiento').count()),
                        str(items_area.filter(estado='danado').count())])

        exporter.add_table(headers, data)

        summary = {'Total general de ítems': total_items, 'Valor total general': format_currency(total_valor)}
        exporter.add_summary_section(summary)

        fecha = timezone.now().strftime('%Y%m%d_%H%M%S')
        return exporter.get_response(f"reporte_por_area_{fecha}.pdf")


class ExportarLeasingExcelView(RateLimitMixin, LoginRequiredMixin, View):
    """Exporta reporte de leasing por vencimiento a Excel"""
    ratelimit_key = 'export'

    def get(self, request):
        from .models import EspecificacionesSistemas

        # Obtener parámetros
        anio = request.GET.get('anio')
        meses_str = request.GET.get('meses', '')
        meses = [int(m) for m in meses_str.split(',') if m.isdigit()]

        # Filtrar items de Sistemas con leasing
        area_sistemas = Area.objects.filter(codigo='sistemas').first()
        items = Item.objects.filter(
            area=area_sistemas,
            es_leasing=True,
            leasing_vencimiento__isnull=False
        ).select_related('colaborador_asignado', 'ambiente', 'tipo_item')

        # Filtrar por año y meses
        if anio:
            items = items.filter(leasing_vencimiento__year=int(anio))
        if meses:
            items = items.filter(leasing_vencimiento__month__in=meses)

        items = items.order_by('leasing_vencimiento')

        # Crear Excel
        exporter = ExcelExporter(title="Reporte de Leasing por Vencimiento")

        meses_nombres = ['', 'Enero', 'Febrero', 'Marzo', 'Abril', 'Mayo', 'Junio',
                         'Julio', 'Agosto', 'Septiembre', 'Octubre', 'Noviembre', 'Diciembre']
        filtros = f"Año: {anio or 'Todos'}"
        if meses:
            filtros += f" | Meses: {', '.join([meses_nombres[m] for m in meses])}"
        exporter.add_title("Reporte de Leasing por Vencimiento", filtros)

        headers = ['Código UTP', 'Nombre', 'Tipo', 'Marca', 'Modelo', 'Procesador', 'RAM', 'Disco', 'Vencimiento Leasing', 'Colaborador', 'Ubicación']
        exporter.add_headers(headers)

        for i, item in enumerate(items):
            specs = getattr(item, 'especificaciones_sistemas', None)
            exporter.add_row([
                item.codigo_utp if not item.codigo_utp_pendiente else item.codigo_interno,
                item.nombre,
                item.tipo_item.nombre if item.tipo_item else '-',
                specs.marca if specs and specs.marca else '-',
                specs.modelo if specs and specs.modelo else '-',
                specs.procesador if specs and specs.procesador else '-',
                specs.ram_display if specs else '-',
                specs.almacenamiento_display if specs else '-',
                format_date(item.leasing_vencimiento),
                item.colaborador_asignado.nombre_completo if item.colaborador_asignado else '-',
                item.ambiente.nombre if item.ambiente else '-'
            ], alternate=(i % 2 == 1))

        exporter.auto_adjust_columns()
        exporter.add_summary({'Total de equipos en leasing': items.count()})

        fecha = timezone.now().strftime('%Y%m%d_%H%M%S')
        return exporter.get_response(f"leasing_vencimiento_{fecha}.xlsx")


class ExportarLeasingPDFView(RateLimitMixin, LoginRequiredMixin, View):
    """Exporta reporte de leasing por vencimiento a PDF"""
    ratelimit_key = 'export'

    def get(self, request):
        from .models import EspecificacionesSistemas

        # Obtener parámetros
        anio = request.GET.get('anio')
        meses_str = request.GET.get('meses', '')
        meses = [int(m) for m in meses_str.split(',') if m.isdigit()]

        # Filtrar items de Sistemas con leasing
        area_sistemas = Area.objects.filter(codigo='sistemas').first()
        items = Item.objects.filter(
            area=area_sistemas,
            es_leasing=True,
            leasing_vencimiento__isnull=False
        ).select_related('colaborador_asignado', 'ambiente', 'tipo_item')

        # Filtrar por año y meses
        if anio:
            items = items.filter(leasing_vencimiento__year=int(anio))
        if meses:
            items = items.filter(leasing_vencimiento__month__in=meses)

        items = items.order_by('leasing_vencimiento')

        # Crear PDF
        exporter = PDFExporter(title="Reporte de Leasing", orientation='landscape')

        meses_nombres = ['', 'Enero', 'Febrero', 'Marzo', 'Abril', 'Mayo', 'Junio',
                         'Julio', 'Agosto', 'Septiembre', 'Octubre', 'Noviembre', 'Diciembre']
        filtros = f"Año: {anio or 'Todos'}"
        if meses:
            filtros += f" | Meses: {', '.join([meses_nombres[m] for m in meses])}"
        exporter.add_title("Reporte de Leasing por Vencimiento", filtros)

        headers = ['Código UTP', 'Nombre', 'Marca/Modelo', 'Procesador', 'RAM', 'Disco', 'Vencimiento', 'Colaborador']
        data = []

        for item in items[:100]:  # Limitar a 100 para PDF
            specs = getattr(item, 'especificaciones_sistemas', None)
            data.append([
                item.codigo_utp if not item.codigo_utp_pendiente else item.codigo_interno,
                item.nombre[:25] + '...' if len(item.nombre) > 25 else item.nombre,
                f"{specs.marca or '-'} / {specs.modelo or '-'}" if specs else '-',
                (specs.procesador[:20] + '...' if specs and specs.procesador and len(specs.procesador) > 20 else specs.procesador) if specs else '-',
                specs.ram_display if specs else '-',
                specs.almacenamiento_display if specs else '-',
                format_date(item.leasing_vencimiento),
                item.colaborador_asignado.nombre_completo[:20] if item.colaborador_asignado else '-'
            ])

        exporter.add_table(headers, data)
        exporter.add_summary_section({'Total de equipos': items.count()})

        fecha = timezone.now().strftime('%Y%m%d_%H%M%S')
        return exporter.get_response(f"leasing_vencimiento_{fecha}.pdf")


class ExportarEspecificacionesExcelView(RateLimitMixin, LoginRequiredMixin, View):
    """Exporta reporte por especificaciones técnicas a Excel"""
    ratelimit_key = 'export'

    def get(self, request):
        from .models import EspecificacionesSistemas

        # Obtener parámetros de filtro
        procesador = request.GET.get('procesador', '')
        ram = request.GET.get('ram', '')
        almacenamiento = request.GET.get('almacenamiento', '')
        tipo_disco = request.GET.get('tipo_disco', '')
        marca = request.GET.get('marca', '')
        modelo = request.GET.get('modelo', '')

        # Filtrar items de Sistemas con especificaciones
        area_sistemas = Area.objects.filter(codigo='sistemas').first()
        items = Item.objects.filter(
            area=area_sistemas,
            especificaciones_sistemas__isnull=False
        ).select_related('especificaciones_sistemas', 'colaborador_asignado', 'ambiente', 'tipo_item')

        # Aplicar filtros
        if procesador:
            items = items.filter(especificaciones_sistemas__procesador=procesador)
        if ram:
            items = items.filter(especificaciones_sistemas__ram_total_gb=int(ram))
        if almacenamiento:
            items = items.filter(especificaciones_sistemas__almacenamiento_gb=int(almacenamiento))
        if tipo_disco:
            items = items.filter(especificaciones_sistemas__almacenamiento_tipo=tipo_disco)
        if marca:
            items = items.filter(especificaciones_sistemas__marca=marca)
        if modelo:
            items = items.filter(especificaciones_sistemas__modelo=modelo)

        items = items.order_by('codigo_interno')

        # Crear Excel
        exporter = ExcelExporter(title="Reporte por Especificaciones Técnicas")

        # Construir descripción de filtros
        filtros_list = []
        if procesador:
            filtros_list.append(f"Procesador: {procesador}")
        if ram:
            filtros_list.append(f"RAM: {ram}GB")
        if almacenamiento:
            filtros_list.append(f"Disco: {almacenamiento}GB")
        if tipo_disco:
            filtros_list.append(f"Tipo: {tipo_disco}")
        if marca:
            filtros_list.append(f"Marca: {marca}")
        if modelo:
            filtros_list.append(f"Modelo: {modelo}")

        filtros_str = ' | '.join(filtros_list) if filtros_list else 'Sin filtros'
        exporter.add_title("Reporte por Especificaciones Técnicas", filtros_str)

        headers = ['Código UTP', 'Nombre', 'Tipo', 'Marca', 'Modelo', 'Procesador', 'Gen.', 'RAM', 'Disco', 'S.O.', 'Estado', 'Colaborador', 'Ubicación']
        exporter.add_headers(headers)

        for i, item in enumerate(items):
            specs = item.especificaciones_sistemas
            exporter.add_row([
                item.codigo_utp if not item.codigo_utp_pendiente else item.codigo_interno,
                item.nombre,
                item.tipo_item.nombre if item.tipo_item else '-',
                specs.marca or '-',
                specs.modelo or '-',
                specs.procesador or '-',
                specs.generacion_procesador or '-',
                specs.ram_display,
                specs.almacenamiento_display,
                specs.sistema_operativo or '-',
                item.get_estado_display(),
                item.colaborador_asignado.nombre_completo if item.colaborador_asignado else '-',
                item.ambiente.nombre if item.ambiente else '-'
            ], alternate=(i % 2 == 1))

        exporter.auto_adjust_columns()
        exporter.add_summary({'Total de equipos': items.count()})

        fecha = timezone.now().strftime('%Y%m%d_%H%M%S')
        return exporter.get_response(f"especificaciones_tecnicas_{fecha}.xlsx")


class ExportarEspecificacionesPDFView(RateLimitMixin, LoginRequiredMixin, View):
    """Exporta reporte por especificaciones técnicas a PDF"""
    ratelimit_key = 'export'

    def get(self, request):
        from .models import EspecificacionesSistemas

        # Obtener parámetros de filtro
        procesador = request.GET.get('procesador', '')
        ram = request.GET.get('ram', '')
        almacenamiento = request.GET.get('almacenamiento', '')
        tipo_disco = request.GET.get('tipo_disco', '')
        marca = request.GET.get('marca', '')
        modelo = request.GET.get('modelo', '')

        # Filtrar items de Sistemas con especificaciones
        area_sistemas = Area.objects.filter(codigo='sistemas').first()
        items = Item.objects.filter(
            area=area_sistemas,
            especificaciones_sistemas__isnull=False
        ).select_related('especificaciones_sistemas', 'colaborador_asignado', 'ambiente', 'tipo_item')

        # Aplicar filtros
        if procesador:
            items = items.filter(especificaciones_sistemas__procesador=procesador)
        if ram:
            items = items.filter(especificaciones_sistemas__ram_total_gb=int(ram))
        if almacenamiento:
            items = items.filter(especificaciones_sistemas__almacenamiento_gb=int(almacenamiento))
        if tipo_disco:
            items = items.filter(especificaciones_sistemas__almacenamiento_tipo=tipo_disco)
        if marca:
            items = items.filter(especificaciones_sistemas__marca=marca)
        if modelo:
            items = items.filter(especificaciones_sistemas__modelo=modelo)

        items = items.order_by('codigo_interno')

        # Crear PDF
        exporter = PDFExporter(title="Reporte Especificaciones", orientation='landscape')

        # Construir descripción de filtros
        filtros_list = []
        if procesador:
            filtros_list.append(f"Procesador: {procesador}")
        if ram:
            filtros_list.append(f"RAM: {ram}GB")
        if almacenamiento:
            filtros_list.append(f"Disco: {almacenamiento}GB")
        if tipo_disco:
            filtros_list.append(f"Tipo: {tipo_disco}")
        if marca:
            filtros_list.append(f"Marca: {marca}")
        if modelo:
            filtros_list.append(f"Modelo: {modelo}")

        filtros_str = ' | '.join(filtros_list) if filtros_list else 'Sin filtros'
        exporter.add_title("Reporte por Especificaciones Técnicas", filtros_str)

        headers = ['Código UTP', 'Nombre', 'Marca/Modelo', 'Procesador', 'RAM', 'Disco', 'Estado', 'Colaborador']
        data = []

        for item in items[:100]:  # Limitar a 100 para PDF
            specs = item.especificaciones_sistemas
            data.append([
                item.codigo_utp if not item.codigo_utp_pendiente else item.codigo_interno,
                item.nombre[:20] + '...' if len(item.nombre) > 20 else item.nombre,
                f"{specs.marca or '-'} / {specs.modelo or '-'}"[:25],
                (specs.procesador[:18] + '...' if specs.procesador and len(specs.procesador) > 18 else specs.procesador) or '-',
                specs.ram_display,
                specs.almacenamiento_display,
                item.get_estado_display(),
                (item.colaborador_asignado.nombre_completo[:18] + '...' if item.colaborador_asignado and len(item.colaborador_asignado.nombre_completo) > 18 else item.colaborador_asignado.nombre_completo) if item.colaborador_asignado else '-'
            ])

        exporter.add_table(headers, data)
        exporter.add_summary_section({'Total de equipos': items.count()})

        fecha = timezone.now().strftime('%Y%m%d_%H%M%S')
        return exporter.get_response(f"especificaciones_tecnicas_{fecha}.pdf")


# ==============================================================================
# VISTAS DE MANTENIMIENTO
# ==============================================================================

class MantenimientoListView(PerfilRequeridoMixin, ListView):
    """Lista de mantenimientos"""
    model = Mantenimiento
    template_name = 'productos/mantenimiento_list.html'
    context_object_name = 'mantenimientos'
    paginate_by = 20

    def get_queryset(self):
        queryset = Mantenimiento.objects.select_related(
            'item', 'item__area', 'item__tipo_item', 'item__ambiente__pabellon__sede',
            'responsable', 'creado_por'
        )
        perfil = getattr(self.request.user, 'perfil', None)

        # Filtrar por área si no es admin
        if perfil and perfil.rol != 'admin' and perfil.area:
            queryset = queryset.filter(item__area=perfil.area)

        # Filtros
        estado = self.request.GET.get('estado')
        tipo = self.request.GET.get('tipo')

        if estado:
            queryset = queryset.filter(estado=estado)
        if tipo:
            queryset = queryset.filter(tipo=tipo)

        return queryset.order_by('-fecha_programada')

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        queryset = self.get_queryset()

        # Estadísticas
        context['total_mantenimientos'] = queryset.count()
        context['pendientes'] = queryset.filter(estado='pendiente').count()
        context['en_proceso'] = queryset.filter(estado='en_proceso').count()
        context['completados'] = queryset.filter(estado='completado').count()

        return context


class MantenimientoDetailView(PerfilRequeridoMixin, DetailView):
    """Detalle de un mantenimiento"""
    model = Mantenimiento
    template_name = 'productos/mantenimiento_detail.html'
    context_object_name = 'mantenimiento'

    def get_queryset(self):
        queryset = super().get_queryset()
        perfil = getattr(self.request.user, 'perfil', None)

        if perfil and perfil.rol != 'admin' and perfil.area:
            queryset = queryset.filter(item__area=perfil.area)

        return queryset


class MantenimientoCreateView(PerfilRequeridoMixin, CreateView):
    """Crear un nuevo mantenimiento"""
    model = Mantenimiento
    form_class = MantenimientoForm
    template_name = 'productos/mantenimiento_form.html'
    success_url = reverse_lazy('productos:mantenimiento-list')

    def get_form_kwargs(self):
        kwargs = super().get_form_kwargs()
        kwargs['user'] = self.request.user
        return kwargs

    def form_valid(self, form):
        mantenimiento = form.save(commit=False)
        mantenimiento.creado_por = self.request.user
        mantenimiento.responsable = self.request.user
        mantenimiento.save()

        messages.success(self.request, f'Mantenimiento programado correctamente para {mantenimiento.item.codigo_interno}.')
        return redirect('productos:mantenimiento-detail', pk=mantenimiento.pk)


class MantenimientoUpdateView(PerfilRequeridoMixin, UpdateView):
    """Editar un mantenimiento"""
    model = Mantenimiento
    form_class = MantenimientoForm
    template_name = 'productos/mantenimiento_form.html'

    def get_queryset(self):
        queryset = super().get_queryset()
        perfil = getattr(self.request.user, 'perfil', None)

        if perfil and perfil.rol != 'admin' and perfil.area:
            queryset = queryset.filter(item__area=perfil.area)

        return queryset

    def get_form_kwargs(self):
        kwargs = super().get_form_kwargs()
        kwargs['user'] = self.request.user
        return kwargs

    def form_valid(self, form):
        mantenimiento = form.save()
        messages.success(self.request, f'Mantenimiento actualizado correctamente.')
        return redirect('productos:mantenimiento-detail', pk=mantenimiento.pk)


class MantenimientoIniciarView(PerfilRequeridoMixin, View):
    """Iniciar un mantenimiento"""

    def post(self, request, pk):
        mantenimiento = get_object_or_404(Mantenimiento, pk=pk)
        perfil = request.user.perfil

        # Verificar permisos
        if perfil.rol != 'admin' and perfil.area and mantenimiento.item.area != perfil.area:
            messages.error(request, 'No tienes permisos para iniciar este mantenimiento.')
            return redirect('productos:mantenimiento-detail', pk=pk)

        if mantenimiento.estado != 'pendiente':
            messages.warning(request, 'Este mantenimiento ya fue iniciado.')
            return redirect('productos:mantenimiento-detail', pk=pk)

        # Cambiar estado del ítem a en_mantenimiento
        mantenimiento.item.estado = 'en_mantenimiento'
        mantenimiento.item.save()

        # Iniciar mantenimiento
        mantenimiento.iniciar(usuario=request.user)

        messages.success(request, f'Mantenimiento iniciado. El ítem {mantenimiento.item.codigo_interno} está ahora en mantenimiento.')
        return redirect('productos:mantenimiento-detail', pk=pk)


class MantenimientoFinalizarView(PerfilRequeridoMixin, View):
    """Finalizar un mantenimiento"""
    template_name = 'productos/mantenimiento_finalizar.html'

    def get(self, request, pk):
        mantenimiento = get_object_or_404(Mantenimiento, pk=pk)
        perfil = request.user.perfil

        # Verificar permisos
        if perfil.rol != 'admin' and perfil.area and mantenimiento.item.area != perfil.area:
            messages.error(request, 'No tienes permisos para finalizar este mantenimiento.')
            return redirect('productos:mantenimiento-detail', pk=pk)

        if mantenimiento.estado not in ['pendiente', 'en_proceso']:
            messages.warning(request, 'Este mantenimiento ya fue finalizado o cancelado.')
            return redirect('productos:mantenimiento-detail', pk=pk)

        form = MantenimientoFinalizarForm()
        return render(request, self.template_name, {
            'mantenimiento': mantenimiento,
            'form': form
        })

    def post(self, request, pk):
        mantenimiento = get_object_or_404(Mantenimiento, pk=pk)
        perfil = request.user.perfil

        # Verificar permisos
        if perfil.rol != 'admin' and perfil.area and mantenimiento.item.area != perfil.area:
            messages.error(request, 'No tienes permisos para finalizar este mantenimiento.')
            return redirect('productos:mantenimiento-detail', pk=pk)

        form = MantenimientoFinalizarForm(request.POST)
        if form.is_valid():
            resultado = form.cleaned_data['resultado']
            trabajo_realizado = form.cleaned_data['trabajo_realizado']
            costo = form.cleaned_data.get('costo')

            mantenimiento.finalizar(resultado, trabajo_realizado, costo)

            messages.success(request, f'Mantenimiento finalizado. Estado del ítem actualizado a {mantenimiento.item.get_estado_display()}.')
            return redirect('productos:mantenimiento-detail', pk=pk)

        return render(request, self.template_name, {
            'mantenimiento': mantenimiento,
            'form': form
        })


class MantenimientoCancelarView(PerfilRequeridoMixin, View):
    """Cancelar un mantenimiento"""

    def post(self, request, pk):
        mantenimiento = get_object_or_404(Mantenimiento, pk=pk)
        perfil = request.user.perfil

        # Verificar permisos
        if perfil.rol not in ['admin', 'supervisor']:
            messages.error(request, 'Solo administradores y supervisores pueden cancelar mantenimientos.')
            return redirect('productos:mantenimiento-detail', pk=pk)

        if perfil.rol != 'admin' and perfil.area and mantenimiento.item.area != perfil.area:
            messages.error(request, 'No tienes permisos para cancelar este mantenimiento.')
            return redirect('productos:mantenimiento-detail', pk=pk)

        if mantenimiento.estado in ['completado', 'cancelado']:
            messages.warning(request, 'Este mantenimiento ya fue completado o cancelado.')
            return redirect('productos:mantenimiento-detail', pk=pk)

        motivo = request.POST.get('motivo', '')
        mantenimiento.cancelar(motivo)

        # Si el ítem estaba en mantenimiento, regresarlo a operativo
        if mantenimiento.item.estado == 'en_mantenimiento':
            mantenimiento.item.estado = 'operativo'
            mantenimiento.item.save()

        messages.success(request, 'Mantenimiento cancelado.')
        return redirect('productos:mantenimiento-detail', pk=pk)


class MantenimientoDeleteView(AdminRequeridoMixin, DeleteView):
    """Eliminar un mantenimiento (solo admin)"""
    model = Mantenimiento
    template_name = 'productos/mantenimiento_confirm_delete.html'
    success_url = reverse_lazy('productos:mantenimiento-list')

    def delete(self, request, *args, **kwargs):
        mantenimiento = self.get_object()
        messages.success(request, f'Mantenimiento eliminado.')
        return super().delete(request, *args, **kwargs)


class MantenimientoLoteView(PerfilRequeridoMixin, View):
    """Crear mantenimientos en lote para múltiples ítems"""
    template_name = 'productos/mantenimiento_lote.html'

    def get(self, request):
        # Si vienen items pre-seleccionados por query params
        items_ids = request.GET.getlist('items')
        
        form = MantenimientoLoteForm(user=request.user)
        
        # Pre-seleccionar items si vienen en query params
        if items_ids:
            form.fields['items'].initial = items_ids
        
        # Obtener items disponibles para contexto
        perfil = request.user.perfil
        if perfil.area and perfil.rol != 'admin':
            items = Item.objects.filter(area=perfil.area)
        else:
            items = Item.objects.all()
        
        return render(request, self.template_name, {
            'form': form,
            'items': items,
            'items_preseleccionados': len(items_ids) if items_ids else 0
        })

    def post(self, request):
        form = MantenimientoLoteForm(request.POST, user=request.user)
        
        if form.is_valid():
            items = form.cleaned_data['items']
            tipo = form.cleaned_data['tipo']
            fecha_programada = form.cleaned_data['fecha_programada']
            descripcion = form.cleaned_data.get('descripcion_problema', '')
            tecnico = form.cleaned_data.get('tecnico_asignado', '')
            proveedor = form.cleaned_data.get('proveedor_servicio', '')
            costo = form.cleaned_data.get('costo_estimado')
            proximo = form.cleaned_data.get('proximo_mantenimiento')
            observaciones = form.cleaned_data.get('observaciones', '')
            
            # Crear mantenimientos para cada ítem seleccionado
            mantenimientos_creados = []
            for item in items:
                mantenimiento = Mantenimiento.objects.create(
                    item=item,
                    tipo=tipo,
                    fecha_programada=fecha_programada,
                    descripcion_problema=descripcion,
                    tecnico_asignado=tecnico,
                    proveedor_servicio=proveedor,
                    costo=costo,
                    proximo_mantenimiento=proximo,
                    observaciones=observaciones,
                    responsable=request.user,
                    creado_por=request.user
                )
                mantenimientos_creados.append(mantenimiento)
            
            messages.success(
                request,
                f'Se programaron {len(mantenimientos_creados)} mantenimientos correctamente.'
            )
            return redirect('productos:mantenimiento-list')
        
        # Si hay errores, volver a mostrar el formulario
        perfil = request.user.perfil
        if perfil.area and perfil.rol != 'admin':
            items = Item.objects.filter(area=perfil.area)
        else:
            items = Item.objects.all()
        
        return render(request, self.template_name, {
            'form': form,
            'items': items
        })


# ==============================================================================
# VISTAS DE GARANTÍAS
# ==============================================================================

class GarantiaListView(PerfilRequeridoMixin, ListView):
    """Lista de equipos con información de garantía."""
    model = Item
    template_name = 'productos/garantia_list.html'
    context_object_name = 'items'
    paginate_by = 20

    def get_queryset(self):
        queryset = Item.objects.select_related(
            'area', 'tipo_item', 'ambiente', 'ambiente__pabellon__sede__campus',
            'lote', 'lote__contrato', 'lote__contrato__proveedor'
        ).filter(
            garantia_hasta__isnull=False
        )

        perfil = getattr(self.request.user, 'perfil', None)

        # Filtrar por área si no es admin
        if perfil and perfil.rol != 'admin' and perfil.area:
            queryset = queryset.filter(area=perfil.area)

        # Filtros
        estado_garantia = self.request.GET.get('estado_garantia')
        area = self.request.GET.get('area')
        q = self.request.GET.get('q', '').strip()

        if estado_garantia == 'vigente':
            queryset = queryset.filter(garantia_hasta__gte=timezone.now().date())
        elif estado_garantia == 'vencida':
            queryset = queryset.filter(garantia_hasta__lt=timezone.now().date())

        if area:
            queryset = queryset.filter(area__codigo=area)

        if q:
            queryset = queryset.filter(
                Q(codigo_interno__icontains=q) |
                Q(codigo_utp__icontains=q) |
                Q(serie__icontains=q) |
                Q(nombre__icontains=q)
            )

        return queryset.order_by('garantia_hasta')

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        hoy = timezone.now().date()
        queryset_base = self.get_queryset()

        # Estadísticas
        context['total_con_garantia'] = queryset_base.count()
        context['garantias_vigentes'] = queryset_base.filter(garantia_hasta__gte=hoy).count()
        context['garantias_vencidas'] = queryset_base.filter(garantia_hasta__lt=hoy).count()
        context['por_vencer_30'] = queryset_base.filter(
            garantia_hasta__gte=hoy,
            garantia_hasta__lte=hoy + timedelta(days=30)
        ).count()

        # Para filtros
        context['areas'] = Area.objects.filter(activo=True)
        context['filtros_activos'] = {
            'q': self.request.GET.get('q', ''),
            'estado_garantia': self.request.GET.get('estado_garantia', ''),
            'area': self.request.GET.get('area', ''),
        }

        return context


class GarantiaPorVencerView(PerfilRequeridoMixin, ListView):
    """Lista de equipos con garantía próxima a vencer."""
    model = Item
    template_name = 'productos/garantia_por_vencer.html'
    context_object_name = 'items'
    paginate_by = 20

    def get_queryset(self):
        hoy = timezone.now().date()
        dias = int(self.request.GET.get('dias', 90))

        queryset = Item.objects.select_related(
            'area', 'tipo_item', 'ambiente', 'ambiente__pabellon__sede__campus',
            'lote', 'lote__contrato', 'lote__contrato__proveedor'
        ).filter(
            garantia_hasta__isnull=False,
            garantia_hasta__gte=hoy,
            garantia_hasta__lte=hoy + timedelta(days=dias)
        )

        perfil = getattr(self.request.user, 'perfil', None)
        if perfil and perfil.rol != 'admin' and perfil.area:
            queryset = queryset.filter(area=perfil.area)

        return queryset.order_by('garantia_hasta')

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        hoy = timezone.now().date()
        dias = int(self.request.GET.get('dias', 90))

        context['dias_filtro'] = dias
        context['total_items'] = self.get_queryset().count()

        # Agrupar por urgencia
        queryset = self.get_queryset()
        context['criticos'] = queryset.filter(garantia_hasta__lte=hoy + timedelta(days=30)).count()
        context['urgentes'] = queryset.filter(
            garantia_hasta__gt=hoy + timedelta(days=30),
            garantia_hasta__lte=hoy + timedelta(days=60)
        ).count()
        context['proximos'] = queryset.filter(garantia_hasta__gt=hoy + timedelta(days=60)).count()

        return context


class GarantiaEnProcesoView(PerfilRequeridoMixin, ListView):
    """Lista de equipos actualmente en proceso de garantía."""
    model = Item
    template_name = 'productos/garantia_en_proceso.html'
    context_object_name = 'items'
    paginate_by = 20

    def get_queryset(self):
        queryset = Item.objects.select_related(
            'area', 'tipo_item', 'ambiente', 'ambiente__pabellon__sede__campus',
            'lote', 'lote__contrato', 'lote__contrato__proveedor'
        ).filter(
            estado='garantia'
        )

        perfil = getattr(self.request.user, 'perfil', None)
        if perfil and perfil.rol != 'admin' and perfil.area:
            queryset = queryset.filter(area=perfil.area)

        return queryset.order_by('-modificado_en')

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['total_en_proceso'] = self.get_queryset().count()

        # Obtener los movimientos de garantía asociados
        items_ids = self.get_queryset().values_list('id', flat=True)
        context['movimientos_garantia'] = Movimiento.objects.filter(
            item__in=items_ids,
            tipo='garantia',
            estado__in=['aprobado', 'en_ejecucion', 'en_transito']
        ).select_related('item', 'solicitado_por')

        return context


# Vistas CRUD para Registro de Garantías
from .models import GarantiaRegistro
from .forms_legacy import GarantiaRegistroForm, GarantiaEnviarForm, GarantiaRecibirForm


class GarantiaRegistroListView(PerfilRequeridoMixin, ListView):
    """Lista de registros de garantía."""
    model = GarantiaRegistro
    template_name = 'productos/garantia_registro_list.html'
    context_object_name = 'registros'
    paginate_by = 20

    def get_queryset(self):
        queryset = GarantiaRegistro.objects.select_related(
            'item', 'item__area', 'proveedor', 'creado_por'
        )

        perfil = getattr(self.request.user, 'perfil', None)
        if perfil and perfil.rol != 'admin' and perfil.area:
            queryset = queryset.filter(item__area=perfil.area)

        # Filtros
        estado = self.request.GET.get('estado')
        if estado:
            queryset = queryset.filter(estado=estado)

        return queryset.order_by('-creado_en')

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        queryset = self.get_queryset()
        context['total_registros'] = queryset.count()
        context['pendientes'] = queryset.filter(estado='pendiente').count()
        context['enviados'] = queryset.filter(estado='enviado').count()
        context['en_revision'] = queryset.filter(estado='en_revision').count()
        return context


class GarantiaRegistroCreateView(PerfilRequeridoMixin, CreateView):
    """Crear un nuevo registro de garantía."""
    model = GarantiaRegistro
    form_class = GarantiaRegistroForm
    template_name = 'productos/garantia_registro_form.html'
    success_url = '/productos/garantias/registros/'

    def get_form_kwargs(self):
        kwargs = super().get_form_kwargs()
        kwargs['user'] = self.request.user
        return kwargs

    def form_valid(self, form):
        form.instance.creado_por = self.request.user
        messages.success(self.request, 'Registro de garantía creado correctamente.')
        return super().form_valid(form)


class GarantiaRegistroDetailView(PerfilRequeridoMixin, DetailView):
    """Detalle de un registro de garantía."""
    model = GarantiaRegistro
    template_name = 'productos/garantia_registro_detail.html'
    context_object_name = 'registro'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['form_enviar'] = GarantiaEnviarForm()
        context['form_recibir'] = GarantiaRecibirForm()
        return context


class GarantiaRegistroEnviarView(PerfilRequeridoMixin, View):
    """Marcar un registro de garantía como enviado."""

    def post(self, request, pk):
        registro = get_object_or_404(GarantiaRegistro, pk=pk)
        form = GarantiaEnviarForm(request.POST)

        if form.is_valid():
            registro.fecha_envio = form.cleaned_data['fecha_envio']
            if form.cleaned_data.get('numero_caso'):
                registro.numero_caso = form.cleaned_data['numero_caso']
            if form.cleaned_data.get('observaciones'):
                registro.observaciones += f"\n{form.cleaned_data['observaciones']}"
            registro.enviar()
            messages.success(request, f'Equipo {registro.item.codigo_interno} marcado como enviado a garantía.')
        else:
            messages.error(request, 'Error al procesar el envío.')

        return redirect('productos:garantia-registro-detail', pk=pk)


class GarantiaRegistroRecibirView(PerfilRequeridoMixin, View):
    """Registrar la recepción de un equipo de garantía."""

    def post(self, request, pk):
        registro = get_object_or_404(GarantiaRegistro, pk=pk)
        form = GarantiaRecibirForm(request.POST)

        if form.is_valid():
            registro.recibir(
                diagnostico=form.cleaned_data['diagnostico_proveedor'],
                solucion=form.cleaned_data['solucion_aplicada'],
                resultado=form.cleaned_data['resultado'],
                fecha_recepcion=form.cleaned_data['fecha_recepcion']
            )
            messages.success(request, f'Recepción del equipo {registro.item.codigo_interno} registrada correctamente.')
        else:
            messages.error(request, 'Error al procesar la recepción.')

        return redirect('productos:garantia-registro-detail', pk=pk)


class GarantiaRegistroCancelarView(PerfilRequeridoMixin, View):
    """Cancelar un registro de garantía."""

    def post(self, request, pk):
        registro = get_object_or_404(GarantiaRegistro, pk=pk)
        motivo = request.POST.get('motivo', '')
        registro.cancelar(motivo)
        messages.success(request, f'Registro de garantía cancelado.')
        return redirect('productos:garantia-registro-list')


# ==============================================================================
# SISTEMA DE ACTAS DE ENTREGA/DEVOLUCIÓN
# ==============================================================================

from .models import Gerencia, Colaborador, SoftwareEstandar, ActaEntrega, ActaItem, ActaFoto, ActaSoftware
from .forms import (
    GerenciaForm, ColaboradorForm, SoftwareEstandarForm, ActaEntregaForm,
    ActaItemFormSet, ActaFotoFormSet, FirmaForm, SeleccionarSoftwareForm
)


class GerenciaListView(PerfilRequeridoMixin, ListView):
    """Lista de gerencias."""
    model = Gerencia
    template_name = 'productos/gerencia_list.html'
    context_object_name = 'gerencias'

    def get_queryset(self):
        return Gerencia.objects.all().order_by('nombre')


class GerenciaCreateView(PerfilRequeridoMixin, CreateView):
    """Crear nueva gerencia."""
    model = Gerencia
    form_class = GerenciaForm
    template_name = 'productos/gerencia_form.html'
    success_url = '/productos/gerencias/'

    def form_valid(self, form):
        messages.success(self.request, 'Gerencia creada correctamente.')
        return super().form_valid(form)


class GerenciaUpdateView(PerfilRequeridoMixin, UpdateView):
    """Editar gerencia."""
    model = Gerencia
    form_class = GerenciaForm
    template_name = 'productos/gerencia_form.html'
    success_url = '/productos/gerencias/'

    def form_valid(self, form):
        messages.success(self.request, 'Gerencia actualizada correctamente.')
        return super().form_valid(form)


class ColaboradorListView(PerfilRequeridoMixin, ListView):
    """Lista de colaboradores."""
    model = Colaborador
    template_name = 'productos/colaborador_list.html'
    context_object_name = 'colaboradores'
    paginate_by = 20

    def get_queryset(self):
        queryset = Colaborador.objects.select_related('gerencia', 'sede').all()

        # Filtros
        buscar = self.request.GET.get('buscar', '')
        gerencia = self.request.GET.get('gerencia', '')
        sede = self.request.GET.get('sede', '')
        activo = self.request.GET.get('activo', '')

        if buscar:
            queryset = queryset.filter(
                models.Q(nombre_completo__icontains=buscar) |
                models.Q(dni__icontains=buscar) |
                models.Q(correo__icontains=buscar)
            )

        if gerencia:
            queryset = queryset.filter(gerencia_id=gerencia)

        if sede:
            queryset = queryset.filter(sede_id=sede)

        if activo == '1':
            queryset = queryset.filter(activo=True)
        elif activo == '0':
            queryset = queryset.filter(activo=False)

        return queryset.order_by('nombre_completo')

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['gerencias'] = Gerencia.objects.filter(activo=True)
        context['sedes'] = Sede.objects.filter(activo=True)
        context['filtros'] = {
            'buscar': self.request.GET.get('buscar', ''),
            'gerencia': self.request.GET.get('gerencia', ''),
            'sede': self.request.GET.get('sede', ''),
            'activo': self.request.GET.get('activo', ''),
        }
        return context


class ColaboradorCreateView(PerfilRequeridoMixin, CreateView):
    """Crear nuevo colaborador."""
    model = Colaborador
    form_class = ColaboradorForm
    template_name = 'productos/colaborador_form.html'
    success_url = '/productos/colaboradores/'

    def form_valid(self, form):
        form.instance.creado_por = self.request.user
        messages.success(self.request, 'Colaborador creado correctamente.')
        return super().form_valid(form)


class ColaboradorUpdateView(PerfilRequeridoMixin, UpdateView):
    """Editar colaborador."""
    model = Colaborador
    form_class = ColaboradorForm
    template_name = 'productos/colaborador_form.html'
    success_url = '/productos/colaboradores/'

    def form_valid(self, form):
        messages.success(self.request, 'Colaborador actualizado correctamente.')
        return super().form_valid(form)


class ColaboradorDetailView(PerfilRequeridoMixin, DetailView):
    """Detalle de colaborador con sus ítems asignados."""
    model = Colaborador
    template_name = 'productos/colaborador_detail.html'
    context_object_name = 'colaborador'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['items_asignados'] = Item.objects.filter(
            colaborador_asignado=self.object
        ).select_related('tipo_item', 'area')
        context['actas'] = ActaEntrega.objects.filter(
            colaborador=self.object
        ).order_by('-fecha')[:10]
        return context


class BuscarColaboradorView(PerfilRequeridoMixin, View):
    """API para buscar colaborador por DNI (AJAX)."""

    def get(self, request):
        dni = request.GET.get('dni', '').strip()

        if not dni:
            return JsonResponse({'error': 'DNI requerido'}, status=400)

        try:
            colaborador = Colaborador.objects.select_related('gerencia', 'sede').get(dni=dni)
            return JsonResponse({
                'id': colaborador.id,
                'dni': colaborador.dni,
                'nombre_completo': colaborador.nombre_completo,
                'cargo': colaborador.cargo,
                'gerencia': colaborador.gerencia.nombre,
                'sede': str(colaborador.sede),
                'anexo': colaborador.anexo,
                'correo': colaborador.correo,
                'items_asignados': colaborador.cantidad_items_asignados,
            })
        except Colaborador.DoesNotExist:
            return JsonResponse({'error': 'Colaborador no encontrado'}, status=404)


class EspecificacionesValoresView(LoginRequiredMixin, View):
    """API para obtener valores únicos de especificaciones técnicas (para autocompletado)."""

    def get(self, request):
        from django.db.models import Count

        # Obtener valores únicos de cada campo, ordenados por frecuencia de uso
        marcas = list(
            EspecificacionesSistemas.objects.exclude(marca='').exclude(marca__isnull=True)
            .values_list('marca', flat=True).annotate(count=Count('marca'))
            .order_by('-count').distinct()[:50]
        )

        modelos = list(
            EspecificacionesSistemas.objects.exclude(modelo='').exclude(modelo__isnull=True)
            .values_list('modelo', flat=True).annotate(count=Count('modelo'))
            .order_by('-count').distinct()[:100]
        )

        procesadores = list(
            EspecificacionesSistemas.objects.exclude(procesador='').exclude(procesador__isnull=True)
            .values_list('procesador', flat=True).annotate(count=Count('procesador'))
            .order_by('-count').distinct()[:100]
        )

        sistemas_operativos = list(
            EspecificacionesSistemas.objects.exclude(sistema_operativo='').exclude(sistema_operativo__isnull=True)
            .values_list('sistema_operativo', flat=True).annotate(count=Count('sistema_operativo'))
            .order_by('-count').distinct()[:30]
        )

        # Modelos filtrados por marca (si se especifica)
        marca_filter = request.GET.get('marca', '')
        if marca_filter:
            modelos = list(
                EspecificacionesSistemas.objects.filter(marca__iexact=marca_filter)
                .exclude(modelo='').exclude(modelo__isnull=True)
                .values_list('modelo', flat=True).distinct()[:50]
            )

        return JsonResponse({
            'marcas': marcas,
            'modelos': modelos,
            'procesadores': procesadores,
            'sistemas_operativos': sistemas_operativos,
        })


class ModelosEquipoPorMarcaView(LoginRequiredMixin, View):
    """API para obtener modelos de equipo filtrados por marca."""

    def get(self, request):
        from .models import ModeloEquipo
        marca_id = request.GET.get('marca_id')
        if marca_id:
            modelos = ModeloEquipo.objects.filter(
                marca_id=marca_id, activo=True
            ).values('id', 'nombre').order_by('nombre')
            return JsonResponse(list(modelos), safe=False)
        return JsonResponse([], safe=False)


class ActaListView(PerfilRequeridoMixin, CampusFilterMixin, ListView):
    """Lista de actas de entrega/devolución según permisos de campus."""
    model = ActaEntrega
    template_name = 'productos/acta_list.html'
    context_object_name = 'actas'
    paginate_by = 20

    def get_queryset(self):
        queryset = ActaEntrega.objects.select_related(
            'colaborador', 'colaborador__gerencia', 'colaborador__sede',
            'creado_por'
        ).prefetch_related(
            'items__item',
            'items__item__tipo_item',
            'software__software'
        )

        # Filtrar por campus - solo actas que contengan items del campus del usuario
        perfil = getattr(self.request.user, 'perfil', None)
        if perfil and perfil.rol != 'admin':
            campus_ids = list(self.get_campus_permitidos().values_list('id', flat=True))
            # Obtener actas que tienen al menos un item en los campus permitidos
            actas_con_items_permitidos = ActaItem.objects.filter(
                item__ambiente__pabellon__sede__campus_id__in=campus_ids
            ).values_list('acta_id', flat=True).distinct()
            queryset = queryset.filter(id__in=actas_con_items_permitidos)

        # Filtros
        tipo = self.request.GET.get('tipo', '')
        buscar = self.request.GET.get('buscar', '')
        fecha_desde = self.request.GET.get('fecha_desde', '')
        fecha_hasta = self.request.GET.get('fecha_hasta', '')

        if tipo:
            queryset = queryset.filter(tipo=tipo)

        if buscar:
            queryset = queryset.filter(
                models.Q(numero_acta__icontains=buscar) |
                models.Q(colaborador__nombre_completo__icontains=buscar) |
                models.Q(colaborador__dni__icontains=buscar)
            )

        if fecha_desde:
            queryset = queryset.filter(fecha__date__gte=fecha_desde)

        if fecha_hasta:
            queryset = queryset.filter(fecha__date__lte=fecha_hasta)

        return queryset.order_by('-fecha')

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['filtros'] = {
            'tipo': self.request.GET.get('tipo', ''),
            'buscar': self.request.GET.get('buscar', ''),
            'fecha_desde': self.request.GET.get('fecha_desde', ''),
            'fecha_hasta': self.request.GET.get('fecha_hasta', ''),
        }
        return context


class ActaDetailView(PerfilRequeridoMixin, DetailView):
    """Detalle de un acta."""
    model = ActaEntrega
    template_name = 'productos/acta_detail.html'
    context_object_name = 'acta'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['acta_items'] = self.object.items.select_related(
            'item', 'item__tipo_item'
        ).all()
        context['acta_software'] = self.object.software.select_related('software').all()
        context['acta_fotos'] = self.object.fotos.all()
        return context


class ActaCreateView(PerfilRequeridoMixin, View):
    """Crear nueva acta de entrega/devolución - Wizard de múltiples pasos."""
    template_name = 'productos/acta_create.html'

    def get(self, request):
        # Verificar si viene desde un movimiento
        movimiento_id = request.GET.get('movimiento')
        movimiento = None
        items_movimiento = []

        if movimiento_id:
            try:
                movimiento = Movimiento.objects.select_related(
                    'colaborador_nuevo', 'item'
                ).get(pk=movimiento_id)

                # Validar que el movimiento requiere acta
                if movimiento.tipo not in ['asignacion', 'prestamo']:
                    messages.warning(
                        request,
                        'Este tipo de movimiento no requiere acta de entrega.'
                    )
                    return redirect('productos:movimiento-detail', pk=movimiento_id)

                # Validar que no tenga acta ya
                if hasattr(movimiento, 'acta_entrega') and movimiento.acta_entrega:
                    messages.info(request, 'Este movimiento ya tiene un acta asociada.')
                    return redirect('productos:acta-detail', pk=movimiento.acta_entrega.pk)

                # Guardar en sesión
                request.session['acta_movimiento_id'] = movimiento.id

                # Obtener items del movimiento
                items_movimiento_qs = movimiento.items_movimiento.select_related('item')
                if items_movimiento_qs.exists():
                    items_movimiento = [mi.item for mi in items_movimiento_qs]
                elif movimiento.item:
                    items_movimiento = [movimiento.item]

                # Precargar datos del form
                initial_data = {
                    'tipo': 'entrega',
                    'colaborador': movimiento.colaborador_nuevo,
                    'ticket': '',
                    'observaciones': f'Generado desde movimiento {movimiento.pk}',
                }

                form = ActaEntregaForm(user=request.user, initial=initial_data)
                software_form = SeleccionarSoftwareForm()

                # Si tiene colaborador destino, ir directo al paso 2
                if movimiento.colaborador_nuevo and items_movimiento:
                    # Guardar datos en sesión
                    request.session['acta_tipo'] = 'entrega'
                    request.session['acta_colaborador_id'] = movimiento.colaborador_nuevo.id
                    request.session['acta_ticket'] = ''
                    request.session['acta_observaciones'] = f'Generado desde movimiento {movimiento.pk}'

                    return render(request, self.template_name, {
                        'form': form,
                        'software_form': software_form,
                        'items_disponibles': items_movimiento,
                        'items_preseleccionados': [item.id for item in items_movimiento],
                        'colaborador': movimiento.colaborador_nuevo,
                        'tipo': 'entrega',
                        'movimiento': movimiento,
                        'paso': 2,
                    })

            except Movimiento.DoesNotExist:
                messages.error(request, 'Movimiento no encontrado.')
                return redirect('productos:movimiento-list')

        # Paso inicial normal: seleccionar tipo y colaborador
        form = ActaEntregaForm(user=request.user)
        software_form = SeleccionarSoftwareForm()

        return render(request, self.template_name, {
            'form': form,
            'software_form': software_form,
            'movimiento': movimiento,
            'paso': 1,
        })

    def post(self, request):
        paso = request.POST.get('paso', '1')

        if paso == '1':
            # Validar tipo y colaborador, mostrar selección de ítems
            form = ActaEntregaForm(request.POST, user=request.user)

            if form.is_valid():
                tipo = form.cleaned_data['tipo']
                colaborador = form.cleaned_data['colaborador']

                # Guardar datos en sesión
                request.session['acta_tipo'] = tipo
                request.session['acta_colaborador_id'] = colaborador.id
                request.session['acta_ticket'] = form.cleaned_data.get('ticket', '')
                request.session['acta_observaciones'] = form.cleaned_data.get('observaciones', '')

                software_form = SeleccionarSoftwareForm()

                # Ya no cargamos items aquí - se buscan por AJAX
                return render(request, self.template_name, {
                    'form': form,
                    'software_form': software_form,
                    'colaborador': colaborador,
                    'tipo': tipo,
                    'paso': 2,
                })

            software_form = SeleccionarSoftwareForm()
            return render(request, self.template_name, {
                'form': form,
                'software_form': software_form,
                'paso': 1,
            })

        elif paso == '2':
            # Validar ítems seleccionados, mostrar formulario de accesorios
            items_ids = request.POST.getlist('items')
            software_ids = request.POST.getlist('software')

            if not items_ids:
                messages.error(request, 'Debe seleccionar al menos un ítem.')
                return redirect('productos:acta-create')

            # Guardar en sesión
            request.session['acta_items_ids'] = items_ids
            request.session['acta_software_ids'] = software_ids

            items = Item.objects.filter(id__in=items_ids).select_related('tipo_item')
            colaborador_id = request.session.get('acta_colaborador_id')
            colaborador = Colaborador.objects.get(id=colaborador_id)

            return render(request, self.template_name, {
                'items': items,
                'colaborador': colaborador,
                'tipo': request.session.get('acta_tipo'),
                'paso': 3,
            })

        elif paso == '3':
            # Validar accesorios y firmas, crear el acta
            import base64
            from django.core.files.base import ContentFile

            firma_receptor_data = request.POST.get('firma_receptor')
            firma_emisor_data = request.POST.get('firma_emisor')

            if not firma_receptor_data or not firma_emisor_data:
                messages.error(request, 'Ambas firmas son obligatorias.')
                return redirect('productos:acta-create')

            # Recuperar datos de sesión
            tipo = request.session.get('acta_tipo')
            colaborador_id = request.session.get('acta_colaborador_id')
            ticket = request.session.get('acta_ticket', '')
            observaciones = request.session.get('acta_observaciones', '')
            items_ids = request.session.get('acta_items_ids', [])
            software_ids = request.session.get('acta_software_ids', [])
            movimiento_id = request.session.get('acta_movimiento_id')

            # Validar datos de sesión
            if not tipo or not colaborador_id:
                messages.error(request, 'Datos de sesión incompletos. Por favor, inicie el proceso nuevamente.')
                return redirect('productos:acta-create')

            if not items_ids:
                messages.error(request, 'No hay ítems seleccionados. Por favor, seleccione al menos un ítem.')
                return redirect('productos:acta-create')

            try:
                colaborador = Colaborador.objects.get(id=colaborador_id)
            except Colaborador.DoesNotExist:
                messages.error(request, 'Colaborador no encontrado. Por favor, inicie el proceso nuevamente.')
                return redirect('productos:acta-create')

            # Obtener movimiento si existe
            movimiento = None
            if movimiento_id:
                try:
                    movimiento = Movimiento.objects.get(pk=movimiento_id)
                except Movimiento.DoesNotExist:
                    pass

            # Procesar firmas (base64 a imagen)
            def base64_to_image(data, filename):
                if ',' in data:
                    data = data.split(',')[1]
                image_data = base64.b64decode(data)
                return ContentFile(image_data, name=filename)

            firma_receptor_file = base64_to_image(
                firma_receptor_data,
                f'firma_receptor_{colaborador.dni}.png'
            )
            firma_emisor_file = base64_to_image(
                firma_emisor_data,
                f'firma_emisor_{request.user.username}.png'
            )

            # Crear el acta y sus relaciones en una transacción atómica
            try:
                with transaction.atomic():
                    acta = ActaEntrega.objects.create(
                        tipo=tipo,
                        colaborador=colaborador,
                        ticket=ticket,
                        observaciones=observaciones,
                        firma_receptor=firma_receptor_file,
                        firma_emisor=firma_emisor_file,
                        creado_por=request.user,
                        movimiento=movimiento  # Vincula con movimiento (None si no hay)
                    )

                    # Crear ActaItems con accesorios
                    items = Item.objects.filter(id__in=items_ids)
                    for item in items:
                        ActaItem.objects.create(
                            acta=acta,
                            item=item,
                            acc_cargador=request.POST.get(f'acc_cargador_{item.id}') == 'on',
                            acc_cable_seguridad=request.POST.get(f'acc_cable_seguridad_{item.id}') == 'on',
                            acc_bateria=request.POST.get(f'acc_bateria_{item.id}') == 'on',
                            acc_maletin=request.POST.get(f'acc_maletin_{item.id}') == 'on',
                            acc_cable_red=request.POST.get(f'acc_cable_red_{item.id}') == 'on',
                            acc_teclado_mouse=request.POST.get(f'acc_teclado_mouse_{item.id}') == 'on',
                        )

                        # Actualizar asignación del ítem
                        if tipo == 'entrega':
                            item.colaborador_asignado = colaborador
                        else:
                            item.colaborador_asignado = None
                        item.save()

                    # Crear ActaSoftware
                    for software_id in software_ids:
                        ActaSoftware.objects.create(
                            acta=acta,
                            software_id=software_id
                        )

                    # Procesar fotos si hay
                    fotos = request.FILES.getlist('fotos')
                    for foto in fotos:
                        ActaFoto.objects.create(
                            acta=acta,
                            foto=foto
                        )

            except Exception as e:
                messages.error(request, f'Error al crear el acta: {str(e)}')
                return redirect('productos:acta-create')

            # Limpiar sesión (fuera de la transacción)
            for key in ['acta_tipo', 'acta_colaborador_id', 'acta_ticket',
                       'acta_observaciones', 'acta_items_ids', 'acta_software_ids',
                       'acta_movimiento_id']:
                request.session.pop(key, None)

            # Mensaje de éxito (diferente si viene de un movimiento)
            if movimiento:
                messages.success(
                    request,
                    f'Acta {acta.numero_acta} creada correctamente. '
                    f'El movimiento {movimiento.pk} ha sido ejecutado automáticamente.'
                )
            else:
                messages.success(
                    request,
                    f'Acta {acta.numero_acta} creada correctamente.'
                )

            return redirect('productos:acta-detail', pk=acta.pk)

        return redirect('productos:acta-create')


class ActaDescargarPDFView(PerfilRequeridoMixin, View):
    """Descargar PDF del acta."""

    def _generar_nombre_archivo(self, acta):
        """Genera nombre descriptivo para el archivo PDF."""
        import re
        # Limpiar nombre del colaborador
        nombre_colaborador = acta.colaborador.nombre_completo
        nombre_limpio = nombre_colaborador.replace(' ', '_')
        nombre_limpio = re.sub(r'[^\w\-_]', '', nombre_limpio)[:30]

        # Generar nombre según tipo de acta
        if acta.tipo == 'entrega':
            return f"asignacion_a_{nombre_limpio}.pdf"
        else:
            return f"devolucion_de_{nombre_limpio}.pdf"

    def get(self, request, pk):
        acta = get_object_or_404(ActaEntrega, pk=pk)
        filename = self._generar_nombre_archivo(acta)

        # Si ya tiene PDF generado, devolverlo
        if acta.pdf_archivo:
            response = HttpResponse(acta.pdf_archivo, content_type='application/pdf')
            response['Content-Disposition'] = f'attachment; filename="{filename}"'
            return response

        # Si no, generar el PDF
        from .utils.acta_pdf import generar_acta_pdf

        try:
            pdf_buffer = generar_acta_pdf(acta)
            pdf_bytes = pdf_buffer.getvalue()

            # Guardar el PDF en el modelo
            from django.core.files.base import ContentFile
            acta.pdf_archivo.save(
                f'{acta.numero_acta}.pdf',
                ContentFile(pdf_bytes),
                save=True
            )

            response = HttpResponse(pdf_bytes, content_type='application/pdf')
            response['Content-Disposition'] = f'attachment; filename="{filename}"'
            return response

        except Exception as e:
            messages.error(request, f'Error al generar PDF: {str(e)}')
            return redirect('productos:acta-detail', pk=pk)


class ActaEnviarCorreoView(PerfilRequeridoMixin, View):
    """Enviar acta por correo al colaborador."""

    def post(self, request, pk):
        acta = get_object_or_404(ActaEntrega, pk=pk)

        from .utils.acta_email import enviar_acta_por_correo
        from .utils.acta_pdf import generar_acta_pdf

        try:
            # Generar PDF si no existe
            if not acta.pdf_archivo:
                from django.core.files.base import ContentFile

                pdf_buffer = generar_acta_pdf(acta)
                pdf_bytes = pdf_buffer.getvalue()
                acta.pdf_archivo.save(
                    f'{acta.numero_acta}.pdf',
                    ContentFile(pdf_bytes),
                    save=True
                )

            # Leer el PDF para enviarlo
            acta.pdf_archivo.seek(0)
            pdf_bytes = acta.pdf_archivo.read()

            # Obtener rutas de fotos adjuntas
            fotos_paths = []
            for foto in acta.fotos.all():
                if foto.foto and hasattr(foto.foto, 'path'):
                    fotos_paths.append(foto.foto.path)

            # Enviar correo
            enviar_acta_por_correo(acta, pdf_bytes, fotos_paths=fotos_paths if fotos_paths else None)

            # Marcar como enviado
            acta.correo_enviado = True
            acta.fecha_envio_correo = timezone.now()
            acta.save()

            messages.success(
                request,
                f'Acta enviada por correo a {acta.colaborador.correo}'
            )

        except Exception as e:
            messages.error(request, f'Error al enviar correo: {str(e)}')

        return redirect('productos:acta-detail', pk=pk)


class SoftwareEstandarListView(PerfilRequeridoMixin, ListView):
    """Lista de software estándar."""
    model = SoftwareEstandar
    template_name = 'productos/software_list.html'
    context_object_name = 'software_list'


class SoftwareEstandarCreateView(PerfilRequeridoMixin, CreateView):
    """Crear nuevo software estándar."""
    model = SoftwareEstandar
    form_class = SoftwareEstandarForm
    template_name = 'productos/software_form.html'
    success_url = '/productos/software/'

    def form_valid(self, form):
        messages.success(self.request, 'Software agregado correctamente.')
        return super().form_valid(form)


class SoftwareEstandarUpdateView(PerfilRequeridoMixin, UpdateView):
    """Editar software estándar."""
    model = SoftwareEstandar
    form_class = SoftwareEstandarForm
    template_name = 'productos/software_form.html'
    success_url = '/productos/software/'

    def form_valid(self, form):
        messages.success(self.request, 'Software actualizado correctamente.')
        return super().form_valid(form)


# ============================================================================
# CATÁLOGOS DE EQUIPOS (Marca, Modelo, Procesador)
# ============================================================================

class CatalogoEquiposView(SupervisorRequeridoMixin, TemplateView):
    """Vista principal de catálogos de equipos."""
    template_name = 'productos/catalogo_equipos.html'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['marcas'] = MarcaEquipo.objects.annotate(
            total_modelos=Count('modelos')
        ).order_by('nombre')
        context['procesadores'] = ProcesadorEquipo.objects.order_by('nombre')
        return context


# --- Marcas ---
class MarcaEquipoCreateView(SupervisorRequeridoMixin, CreateView):
    """Crear nueva marca de equipo."""
    model = MarcaEquipo
    fields = ['nombre', 'activo']
    template_name = 'productos/marca_form.html'
    success_url = reverse_lazy('productos:catalogo-equipos')

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['titulo'] = 'Nueva Marca'
        return context

    def form_valid(self, form):
        messages.success(self.request, 'Marca creada correctamente.')
        return super().form_valid(form)


class MarcaEquipoUpdateView(SupervisorRequeridoMixin, UpdateView):
    """Editar marca de equipo."""
    model = MarcaEquipo
    fields = ['nombre', 'activo']
    template_name = 'productos/marca_form.html'
    success_url = reverse_lazy('productos:catalogo-equipos')

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['titulo'] = 'Editar Marca'
        context['modelos'] = self.object.modelos.order_by('nombre')
        return context

    def form_valid(self, form):
        messages.success(self.request, 'Marca actualizada correctamente.')
        return super().form_valid(form)


# --- Modelos ---
class ModeloEquipoCreateView(SupervisorRequeridoMixin, CreateView):
    """Crear nuevo modelo de equipo."""
    model = ModeloEquipo
    fields = ['marca', 'nombre', 'activo']
    template_name = 'productos/modelo_form.html'
    success_url = reverse_lazy('productos:catalogo-equipos')

    def get_initial(self):
        initial = super().get_initial()
        marca_id = self.request.GET.get('marca')
        if marca_id:
            initial['marca'] = marca_id
        return initial

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['titulo'] = 'Nuevo Modelo'
        return context

    def form_valid(self, form):
        messages.success(self.request, 'Modelo creado correctamente.')
        return super().form_valid(form)


class ModeloEquipoUpdateView(SupervisorRequeridoMixin, UpdateView):
    """Editar modelo de equipo."""
    model = ModeloEquipo
    fields = ['marca', 'nombre', 'activo']
    template_name = 'productos/modelo_form.html'
    success_url = reverse_lazy('productos:catalogo-equipos')

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['titulo'] = 'Editar Modelo'
        return context

    def form_valid(self, form):
        messages.success(self.request, 'Modelo actualizado correctamente.')
        return super().form_valid(form)


# --- Procesadores ---
class ProcesadorEquipoCreateView(SupervisorRequeridoMixin, CreateView):
    """Crear nuevo procesador."""
    model = ProcesadorEquipo
    fields = ['nombre', 'activo']
    template_name = 'productos/procesador_form.html'
    success_url = reverse_lazy('productos:catalogo-equipos')

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['titulo'] = 'Nuevo Procesador'
        return context

    def form_valid(self, form):
        messages.success(self.request, 'Procesador creado correctamente.')
        return super().form_valid(form)


class ProcesadorEquipoUpdateView(SupervisorRequeridoMixin, UpdateView):
    """Editar procesador."""
    model = ProcesadorEquipo
    fields = ['nombre', 'activo']
    template_name = 'productos/procesador_form.html'
    success_url = reverse_lazy('productos:catalogo-equipos')

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['titulo'] = 'Editar Procesador'
        return context

    def form_valid(self, form):
        messages.success(self.request, 'Procesador actualizado correctamente.')
        return super().form_valid(form)


# ============================================================================
# FORMATO DE TRASLADO
# ============================================================================

class FormatoTrasladoMovimientoView(PerfilRequeridoMixin, View):
    """Genera formato de traslado Excel desde un movimiento existente."""

    def get(self, request, pk):
        from .utils.export_utils import generar_formato_traslado

        movimiento = get_object_or_404(Movimiento, pk=pk)

        # Obtener todos los ítems del movimiento (compatible con modelo nuevo y antiguo)
        items = movimiento.get_items()

        # Preparar datos de todos los ítems
        items_data = []
        for item in items:
            item_data = {
                'codigo_utp': item.codigo_utp or item.codigo_interno,
                'descripcion': item.nombre or (item.tipo_item.nombre if item.tipo_item else ''),
                'marca': '',
                'modelo': '',
                'serie': item.serie or '',
                'estado': 'OPTIMO' if item.estado == 'instalado' else item.get_estado_display().upper(),
                'observaciones': '',
            }

            # Si tiene especificaciones de sistemas, obtener marca y modelo
            if hasattr(item, 'especificaciones_sistemas'):
                try:
                    specs = item.especificaciones_sistemas
                    if specs and specs.marca_equipo:
                        item_data['marca'] = specs.marca_equipo.nombre
                    if specs and specs.modelo_equipo:
                        item_data['modelo'] = specs.modelo_equipo.nombre
                except AttributeError:
                    pass

            items_data.append(item_data)

        # Usar el primer ítem para datos de origen (referencia)
        primer_item = items[0] if items else movimiento.item

        # Preparar datos de origen
        origen_data = {
            'sede': '',
            'piso': '',
            'ubicacion': '',
            'usuario': '',
        }
        if movimiento.ambiente_origen:
            amb = movimiento.ambiente_origen
            origen_data['sede'] = f"{amb.pabellon.sede.campus.nombre} - {amb.pabellon.sede.nombre}" if amb.pabellon and amb.pabellon.sede else ''
            origen_data['piso'] = f"Pab. {amb.pabellon.letra}" if amb.pabellon else ''
            origen_data['ubicacion'] = amb.nombre
            if primer_item and primer_item.colaborador_asignado:
                origen_data['usuario'] = primer_item.colaborador_asignado.nombre_completo
            elif primer_item and primer_item.usuario_asignado:
                origen_data['usuario'] = primer_item.usuario_asignado.get_full_name()

        # Preparar datos de destino
        destino_data = {
            'sede': '',
            'piso': '',
            'ubicacion': '',
            'usuario': '',
        }
        if movimiento.ambiente_destino:
            amb = movimiento.ambiente_destino
            destino_data['sede'] = f"{amb.pabellon.sede.campus.nombre} - {amb.pabellon.sede.nombre}" if amb.pabellon and amb.pabellon.sede else ''
            destino_data['piso'] = f"Pab. {amb.pabellon.letra}" if amb.pabellon else ''
            destino_data['ubicacion'] = amb.nombre
            if movimiento.colaborador_nuevo:
                destino_data['usuario'] = movimiento.colaborador_nuevo.nombre_completo

        # Generar Excel
        buffer = generar_formato_traslado(
            items_data=items_data,
            origen_data=origen_data,
            destino_data=destino_data,
            fecha=movimiento.fecha_solicitud
        )

        # Preparar respuesta con nombre descriptivo
        fecha_str = timezone.now().strftime('%Y%m%d')

        # Función para limpiar nombres (reemplazar espacios y caracteres especiales)
        def limpiar_nombre(nombre):
            import re
            nombre = nombre.replace(' ', '_')
            nombre = re.sub(r'[^\w\-_]', '', nombre)
            return nombre[:30]  # Limitar longitud

        # Generar nombre descriptivo: "FT de [origen] a [destino]"
        origen_nombre = movimiento.ambiente_origen.nombre if movimiento.ambiente_origen else 'Almacen'
        destino_nombre = movimiento.ambiente_destino.nombre if movimiento.ambiente_destino else 'Destino'

        origen_limpio = limpiar_nombre(origen_nombre)
        destino_limpio = limpiar_nombre(destino_nombre)

        filename = f"FT_de_{origen_limpio}_a_{destino_limpio}_{fecha_str}.xlsx"

        response = HttpResponse(
            buffer.read(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        return response


class FormatoTrasladoManualView(PerfilRequeridoMixin, TemplateView):
    """Vista para crear formato de traslado manualmente."""
    template_name = 'productos/formato_traslado.html'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['campus_list'] = Campus.objects.filter(activo=True)
        return context


class FormatoTrasladoGenerarView(PerfilRequeridoMixin, View):
    """Genera formato de traslado Excel desde selección manual."""

    def post(self, request):
        from .utils.export_utils import generar_formato_traslado

        # Obtener items seleccionados
        item_ids = request.POST.getlist('items')
        items = Item.objects.filter(pk__in=item_ids).select_related(
            'tipo_item', 'ambiente', 'ambiente__pabellon', 'ambiente__pabellon__sede',
            'ambiente__pabellon__sede__campus', 'colaborador_asignado', 'usuario_asignado'
        )

        items_data = []
        primer_item_con_ubicacion = None

        for item in items:
            item_dict = {
                'codigo_utp': item.codigo_utp or item.codigo_interno,
                'descripcion': item.nombre or (item.tipo_item.nombre if item.tipo_item else ''),
                'marca': '',
                'modelo': '',
            }
            # Si tiene especificaciones de sistemas
            if hasattr(item, 'especificaciones_sistemas'):
                try:
                    specs = item.especificaciones_sistemas
                    if specs and specs.marca_equipo:
                        item_dict['marca'] = specs.marca_equipo.nombre
                    if specs and specs.modelo_equipo:
                        item_dict['modelo'] = specs.modelo_equipo.nombre
                except AttributeError:
                    pass
            items_data.append(item_dict)

            # Guardar el primer item con ubicación para obtener datos de origen
            if not primer_item_con_ubicacion and item.ambiente:
                primer_item_con_ubicacion = item

        # Datos de origen - se obtienen automáticamente del primer item
        origen_data = {'sede': '', 'piso': '', 'ubicacion': '', 'usuario': ''}

        if primer_item_con_ubicacion and primer_item_con_ubicacion.ambiente:
            amb = primer_item_con_ubicacion.ambiente
            origen_data['sede'] = f"{amb.pabellon.sede.campus.nombre} - {amb.pabellon.sede.nombre}"
            origen_data['piso'] = f"Pab. {amb.pabellon.letra}"
            origen_data['ubicacion'] = amb.nombre
            # Si el item tiene usuario asignado, usarlo como usuario de origen
            if primer_item_con_ubicacion.usuario_asignado:
                origen_data['usuario'] = primer_item_con_ubicacion.usuario_asignado.get_full_name()

        # Datos de destino - pueden venir de ambiente, pabellón o solo sede
        ambiente_destino_id = request.POST.get('ambiente_destino')
        pabellon_destino_id = request.POST.get('pabellon_destino')
        sede_destino_id = request.POST.get('sede_destino')
        destino_data = {'sede': '', 'piso': '', 'ubicacion': '', 'usuario': ''}

        if ambiente_destino_id:
            try:
                amb = Ambiente.objects.select_related('pabellon__sede__campus').get(pk=ambiente_destino_id)
                destino_data['sede'] = f"{amb.pabellon.sede.campus.nombre} - {amb.pabellon.sede.nombre}"
                destino_data['piso'] = f"Pab. {amb.pabellon.letra}"
                destino_data['ubicacion'] = amb.nombre
            except Ambiente.DoesNotExist:
                pass
        elif pabellon_destino_id:
            try:
                pab = Pabellon.objects.select_related('sede__campus').get(pk=pabellon_destino_id)
                destino_data['sede'] = f"{pab.sede.campus.nombre} - {pab.sede.nombre}"
                destino_data['piso'] = f"Pab. {pab.letra}"
            except Pabellon.DoesNotExist:
                pass
        elif sede_destino_id:
            try:
                sede = Sede.objects.select_related('campus').get(pk=sede_destino_id)
                destino_data['sede'] = f"{sede.campus.nombre} - {sede.nombre}"
            except Sede.DoesNotExist:
                pass

        # Generar Excel
        buffer = generar_formato_traslado(
            items_data=items_data,
            origen_data=origen_data,
            destino_data=destino_data,
            fecha=timezone.now()
        )

        # Preparar respuesta con nombre descriptivo
        import re
        fecha_str = timezone.now().strftime('%Y%m%d')

        # Función para limpiar nombres
        def limpiar_nombre(nombre):
            nombre = nombre.replace(' ', '_')
            nombre = re.sub(r'[^\w\-_]', '', nombre)
            return nombre[:30]

        # Generar nombre descriptivo: "FT de [origen] a [destino]"
        origen_nombre = origen_data.get('ubicacion') or 'Origen'
        destino_nombre = destino_data.get('ubicacion') or destino_data.get('piso') or destino_data.get('sede') or 'Destino'

        origen_limpio = limpiar_nombre(origen_nombre)
        destino_limpio = limpiar_nombre(destino_nombre)

        filename = f"FT_de_{origen_limpio}_a_{destino_limpio}_{fecha_str}.xlsx"

        response = HttpResponse(
            buffer.read(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        return response
