from django.shortcuts import render, redirect, get_object_or_404
from django.views.generic import (
    ListView, DetailView, CreateView, UpdateView, DeleteView, TemplateView, View
)
from django.contrib.auth.mixins import LoginRequiredMixin, UserPassesTestMixin
from django.contrib.auth.models import User
from django.urls import reverse_lazy, reverse
from django.db.models import Q, Sum, Count
from django.contrib import messages
from django.http import JsonResponse, HttpResponse
from django.utils import timezone
from datetime import timedelta, date

from .models import (
    Area, Campus, Sede, Pabellon, Ambiente, TipoItem, Item, EspecificacionesSistemas,
    Movimiento, HistorialCambio, Notificacion, PerfilUsuario,
    Proveedor, Contrato, AnexoContrato, Lote, Mantenimiento
)
from .forms import (
    ItemForm, ItemSistemasForm, MovimientoForm, TipoItemForm, AmbienteForm,
    CampusForm, SedeForm, PabellonForm, MantenimientoForm, MantenimientoFinalizarForm,
    MantenimientoLoteForm
)
from .signals import set_current_user
from .ratelimit import RateLimitMixin


# ============================================================================
# MIXINS DE PERMISOS
# ============================================================================

class PerfilRequeridoMixin(LoginRequiredMixin):
    """Mixin que verifica que el usuario tenga perfil."""
    
    def dispatch(self, request, *args, **kwargs):
        if request.user.is_authenticated:
            # Establecer usuario actual para signals
            set_current_user(request.user)
            
            # Crear perfil si no existe
            if not hasattr(request.user, 'perfil'):
                # Superusuarios obtienen rol admin automáticamente
                rol = 'admin' if request.user.is_superuser else 'operador'
                PerfilUsuario.objects.create(usuario=request.user, rol=rol)
        return super().dispatch(request, *args, **kwargs)
    
    def get_user_area(self):
        """Obtiene el área del usuario actual."""
        if hasattr(self.request.user, 'perfil'):
            return self.request.user.perfil.area
        return None
    
    def get_user_rol(self):
        """Obtiene el rol del usuario actual."""
        if hasattr(self.request.user, 'perfil'):
            return self.request.user.perfil.rol
        return 'operador'
    
    def es_admin(self):
        """Verifica si el usuario es admin."""
        return self.get_user_rol() == 'admin'
    
    def es_supervisor(self):
        """Verifica si el usuario es supervisor."""
        return self.get_user_rol() == 'supervisor'


class SupervisorRequeridoMixin(PerfilRequeridoMixin, UserPassesTestMixin):
    """Solo supervisores y admins pueden acceder."""
    
    def test_func(self):
        return self.get_user_rol() in ['admin', 'supervisor']


class AdminRequeridoMixin(PerfilRequeridoMixin, UserPassesTestMixin):
    """Solo admins pueden acceder."""

    def test_func(self):
        return self.get_user_rol() == 'admin'


class CampusFilterMixin:
    """
    Mixin para filtrar querysets según los campus permitidos del usuario.

    Uso:
    - Admin: ve todo (sin filtro)
    - Supervisor: ve solo los campus que tiene asignados
    - Operador: ve solo su campus asignado
    """

    def get_campus_permitidos(self):
        """Retorna los campus que el usuario puede ver."""
        if hasattr(self.request.user, 'perfil'):
            return self.request.user.perfil.get_campus_permitidos()
        return Campus.objects.none()

    def filtrar_por_campus(self, queryset, campo_campus='ambiente__pabellon__sede__campus'):
        """
        Filtra un queryset según los campus permitidos del usuario.

        Args:
            queryset: El queryset a filtrar
            campo_campus: El campo que relaciona con campus (ej: 'ambiente__pabellon__sede__campus')

        Returns:
            Queryset filtrado
        """
        if not hasattr(self.request.user, 'perfil'):
            return queryset.none()

        perfil = self.request.user.perfil

        # Admin ve todo
        if perfil.rol == 'admin':
            return queryset

        # Obtener IDs de campus permitidos
        campus_ids = list(self.get_campus_permitidos().values_list('id', flat=True))

        if not campus_ids:
            return queryset.none()

        # Construir filtro dinámico
        filtro = {f'{campo_campus}__in': campus_ids}
        return queryset.filter(**filtro)


# ============================================================================
# VISTA DE INICIO
# ============================================================================

class HomeView(TemplateView):
    """Página de inicio - accesible a todos."""
    template_name = 'index.html'
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        
        if self.request.user.is_authenticated:
            user = self.request.user
            perfil = getattr(user, 'perfil', None)
            
            # Filtrar por área si no es admin
            items = Item.objects.all()
            if perfil and perfil.rol != 'admin' and perfil.area:
                items = items.filter(area=perfil.area)
            
            context['total_items'] = items.count()
            context['items_nuevos'] = items.filter(estado='nuevo').count()
            context['items_instalados'] = items.filter(estado='instalado').count()
            context['items_dañados'] = items.filter(estado='dañado').count()
            context['valor_total'] = items.aggregate(valor=Sum('precio'))['valor'] or 0
            
            # Notificaciones no leídas
            context['notificaciones_count'] = Notificacion.objects.filter(
                usuario=user, leida=False
            ).count()
            
            # Movimientos pendientes (para supervisores)
            if perfil and perfil.rol in ['admin', 'supervisor']:
                pendientes = Movimiento.objects.filter(estado='pendiente')
                if perfil.rol == 'supervisor' and perfil.area:
                    pendientes = pendientes.filter(item__area=perfil.area)
                context['movimientos_pendientes'] = pendientes.count()
        
        return context


# ============================================================================
# DASHBOARD
# ============================================================================

class DashboardView(PerfilRequeridoMixin, CampusFilterMixin, TemplateView):
    """Dashboard principal del inventario."""
    template_name = 'productos/dashboard.html'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)

        user = self.request.user
        perfil = getattr(user, 'perfil', None)

        # Base queryset - filtrado por campus permitidos
        items = Item.objects.all()

        # Filtrar por campus según permisos
        items = self.filtrar_por_campus(items, 'ambiente__pabellon__sede__campus')

        # Filtrar por área si no es admin
        if perfil and perfil.rol != 'admin' and perfil.area:
            items = items.filter(area=perfil.area)

        # Estadísticas generales
        context['total_items'] = items.count()
        context['items_por_estado'] = {
            'nuevo': items.filter(estado='nuevo').count(),
            'instalado': items.filter(estado='instalado').count(),
            'dañado': items.filter(estado='dañado').count(),
            'obsoleto': items.filter(estado='obsoleto').count(),
        }
        context['valor_total'] = items.aggregate(valor=Sum('precio'))['valor'] or 0

        # Items por área (filtrado por campus)
        campus_ids = list(self.get_campus_permitidos().values_list('id', flat=True))
        if perfil and perfil.rol == 'admin':
            context['items_por_area'] = Area.objects.annotate(
                total=Count('items')
            ).values('nombre', 'total')
        else:
            context['items_por_area'] = Area.objects.annotate(
                total=Count('items', filter=Q(items__ambiente__pabellon__sede__campus_id__in=campus_ids))
            ).values('nombre', 'total')

        # Garantías próximas a vencer (30 días)
        fecha_limite = timezone.now().date() + timezone.timedelta(days=30)
        context['garantias_proximas'] = items.filter(
            garantia_hasta__lte=fecha_limite,
            garantia_hasta__gte=timezone.now().date()
        ).count()

        # Ítems sin código UTP (pendientes de etiqueta de logística)
        context['items_sin_codigo_utp'] = items.filter(codigo_utp='PENDIENTE').count()

        # Últimos movimientos - filtrados por campus
        movimientos = Movimiento.objects.select_related('item', 'solicitado_por')
        movimientos = self.filtrar_por_campus(movimientos, 'item__ambiente__pabellon__sede__campus')
        if perfil and perfil.rol != 'admin' and perfil.area:
            movimientos = movimientos.filter(item__area=perfil.area)
        context['ultimos_movimientos'] = movimientos[:10]

        # Movimientos pendientes de aprobar
        if perfil and perfil.rol in ['admin', 'supervisor']:
            pendientes = Movimiento.objects.filter(estado='pendiente')
            pendientes = self.filtrar_por_campus(pendientes, 'item__ambiente__pabellon__sede__campus')
            if perfil.rol == 'supervisor' and perfil.area:
                pendientes = pendientes.filter(item__area=perfil.area)
            context['movimientos_pendientes'] = pendientes[:5]

        # Notificaciones
        context['notificaciones'] = Notificacion.objects.filter(
            usuario=user, leida=False
        )[:5]

        # Mantenimientos - filtrados por campus
        mantenimientos = Mantenimiento.objects.select_related('item')
        mantenimientos = self.filtrar_por_campus(mantenimientos, 'item__ambiente__pabellon__sede__campus')
        if perfil and perfil.rol != 'admin' and perfil.area:
            mantenimientos = mantenimientos.filter(item__area=perfil.area)

        context['mantenimientos_pendientes'] = mantenimientos.filter(estado='pendiente').count()
        context['mantenimientos_vencidos'] = mantenimientos.filter(
            estado='pendiente',
            fecha_programada__lt=timezone.now().date()
        ).count()
        context['ultimos_mantenimientos'] = mantenimientos.order_by('-fecha_programada')[:5]

        # Campus del usuario para mostrar en dashboard
        context['campus_usuario'] = self.get_campus_permitidos()

        return context


# ============================================================================
# VISTAS DE ITEMS
# ============================================================================

class ItemListView(PerfilRequeridoMixin, CampusFilterMixin, ListView):
    """Lista de ítems del inventario."""
    model = Item
    template_name = 'productos/item_list.html'
    context_object_name = 'items'
    paginate_by = 20

    def get_queryset(self):
        queryset = Item.objects.select_related(
            'area', 'tipo_item', 'ambiente', 'usuario_asignado',
            'ambiente__pabellon__sede__campus', 'colaborador_asignado'
        )

        perfil = getattr(self.request.user, 'perfil', None)

        # Filtrar por campus según permisos del usuario
        queryset = self.filtrar_por_campus(queryset, 'ambiente__pabellon__sede__campus')

        # Filtrar por área si no es admin - SIEMPRE aplicar esta restricción
        if perfil and perfil.rol != 'admin' and perfil.area:
            # Operadores y supervisores SOLO ven su área, no pueden filtrar otras
            queryset = queryset.filter(area=perfil.area)
        else:
            # Solo admin puede filtrar por cualquier área
            area = self.request.GET.get('area')
            if area:
                queryset = queryset.filter(area__codigo=area)

        # ===== FILTROS AVANZADOS =====

        # Búsqueda general (código interno, código UTP, serie, nombre)
        search = self.request.GET.get('q')
        if search:
            queryset = queryset.filter(
                Q(codigo_interno__icontains=search) |
                Q(codigo_utp__icontains=search) |
                Q(serie__icontains=search) |
                Q(nombre__icontains=search) |
                Q(descripcion__icontains=search)
            )

        # Filtro por estado (puede ser múltiple)
        estados = self.request.GET.getlist('estado')
        if estados:
            queryset = queryset.filter(estado__in=estados)

        # Filtro por tipo de ítem
        tipo_item = self.request.GET.get('tipo_item')
        if tipo_item:
            queryset = queryset.filter(tipo_item_id=tipo_item)

        # Filtro por código UTP pendiente
        utp_pendiente = self.request.GET.get('utp_pendiente')
        if utp_pendiente == '1':
            queryset = queryset.filter(codigo_utp='PENDIENTE')
        elif utp_pendiente == '0':
            queryset = queryset.exclude(codigo_utp='PENDIENTE')

        # Filtro por usuario asignado
        usuario_asignado = self.request.GET.get('usuario_asignado')
        if usuario_asignado == 'sin_asignar':
            queryset = queryset.filter(usuario_asignado__isnull=True)
        elif usuario_asignado == 'asignado':
            queryset = queryset.filter(usuario_asignado__isnull=False)
        elif usuario_asignado:
            queryset = queryset.filter(usuario_asignado_id=usuario_asignado)

        # Filtro por ambiente
        ambiente = self.request.GET.get('ambiente')
        if ambiente == 'sin_ubicacion':
            queryset = queryset.filter(ambiente__isnull=True)
        elif ambiente:
            queryset = queryset.filter(ambiente_id=ambiente)

        # Filtro por campus
        campus = self.request.GET.get('campus')
        if campus:
            queryset = queryset.filter(ambiente__pabellon__sede__campus_id=campus)

        # Filtro por rango de precios
        precio_min = self.request.GET.get('precio_min')
        precio_max = self.request.GET.get('precio_max')
        if precio_min:
            queryset = queryset.filter(precio__gte=precio_min)
        if precio_max:
            queryset = queryset.filter(precio__lte=precio_max)

        # Filtro por rango de fechas de adquisición
        fecha_adq_desde = self.request.GET.get('fecha_adq_desde')
        fecha_adq_hasta = self.request.GET.get('fecha_adq_hasta')
        if fecha_adq_desde:
            queryset = queryset.filter(fecha_adquisicion__gte=fecha_adq_desde)
        if fecha_adq_hasta:
            queryset = queryset.filter(fecha_adquisicion__lte=fecha_adq_hasta)

        # Filtro por garantía
        garantia = self.request.GET.get('garantia')
        if garantia == 'vigente':
            queryset = queryset.filter(garantia_hasta__gte=timezone.now().date())
        elif garantia == 'vencida':
            queryset = queryset.filter(garantia_hasta__lt=timezone.now().date())
        elif garantia == 'sin_garantia':
            queryset = queryset.filter(garantia_hasta__isnull=True)
        elif garantia == 'proxima_vencer':  # 30 días
            fecha_limite = timezone.now().date() + timedelta(days=30)
            queryset = queryset.filter(
                garantia_hasta__lte=fecha_limite,
                garantia_hasta__gte=timezone.now().date()
            )

        # Filtro por leasing
        es_leasing = self.request.GET.get('es_leasing')
        if es_leasing == '1':
            queryset = queryset.filter(es_leasing=True)
        elif es_leasing == '0':
            queryset = queryset.filter(es_leasing=False)

        # Filtro por lote
        lote = self.request.GET.get('lote')
        if lote == 'sin_lote':
            queryset = queryset.filter(lote__isnull=True)
        elif lote:
            queryset = queryset.filter(lote_id=lote)

        # Ordenamiento
        order_by = self.request.GET.get('order_by', '-creado_en')
        if order_by:
            queryset = queryset.order_by(order_by)

        return queryset
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)

        # Opciones para filtros
        context['areas'] = Area.objects.filter(activo=True)
        context['campus_list'] = Campus.objects.filter(activo=True)
        context['ambientes'] = Ambiente.objects.filter(activo=True).select_related('pabellon__sede__campus')
        context['estados'] = Item.ESTADOS
        context['tipos_item'] = TipoItem.objects.filter(activo=True).select_related('area')
        context['usuarios'] = User.objects.filter(is_active=True).order_by('first_name', 'last_name')
        context['lotes'] = Lote.objects.filter(activo=True).order_by('-creado_en')[:50]

        # Filtros activos (para mostrar chips)
        context['filtros_activos'] = {
            'q': self.request.GET.get('q', ''),
            'area': self.request.GET.get('area', ''),
            'estado': self.request.GET.getlist('estado'),
            'tipo_item': self.request.GET.get('tipo_item', ''),
            'utp_pendiente': self.request.GET.get('utp_pendiente', ''),
            'usuario_asignado': self.request.GET.get('usuario_asignado', ''),
            'ambiente': self.request.GET.get('ambiente', ''),
            'campus': self.request.GET.get('campus', ''),
            'precio_min': self.request.GET.get('precio_min', ''),
            'precio_max': self.request.GET.get('precio_max', ''),
            'fecha_adq_desde': self.request.GET.get('fecha_adq_desde', ''),
            'fecha_adq_hasta': self.request.GET.get('fecha_adq_hasta', ''),
            'garantia': self.request.GET.get('garantia', ''),
            'es_leasing': self.request.GET.get('es_leasing', ''),
            'lote': self.request.GET.get('lote', ''),
            'order_by': self.request.GET.get('order_by', '-creado_en'),
        }

        # Contar filtros activos (para badge)
        filtros_count = sum(1 for k, v in context['filtros_activos'].items()
                           if v and k != 'order_by' and v != [])
        context['filtros_count'] = filtros_count

        # Contar solo filtros avanzados (excluyendo búsqueda simple 'q')
        filtros_avanzados = sum(1 for k, v in context['filtros_activos'].items()
                               if v and k not in ['order_by', 'q'] and v != [])
        context['filtros_avanzados_activos'] = filtros_avanzados > 0

        return context


class ItemDetailView(PerfilRequeridoMixin, DetailView):
    """Detalle de un ítem."""
    model = Item
    template_name = 'productos/item_detail.html'
    context_object_name = 'item'
    slug_field = 'codigo_interno'
    slug_url_kwarg = 'codigo'
    
    def get_queryset(self):
        """Restringir acceso por área si no es admin."""
        queryset = super().get_queryset()
        perfil = getattr(self.request.user, 'perfil', None)
        
        if perfil and perfil.rol != 'admin' and perfil.area:
            queryset = queryset.filter(area=perfil.area)
        
        return queryset
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        item = self.object
        
        # Especificaciones de sistemas si aplica
        if item.area.codigo == 'sistemas':
            context['especificaciones'] = getattr(item, 'especificaciones_sistemas', None)
        
        # Historial de movimientos
        context['movimientos'] = item.movimientos.select_related(
            'solicitado_por', 'autorizado_por'
        )[:10]
        
        # Historial de cambios
        context['historial'] = item.historial_cambios.select_related('usuario')[:10]
        
        return context


class ItemCreateView(PerfilRequeridoMixin, CreateView):
    """Crear un nuevo ítem."""
    model = Item
    form_class = ItemForm
    template_name = 'productos/item_form.html'
    
    def get_form_kwargs(self):
        kwargs = super().get_form_kwargs()
        kwargs['user'] = self.request.user
        return kwargs
    
    def get_initial(self):
        """Pre-seleccionar el área del usuario si no es admin."""
        initial = super().get_initial()
        perfil = getattr(self.request.user, 'perfil', None)
        
        if perfil and perfil.rol != 'admin' and perfil.area:
            initial['area'] = perfil.area
        
        return initial
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['titulo'] = 'Nuevo Ítem'
        context['es_sistemas'] = self.request.GET.get('area') == 'sistemas'
        return context
    
    def form_valid(self, form):
        item = form.save(commit=False)
        item.creado_por = self.request.user
        item.modificado_por = self.request.user
        
        # Si no es admin, forzar el área del usuario
        perfil = getattr(self.request.user, 'perfil', None)
        if perfil and perfil.rol != 'admin' and perfil.area:
            item.area = perfil.area
        
        # Auto-generar código UTP
        if not item.codigo_utp:
            item.codigo_utp = Item.generar_codigo_utp(item.area.codigo)
        
        item.save()

        messages.success(self.request, f'Ítem {item.codigo_interno} creado correctamente.')
        return redirect('productos:item-detail', codigo=item.codigo_interno)


class ItemUpdateView(PerfilRequeridoMixin, UpdateView):
    """Editar un ítem existente."""
    model = Item
    form_class = ItemForm
    template_name = 'productos/item_form.html'
    slug_field = 'codigo_interno'
    slug_url_kwarg = 'codigo'
    
    def get_queryset(self):
        """Restringir acceso por área si no es admin."""
        queryset = super().get_queryset()
        perfil = getattr(self.request.user, 'perfil', None)
        
        if perfil and perfil.rol != 'admin' and perfil.area:
            queryset = queryset.filter(area=perfil.area)
        
        return queryset
    
    def get_form_kwargs(self):
        kwargs = super().get_form_kwargs()
        kwargs['user'] = self.request.user
        return kwargs
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['titulo'] = f'Editar {self.object.codigo_utp}'
        context['es_sistemas'] = self.object.area.codigo == 'sistemas'
        return context
    
    def form_valid(self, form):
        item = form.save(commit=False)
        item.modificado_por = self.request.user
        item.save()

        messages.success(self.request, f'Ítem {item.codigo_interno} actualizado correctamente.')
        return redirect('productos:item-detail', codigo=item.codigo_interno)


class ItemDeleteView(AdminRequeridoMixin, DeleteView):
    """Eliminar un ítem (solo admin)."""
    model = Item
    template_name = 'productos/item_confirm_delete.html'
    success_url = reverse_lazy('productos:item-list')
    slug_field = 'codigo_interno'
    slug_url_kwarg = 'codigo'

    def delete(self, request, *args, **kwargs):
        item = self.get_object()
        messages.success(request, f'Ítem {item.codigo_interno} eliminado.')
        return super().delete(request, *args, **kwargs)


# ============================================================================
# VISTAS DE MOVIMIENTOS
# ============================================================================

class MovimientoListView(PerfilRequeridoMixin, ListView):
    """Lista de movimientos."""
    model = Movimiento
    template_name = 'productos/movimiento_list.html'
    context_object_name = 'movimientos'
    paginate_by = 20
    
    def get_queryset(self):
        queryset = Movimiento.objects.select_related(
            'item', 'solicitado_por', 'autorizado_por'
        )
        
        perfil = getattr(self.request.user, 'perfil', None)
        
        # Filtrar por área si no es admin
        if perfil and perfil.rol != 'admin' and perfil.area:
            queryset = queryset.filter(item__area=perfil.area)
        
        # Filtro por estado
        estado = self.request.GET.get('estado')
        if estado:
            queryset = queryset.filter(estado=estado)
        
        return queryset
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['estados'] = Movimiento.ESTADOS_MOVIMIENTO
        return context


class MovimientoPendientesView(SupervisorRequeridoMixin, ListView):
    """Movimientos pendientes de aprobar."""
    model = Movimiento
    template_name = 'productos/movimiento_pendientes.html'
    context_object_name = 'movimientos'
    
    def get_queryset(self):
        queryset = Movimiento.objects.filter(estado='pendiente').select_related(
            'item', 'solicitado_por'
        )
        
        perfil = getattr(self.request.user, 'perfil', None)
        
        # Supervisor solo ve de su área
        if perfil and perfil.rol == 'supervisor' and perfil.area:
            queryset = queryset.filter(item__area=perfil.area)
        
        return queryset


class MovimientoCreateView(PerfilRequeridoMixin, CreateView):
    """Crear un nuevo movimiento."""
    model = Movimiento
    form_class = MovimientoForm
    template_name = 'productos/movimiento_form.html'
    
    def get_form_kwargs(self):
        kwargs = super().get_form_kwargs()
        kwargs['user'] = self.request.user
        
        # Pre-cargar ítem si viene en la URL
        item_pk = self.request.GET.get('item')
        if item_pk:
            kwargs['item'] = get_object_or_404(Item, pk=item_pk)
        
        return kwargs
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['titulo'] = 'Nueva Solicitud de Movimiento'
        
        # Datos para los filtros de búsqueda de ítems
        context['campus_list'] = Campus.objects.filter(activo=True)
        context['areas_list'] = Area.objects.filter(activo=True)
        
        # Ítem preseleccionado si viene en URL
        item_pk = self.request.GET.get('item')
        if item_pk:
            context['item'] = get_object_or_404(Item, pk=item_pk)
        
        return context
    
    def form_valid(self, form):
        movimiento = form.save(commit=False)
        movimiento.solicitado_por = self.request.user
        
        # Si es emergencia, ejecutar inmediatamente
        if movimiento.es_emergencia:
            movimiento.estado = 'ejecutado_emergencia'
            movimiento.fecha_ejecucion = timezone.now()
            
            # Aplicar cambios al ítem
            item = movimiento.item
            if movimiento.tipo == 'traslado' and movimiento.ambiente_destino:
                movimiento.ambiente_origen = item.ambiente
                item.ambiente = movimiento.ambiente_destino
            if movimiento.tipo == 'cambio_estado' and movimiento.estado_item_nuevo:
                movimiento.estado_item_anterior = item.estado
                item.estado = movimiento.estado_item_nuevo
            if movimiento.tipo == 'asignacion':
                movimiento.usuario_anterior = item.usuario_asignado
                item.usuario_asignado = movimiento.usuario_nuevo
            item.save()
        else:
            # Guardar estado actual como origen
            if movimiento.tipo == 'traslado':
                movimiento.ambiente_origen = movimiento.item.ambiente
            if movimiento.tipo == 'cambio_estado':
                movimiento.estado_item_anterior = movimiento.item.estado
            if movimiento.tipo == 'asignacion':
                movimiento.usuario_anterior = movimiento.item.usuario_asignado
        
        movimiento.save()
        
        if movimiento.es_emergencia:
            messages.warning(
                self.request, 
                f'Movimiento de EMERGENCIA ejecutado. Pendiente de validación por {movimiento.autorizado_por}.'
            )
        else:
            messages.success(
                self.request, 
                f'Solicitud creada. Pendiente de aprobación por {movimiento.autorizado_por}.'
            )
        
        return redirect('productos:movimiento-list')


class MovimientoDetailView(PerfilRequeridoMixin, DetailView):
    """Detalle de un movimiento."""
    model = Movimiento
    template_name = 'productos/movimiento_detail.html'
    context_object_name = 'movimiento'
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        # Pasar el perfil del usuario al contexto
        context['perfil'] = getattr(self.request.user, 'perfil', None)
        # Detectar si viene de pendientes para el breadcrumb
        context['from_pendientes'] = 'pendientes' in self.request.META.get('HTTP_REFERER', '')
        return context


class MovimientoAprobarView(SupervisorRequeridoMixin, View):
    """Aprobar un movimiento."""
    
    def post(self, request, pk):
        movimiento = get_object_or_404(Movimiento, pk=pk)
        
        # Verificar que puede aprobar
        perfil = getattr(request.user, 'perfil', None)
        if perfil and perfil.rol == 'supervisor':
            if perfil.area != movimiento.item.area:
                messages.error(request, 'No puedes aprobar movimientos de otra área.')
                return redirect('productos:movimiento-detail', pk=pk)
        
        if movimiento.estado not in ['pendiente', 'ejecutado_emergencia']:
            messages.error(request, 'Este movimiento no puede ser aprobado.')
            return redirect('productos:movimiento-detail', pk=pk)
        
        # Aprobar
        movimiento.aprobar(request.user)
        
        # Si no era emergencia, ejecutar
        if movimiento.estado != 'ejecutado_emergencia':
            movimiento.ejecutar()
        
        messages.success(request, 'Movimiento aprobado y ejecutado.')
        return redirect('productos:movimiento-pendientes')


class MovimientoRechazarView(SupervisorRequeridoMixin, View):
    """Rechazar un movimiento."""
    
    def post(self, request, pk):
        movimiento = get_object_or_404(Movimiento, pk=pk)
        motivo = request.POST.get('motivo_rechazo', 'Sin motivo especificado')
        
        # Verificar que puede rechazar
        perfil = getattr(request.user, 'perfil', None)
        if perfil and perfil.rol == 'supervisor':
            if perfil.area != movimiento.item.area:
                messages.error(request, 'No puedes rechazar movimientos de otra área.')
                return redirect('productos:movimiento-detail', pk=pk)
        
        if movimiento.estado not in ['pendiente', 'ejecutado_emergencia']:
            messages.error(request, 'Este movimiento no puede ser rechazado.')
            return redirect('productos:movimiento-detail', pk=pk)
        
        # Si era emergencia, revertir cambios
        if movimiento.estado == 'ejecutado_emergencia':
            item = movimiento.item
            if movimiento.ambiente_origen:
                item.ambiente = movimiento.ambiente_origen
            if movimiento.estado_item_anterior:
                item.estado = movimiento.estado_item_anterior
            if movimiento.usuario_anterior:
                item.usuario_asignado = movimiento.usuario_anterior
            item.save()
            movimiento.estado = 'revertido'
        
        movimiento.rechazar(request.user, motivo)
        
        messages.success(request, 'Movimiento rechazado.')
        return redirect('productos:movimiento-pendientes')


# ============================================================================
# VISTAS DE NOTIFICACIONES
# ============================================================================

class NotificacionListView(PerfilRequeridoMixin, ListView):
    """Lista de notificaciones del usuario."""
    model = Notificacion
    template_name = 'productos/notificacion_list.html'
    context_object_name = 'notificaciones'
    paginate_by = 20
    
    def get_queryset(self):
        return Notificacion.objects.filter(usuario=self.request.user)


class NotificacionMarcarLeidaView(PerfilRequeridoMixin, View):
    """Marcar notificación como leída."""
    
    def post(self, request, pk):
        notificacion = get_object_or_404(Notificacion, pk=pk, usuario=request.user)
        notificacion.marcar_leida()
        
        # Redirigir a la URL de la notificación si existe
        if notificacion.url:
            return redirect(notificacion.url)
        return redirect('productos:notificacion-list')


class NotificacionMarcarTodasLeidasView(PerfilRequeridoMixin, View):
    """Marcar todas las notificaciones como leídas."""
    
    def post(self, request):
        Notificacion.objects.filter(usuario=request.user, leida=False).update(leida=True)
        messages.success(request, 'Todas las notificaciones marcadas como leídas.')
        return redirect('productos:notificacion-list')


# ============================================================================
# API ENDPOINTS (JSON)
# ============================================================================

class TiposItemPorAreaView(PerfilRequeridoMixin, View):
    """Obtiene los tipos de ítem para un área (para formularios dinámicos)."""
    
    def get(self, request):
        area_id = request.GET.get('area')
        tipos = []
        
        if area_id:
            tipos = list(TipoItem.objects.filter(
                area_id=area_id, activo=True
            ).values('id', 'nombre'))
        
        return JsonResponse({'tipos': tipos})


class SupervisoresPorAreaView(PerfilRequeridoMixin, View):
    """Obtiene los supervisores de un área (para seleccionar autorizador)."""
    
    def get(self, request):
        area_id = request.GET.get('area')
        supervisores = []
        
        if area_id:
            from django.contrib.auth.models import User
            supervisores = list(
                User.objects.filter(
                    perfil__area_id=area_id,
                    perfil__rol='supervisor',
                    perfil__activo=True,
                    is_active=True
                ).values('id', 'first_name', 'last_name', 'username')
            )
        
        # Agregar admins siempre
        admins = list(
            User.objects.filter(
                perfil__rol='admin',
                perfil__activo=True,
                is_active=True
            ).values('id', 'first_name', 'last_name', 'username')
        )
        
        return JsonResponse({'supervisores': supervisores + admins})


# ============================================================================
# VISTAS PARA CREAR TIPO ITEM (accesible a operadores)
# ============================================================================

class TipoItemCreateView(PerfilRequeridoMixin, CreateView):
    """Vista para crear tipos de ítem (accesible a todos los roles excepto externos)."""
    model = TipoItem
    form_class = TipoItemForm
    template_name = 'productos/tipoitem_form.html'
    
    def get_form_kwargs(self):
        kwargs = super().get_form_kwargs()
        kwargs['user'] = self.request.user
        return kwargs
    
    def form_valid(self, form):
        # Si el usuario tiene área asignada y no es admin, usar esa área
        perfil = getattr(self.request.user, 'perfil', None)
        if perfil and perfil.area and perfil.rol != 'admin':
            form.instance.area = perfil.area
        
        messages.success(
            self.request, 
            f'Tipo de ítem "{form.instance.nombre}" creado exitosamente.'
        )
        return super().form_valid(form)
    
    def get_success_url(self):
        # Redirigir a donde vino o al listado de ítems
        next_url = self.request.GET.get('next')
        if next_url:
            return next_url
        return reverse_lazy('productos:item-create')
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['titulo'] = 'Crear Tipo de Ítem'
        
        # Verificar si hay advertencia de tipos similares
        if hasattr(self.get_form(), 'advertencia_similares'):
            context['advertencia_similares'] = self.get_form().advertencia_similares
        
        return context


class TipoItemListView(PerfilRequeridoMixin, ListView):
    """Lista de tipos de ítem."""
    model = TipoItem
    template_name = 'productos/tipoitem_list.html'
    context_object_name = 'tipos'
    paginate_by = 20
    
    def get_queryset(self):
        queryset = TipoItem.objects.select_related('area').filter(activo=True)
        
        # Filtrar por área si el usuario no es admin
        perfil = getattr(self.request.user, 'perfil', None)
        if perfil and perfil.rol != 'admin' and perfil.area:
            queryset = queryset.filter(area=perfil.area)
        
        # Filtro por área desde request
        area_id = self.request.GET.get('area')
        if area_id:
            queryset = queryset.filter(area_id=area_id)
        
        # Búsqueda
        search = self.request.GET.get('q')
        if search:
            queryset = queryset.filter(nombre__icontains=search)
        
        return queryset.order_by('area', 'nombre')
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['areas'] = Area.objects.filter(activo=True)
        return context


# ============================================================================
# VISTAS PARA UBICACIONES
# ============================================================================

class UbicacionListView(PerfilRequeridoMixin, CampusFilterMixin, ListView):
    """Lista de ambientes (ubicaciones) según permisos de campus."""
    model = Ambiente
    template_name = 'productos/ubicacion_list.html'
    context_object_name = 'ambientes'
    paginate_by = 20

    def get_queryset(self):
        queryset = Ambiente.objects.filter(activo=True).select_related(
            'pabellon', 'pabellon__sede', 'pabellon__sede__campus'
        )

        # Filtrar por campus permitidos del usuario
        queryset = self.filtrar_por_campus(queryset, 'pabellon__sede__campus')

        # Filtros adicionales
        campus = self.request.GET.get('campus')
        sede = self.request.GET.get('sede')
        pabellon = self.request.GET.get('pabellon')
        tipo = self.request.GET.get('tipo')
        search = self.request.GET.get('q')

        if campus:
            queryset = queryset.filter(pabellon__sede__campus_id=campus)
        if sede:
            queryset = queryset.filter(pabellon__sede_id=sede)
        if pabellon:
            queryset = queryset.filter(pabellon_id=pabellon)
        if tipo:
            queryset = queryset.filter(tipo=tipo)
        if search:
            queryset = queryset.filter(
                Q(nombre__icontains=search) |
                Q(codigo__icontains=search)
            )

        return queryset.order_by('pabellon__sede__campus__nombre', 'pabellon__sede__nombre', 'pabellon__nombre', 'piso')

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        # Solo mostrar campus/sedes/pabellones que el usuario puede ver
        campus_permitidos = self.get_campus_permitidos()
        campus_ids = list(campus_permitidos.values_list('id', flat=True))

        context['campus_list'] = campus_permitidos
        context['sedes_list'] = Sede.objects.filter(activo=True, campus_id__in=campus_ids).select_related('campus')
        context['pabellones_list'] = Pabellon.objects.filter(activo=True, sede__campus_id__in=campus_ids).select_related('sede')
        context['tipos_ambiente'] = Ambiente.TIPOS_AMBIENTE
        return context


class UbicacionDetailView(PerfilRequeridoMixin, DetailView):
    """Detalle de ambiente (ubicación) con ítems."""
    model = Ambiente
    template_name = 'productos/ubicacion_detail.html'
    context_object_name = 'ambiente'
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        # Ítems en este ambiente
        context['items'] = Item.objects.filter(ambiente=self.object).select_related('area', 'tipo_item')
        return context


class UbicacionCreateView(SupervisorRequeridoMixin, CreateView):
    """Crear ambiente (solo supervisores y admins)."""
    model = Ambiente
    form_class = AmbienteForm
    template_name = 'productos/ubicacion_form.html'
    success_url = reverse_lazy('productos:ubicacion-list')
    
    def form_valid(self, form):
        messages.success(self.request, 'Ambiente creado exitosamente.')
        return super().form_valid(form)
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['titulo'] = 'Crear Ambiente'
        return context


class UbicacionUpdateView(SupervisorRequeridoMixin, UpdateView):
    """Editar ambiente (solo supervisores y admins)."""
    model = Ambiente
    form_class = AmbienteForm
    template_name = 'productos/ubicacion_form.html'
    
    def get_success_url(self):
        return reverse_lazy('productos:ubicacion-detail', kwargs={'pk': self.object.pk})
    
    def form_valid(self, form):
        messages.success(self.request, 'Ambiente actualizado exitosamente.')
        return super().form_valid(form)
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['titulo'] = 'Editar Ambiente'
        context['editando'] = True
        return context


# ============================================================================
# APIs AJAX PARA DROPDOWNS EN CASCADA
# ============================================================================

class SedesPorCampusView(LoginRequiredMixin, View):
    """API para obtener sedes de un campus."""
    
    def get(self, request):
        campus_id = request.GET.get('campus_id')
        if campus_id:
            sedes = Sede.objects.filter(campus_id=campus_id, activo=True).values('id', 'nombre')
            return JsonResponse(list(sedes), safe=False)
        return JsonResponse([], safe=False)


class PabellonesPorSedeView(LoginRequiredMixin, View):
    """API para obtener pabellones de una sede."""
    
    def get(self, request):
        sede_id = request.GET.get('sede_id')
        if sede_id:
            pabellones = Pabellon.objects.filter(sede_id=sede_id, activo=True).values('id', 'nombre')
            return JsonResponse(list(pabellones), safe=False)
        return JsonResponse([], safe=False)


class AmbientesPorPabellonView(LoginRequiredMixin, View):
    """API para obtener ambientes de un pabellón."""
    
    def get(self, request):
        pabellon_id = request.GET.get('pabellon_id')
        if pabellon_id:
            ambientes = Ambiente.objects.filter(pabellon_id=pabellon_id, activo=True).values('id', 'nombre', 'piso', 'tipo')
            return JsonResponse(list(ambientes), safe=False)
        return JsonResponse([], safe=False)


class BuscarItemsView(RateLimitMixin, LoginRequiredMixin, View):
    """API para buscar ítems con autocompletado."""
    ratelimit_key = 'search'

    def get(self, request):
        query = request.GET.get('q', '').strip()
        area_id = request.GET.get('area')
        tipo_id = request.GET.get('tipo')
        campus_id = request.GET.get('campus')
        sede_id = request.GET.get('sede')
        pabellon_id = request.GET.get('pabellon')
        ambiente_id = request.GET.get('ambiente')
        
        items = Item.objects.select_related(
            'area', 'tipo_item', 'ambiente', 'ambiente__pabellon',
            'ambiente__pabellon__sede', 'ambiente__pabellon__sede__campus',
            'usuario_asignado'
        )
        
        # Filtrar por área del usuario si no es admin
        perfil = getattr(request.user, 'perfil', None)
        if perfil and perfil.rol != 'admin' and perfil.area:
            items = items.filter(area=perfil.area)
        
        # Búsqueda por texto
        if query:
            items = items.filter(
                Q(codigo_utp__icontains=query) |
                Q(serie__icontains=query) |
                Q(nombre__icontains=query)
            )
        
        # Filtros adicionales
        if area_id:
            items = items.filter(area_id=area_id)
        if tipo_id:
            items = items.filter(tipo_item_id=tipo_id)
        if campus_id:
            items = items.filter(ambiente__pabellon__sede__campus_id=campus_id)
        if sede_id:
            items = items.filter(ambiente__pabellon__sede_id=sede_id)
        if pabellon_id:
            items = items.filter(ambiente__pabellon_id=pabellon_id)
        if ambiente_id:
            items = items.filter(ambiente_id=ambiente_id)
        
        # Limitar resultados
        items = items[:20]
        
        # Formatear respuesta
        resultados = []
        for item in items:
            ubicacion = ""
            if item.ambiente:
                amb = item.ambiente
                ubicacion = f"{amb.pabellon.sede.campus.codigo} > {amb.pabellon.sede.nombre} > Pab. {amb.pabellon.nombre} > {amb.nombre}"
            
            resultados.append({
                'id': item.id,
                'codigo_utp': item.codigo_utp,
                'serie': item.serie,
                'nombre': item.nombre,
                'area': item.area.nombre,
                'tipo': item.tipo_item.nombre if item.tipo_item else '',
                'estado': item.estado,
                'estado_display': item.get_estado_display(),
                'ubicacion': ubicacion,
                'ambiente_id': item.ambiente_id,
                'usuario_asignado': item.usuario_asignado.get_full_name() if item.usuario_asignado else None,
                'texto': f"{item.codigo_utp} - {item.nombre}"
            })
        
        return JsonResponse({'items': resultados})


class ObtenerItemDetalleView(LoginRequiredMixin, View):
    """API para obtener detalles de un ítem específico."""
    
    def get(self, request):
        item_id = request.GET.get('id')
        if not item_id:
            return JsonResponse({'error': 'ID requerido'}, status=400)
        
        try:
            item = Item.objects.select_related(
                'area', 'tipo_item', 'ambiente', 'ambiente__pabellon',
                'ambiente__pabellon__sede', 'ambiente__pabellon__sede__campus',
                'usuario_asignado'
            ).get(pk=item_id)
        except Item.DoesNotExist:
            return JsonResponse({'error': 'Ítem no encontrado'}, status=404)
        
        ubicacion = ""
        ubicacion_completa = {}
        if item.ambiente:
            amb = item.ambiente
            ubicacion = f"{amb.pabellon.sede.campus.nombre} > {amb.pabellon.sede.nombre} > Pabellón {amb.pabellon.nombre} > {amb.nombre}"
            ubicacion_completa = {
                'campus_id': amb.pabellon.sede.campus_id,
                'campus': amb.pabellon.sede.campus.nombre,
                'sede_id': amb.pabellon.sede_id,
                'sede': amb.pabellon.sede.nombre,
                'pabellon_id': amb.pabellon_id,
                'pabellon': amb.pabellon.nombre,
                'ambiente_id': amb.id,
                'ambiente': amb.nombre,
                'piso': amb.piso
            }
        
        return JsonResponse({
            'id': item.id,
            'codigo_utp': item.codigo_utp,
            'serie': item.serie,
            'nombre': item.nombre,
            'descripcion': item.descripcion,
            'area': item.area.nombre,
            'area_id': item.area_id,
            'tipo': item.tipo_item.nombre if item.tipo_item else '',
            'estado': item.estado,
            'estado_display': item.get_estado_display(),
            'ubicacion': ubicacion,
            'ubicacion_completa': ubicacion_completa,
            'usuario_asignado': item.usuario_asignado.get_full_name() if item.usuario_asignado else None,
            'usuario_asignado_id': item.usuario_asignado_id,
        })


# ============================================================================
# VISTAS CRUD PARA CAMPUS
# ============================================================================

class CampusListView(AdminRequeridoMixin, ListView):
    """Listar todos los campus (solo admins)."""
    model = Campus
    template_name = 'productos/campus_list.html'
    context_object_name = 'campus_list'
    
    def get_queryset(self):
        queryset = Campus.objects.annotate(
            total_sedes=Count('sedes')
        ).order_by('nombre')
        
        # Búsqueda
        q = self.request.GET.get('q')
        if q:
            queryset = queryset.filter(
                Q(nombre__icontains=q) | Q(codigo__icontains=q)
            )
        
        # Filtro por estado
        activo = self.request.GET.get('activo')
        if activo == '1':
            queryset = queryset.filter(activo=True)
        elif activo == '0':
            queryset = queryset.filter(activo=False)
        
        return queryset
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['titulo'] = 'Gestión de Campus'
        return context


class CampusCreateView(AdminRequeridoMixin, CreateView):
    """Crear campus (solo admins)."""
    model = Campus
    form_class = CampusForm
    template_name = 'productos/campus_form.html'
    success_url = reverse_lazy('productos:campus-list')
    
    def form_valid(self, form):
        messages.success(self.request, 'Campus creado exitosamente.')
        return super().form_valid(form)
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['titulo'] = 'Crear Campus'
        return context


class CampusUpdateView(AdminRequeridoMixin, UpdateView):
    """Editar campus (solo admins)."""
    model = Campus
    form_class = CampusForm
    template_name = 'productos/campus_form.html'
    success_url = reverse_lazy('productos:campus-list')
    
    def form_valid(self, form):
        messages.success(self.request, 'Campus actualizado exitosamente.')
        return super().form_valid(form)
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['titulo'] = 'Editar Campus'
        context['editando'] = True
        return context


class CampusDeleteView(AdminRequeridoMixin, DeleteView):
    """Eliminar campus (solo admins)."""
    model = Campus
    template_name = 'productos/campus_confirm_delete.html'
    success_url = reverse_lazy('productos:campus-list')
    
    def form_valid(self, form):
        messages.success(self.request, 'Campus eliminado exitosamente.')
        return super().form_valid(form)


# ============================================================================
# VISTAS CRUD PARA SEDES
# ============================================================================

class SedeListView(PerfilRequeridoMixin, CampusFilterMixin, ListView):
    """Listar sedes según permisos de campus del usuario."""
    model = Sede
    template_name = 'productos/sede_list.html'
    context_object_name = 'sedes'

    def get_queryset(self):
        queryset = Sede.objects.select_related('campus').annotate(
            total_pabellones=Count('pabellones')
        ).order_by('campus__nombre', 'nombre')

        # Filtrar por campus permitidos del usuario
        queryset = self.filtrar_por_campus(queryset, 'campus')

        # Búsqueda
        q = self.request.GET.get('q')
        if q:
            queryset = queryset.filter(
                Q(nombre__icontains=q) | Q(codigo__icontains=q) | Q(campus__nombre__icontains=q)
            )

        # Filtro por campus (solo si tiene acceso a múltiples)
        campus_id = self.request.GET.get('campus')
        if campus_id:
            queryset = queryset.filter(campus_id=campus_id)

        # Filtro por estado
        activo = self.request.GET.get('activo')
        if activo == '1':
            queryset = queryset.filter(activo=True)
        elif activo == '0':
            queryset = queryset.filter(activo=False)

        return queryset

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['titulo'] = 'Gestión de Sedes'
        # Solo mostrar campus que el usuario puede ver
        context['campus_list'] = self.get_campus_permitidos()
        return context


class SedeCreateView(AdminRequeridoMixin, CreateView):
    """Crear sede (solo admins)."""
    model = Sede
    form_class = SedeForm
    template_name = 'productos/sede_form.html'
    success_url = reverse_lazy('productos:sede-list')
    
    def form_valid(self, form):
        messages.success(self.request, 'Sede creada exitosamente.')
        return super().form_valid(form)
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['titulo'] = 'Crear Sede'
        return context


class SedeUpdateView(AdminRequeridoMixin, UpdateView):
    """Editar sede (solo admins)."""
    model = Sede
    form_class = SedeForm
    template_name = 'productos/sede_form.html'
    success_url = reverse_lazy('productos:sede-list')
    
    def form_valid(self, form):
        messages.success(self.request, 'Sede actualizada exitosamente.')
        return super().form_valid(form)
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['titulo'] = 'Editar Sede'
        context['editando'] = True
        return context


class SedeDeleteView(AdminRequeridoMixin, DeleteView):
    """Eliminar sede (solo admins)."""
    model = Sede
    template_name = 'productos/sede_confirm_delete.html'
    success_url = reverse_lazy('productos:sede-list')
    
    def form_valid(self, form):
        messages.success(self.request, 'Sede eliminada exitosamente.')
        return super().form_valid(form)


# ============================================================================
# VISTAS CRUD PARA PABELLONES
# ============================================================================

class PabellonListView(PerfilRequeridoMixin, CampusFilterMixin, ListView):
    """Listar pabellones según permisos de campus del usuario."""
    model = Pabellon
    template_name = 'productos/pabellon_list.html'
    context_object_name = 'pabellones'

    def get_queryset(self):
        queryset = Pabellon.objects.select_related('sede', 'sede__campus').annotate(
            total_ambientes=Count('ambientes')
        ).order_by('sede__campus__nombre', 'sede__nombre', 'nombre')

        # Filtrar por campus permitidos del usuario
        queryset = self.filtrar_por_campus(queryset, 'sede__campus')

        # Búsqueda
        q = self.request.GET.get('q')
        if q:
            queryset = queryset.filter(
                Q(nombre__icontains=q) | Q(sede__nombre__icontains=q) | Q(sede__campus__nombre__icontains=q)
            )

        # Filtro por campus
        campus_id = self.request.GET.get('campus')
        if campus_id:
            queryset = queryset.filter(sede__campus_id=campus_id)

        # Filtro por sede
        sede_id = self.request.GET.get('sede')
        if sede_id:
            queryset = queryset.filter(sede_id=sede_id)

        # Filtro por estado
        activo = self.request.GET.get('activo')
        if activo == '1':
            queryset = queryset.filter(activo=True)
        elif activo == '0':
            queryset = queryset.filter(activo=False)

        return queryset

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['titulo'] = 'Gestión de Pabellones'
        # Solo campus permitidos
        context['campus_list'] = self.get_campus_permitidos()
        # Sedes filtradas por campus permitidos
        campus_ids = list(self.get_campus_permitidos().values_list('id', flat=True))
        context['sedes'] = Sede.objects.filter(activo=True, campus_id__in=campus_ids).select_related('campus')
        return context


class PabellonCreateView(AdminRequeridoMixin, CreateView):
    """Crear pabellón (solo admins)."""
    model = Pabellon
    form_class = PabellonForm
    template_name = 'productos/pabellon_form.html'
    success_url = reverse_lazy('productos:pabellon-list')
    
    def form_valid(self, form):
        messages.success(self.request, 'Pabellón creado exitosamente.')
        return super().form_valid(form)
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['titulo'] = 'Crear Pabellón'
        return context


class PabellonUpdateView(AdminRequeridoMixin, UpdateView):
    """Editar pabellón (solo admins)."""
    model = Pabellon
    form_class = PabellonForm
    template_name = 'productos/pabellon_form.html'
    success_url = reverse_lazy('productos:pabellon-list')
    
    def form_valid(self, form):
        messages.success(self.request, 'Pabellón actualizado exitosamente.')
        return super().form_valid(form)
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['titulo'] = 'Editar Pabellón'
        context['editando'] = True
        return context


class PabellonDeleteView(AdminRequeridoMixin, DeleteView):
    """Eliminar pabellón (solo admins)."""
    model = Pabellon
    template_name = 'productos/pabellon_confirm_delete.html'
    success_url = reverse_lazy('productos:pabellon-list')
    
    def form_valid(self, form):
        messages.success(self.request, 'Pabellón eliminado exitosamente.')
        return super().form_valid(form)


# ============================================================================
# PROVEEDORES (Solo supervisor/admin)
# ============================================================================

class ProveedorListView(SupervisorRequeridoMixin, ListView):
    """Lista de proveedores."""
    model = Proveedor
    template_name = 'productos/proveedor_list.html'
    context_object_name = 'proveedores'
    paginate_by = 20
    
    def get_queryset(self):
        queryset = Proveedor.objects.all()
        q = self.request.GET.get('q')
        if q:
            queryset = queryset.filter(
                Q(ruc__icontains=q) |
                Q(razon_social__icontains=q) |
                Q(nombre_comercial__icontains=q)
            )
        return queryset


class ProveedorCreateView(SupervisorRequeridoMixin, CreateView):
    """Crear proveedor."""
    model = Proveedor
    template_name = 'productos/proveedor_form.html'
    fields = ['ruc', 'razon_social', 'nombre_comercial', 'direccion', 'telefono', 'email', 'contacto']
    success_url = reverse_lazy('productos:proveedor-list')
    
    def form_valid(self, form):
        messages.success(self.request, 'Proveedor creado exitosamente.')
        return super().form_valid(form)


class ProveedorDetailView(SupervisorRequeridoMixin, DetailView):
    """Detalle de proveedor."""
    model = Proveedor
    template_name = 'productos/proveedor_detail.html'
    context_object_name = 'proveedor'
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['contratos'] = self.object.contratos.all()[:10]
        return context


class ProveedorUpdateView(SupervisorRequeridoMixin, UpdateView):
    """Editar proveedor."""
    model = Proveedor
    template_name = 'productos/proveedor_form.html'
    fields = ['ruc', 'razon_social', 'nombre_comercial', 'direccion', 'telefono', 'email', 'contacto', 'activo']
    
    def get_success_url(self):
        return reverse('productos:proveedor-detail', kwargs={'pk': self.object.pk})
    
    def form_valid(self, form):
        messages.success(self.request, 'Proveedor actualizado exitosamente.')
        return super().form_valid(form)


# ============================================================================
# CONTRATOS (Solo supervisor/admin)
# ============================================================================

class ContratoListView(SupervisorRequeridoMixin, ListView):
    """Lista de contratos."""
    model = Contrato
    template_name = 'productos/contrato_list.html'
    context_object_name = 'contratos'
    paginate_by = 20
    
    def get_queryset(self):
        queryset = Contrato.objects.select_related('proveedor')
        q = self.request.GET.get('q')
        if q:
            queryset = queryset.filter(
                Q(numero_contrato__icontains=q) |
                Q(proveedor__razon_social__icontains=q)
            )
        estado = self.request.GET.get('estado')
        if estado:
            queryset = queryset.filter(estado=estado)
        return queryset
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['estados'] = Contrato.ESTADOS
        return context


class ContratoCreateView(SupervisorRequeridoMixin, CreateView):
    """Crear contrato."""
    model = Contrato
    template_name = 'productos/contrato_form.html'
    fields = ['numero_contrato', 'proveedor', 'descripcion', 'fecha_inicio', 'fecha_fin', 'monto_total', 'estado', 'observaciones']
    success_url = reverse_lazy('productos:contrato-list')
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['proveedores'] = Proveedor.objects.filter(activo=True)
        return context
    
    def form_valid(self, form):
        form.instance.creado_por = self.request.user
        messages.success(self.request, 'Contrato creado exitosamente.')
        return super().form_valid(form)


class ContratoDetailView(SupervisorRequeridoMixin, DetailView):
    """Detalle de contrato."""
    model = Contrato
    template_name = 'productos/contrato_detail.html'
    context_object_name = 'contrato'
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['anexos'] = self.object.anexos.all()
        context['lotes'] = self.object.lotes.all()
        context['items_directos'] = self.object.items_directos.all()[:20]
        return context


class ContratoUpdateView(SupervisorRequeridoMixin, UpdateView):
    """Editar contrato."""
    model = Contrato
    template_name = 'productos/contrato_form.html'
    fields = ['numero_contrato', 'proveedor', 'descripcion', 'fecha_inicio', 'fecha_fin', 'monto_total', 'estado', 'observaciones']
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['proveedores'] = Proveedor.objects.filter(activo=True)
        context['editando'] = True
        return context
    
    def get_success_url(self):
        return reverse('productos:contrato-detail', kwargs={'pk': self.object.pk})
    
    def form_valid(self, form):
        messages.success(self.request, 'Contrato actualizado exitosamente.')
        return super().form_valid(form)


class AnexoContratoCreateView(SupervisorRequeridoMixin, CreateView):
    """Crear anexo de contrato."""
    model = AnexoContrato
    template_name = 'productos/anexo_form.html'
    fields = ['numero_anexo', 'fecha', 'descripcion', 'monto_modificacion']
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['contrato'] = get_object_or_404(Contrato, pk=self.kwargs['pk'])
        return context
    
    def form_valid(self, form):
        form.instance.contrato = get_object_or_404(Contrato, pk=self.kwargs['pk'])
        messages.success(self.request, 'Anexo creado exitosamente.')
        return super().form_valid(form)
    
    def get_success_url(self):
        return reverse('productos:contrato-detail', kwargs={'pk': self.kwargs['pk']})


# ============================================================================
# LOTES
# ============================================================================

class LoteListView(PerfilRequeridoMixin, ListView):
    """Lista de lotes."""
    model = Lote
    template_name = 'productos/lote_list.html'
    context_object_name = 'lotes'
    paginate_by = 20
    
    def get_queryset(self):
        queryset = Lote.objects.select_related('contrato__proveedor').annotate(
            num_items=Count('items')
        )
        q = self.request.GET.get('q')
        if q:
            queryset = queryset.filter(
                Q(codigo_lote__icontains=q) |
                Q(codigo_interno__icontains=q) |
                Q(descripcion__icontains=q)
            )
        return queryset
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        # Ocultar info de contrato para operadores
        context['mostrar_contrato'] = self.get_user_rol() in ['admin', 'supervisor']
        return context


class LoteCreateView(SupervisorRequeridoMixin, CreateView):
    """Crear lote - solo supervisores y admins."""
    model = Lote
    template_name = 'productos/lote_form.html'
    fields = ['codigo_lote', 'contrato', 'descripcion', 'fecha_adquisicion', 'observaciones']
    success_url = reverse_lazy('productos:lote-list')
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['contratos'] = Contrato.objects.filter(estado='vigente')
        context['mostrar_contrato'] = True
        return context
    
    def form_valid(self, form):
        form.instance.creado_por = self.request.user
        messages.success(self.request, 'Lote creado exitosamente.')
        return super().form_valid(form)


class LoteDetailView(PerfilRequeridoMixin, DetailView):
    """Detalle de lote."""
    model = Lote
    template_name = 'productos/lote_detail.html'
    context_object_name = 'lote'
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['items'] = self.object.items.select_related('area', 'tipo_item', 'ambiente')[:50]
        context['mostrar_contrato'] = self.get_user_rol() in ['admin', 'supervisor']
        
        # Alertas de garantía agrupadas
        context['alertas_garantia'] = self.object.items_por_garantia
        return context


class LoteUpdateView(SupervisorRequeridoMixin, UpdateView):
    """Editar lote - solo supervisores y admins."""
    model = Lote
    template_name = 'productos/lote_form.html'
    fields = ['codigo_lote', 'contrato', 'descripcion', 'fecha_adquisicion', 'observaciones', 'activo']

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['editando'] = True
        context['contratos'] = Contrato.objects.filter(estado='vigente')
        context['mostrar_contrato'] = True
        return context

    def get_success_url(self):
        return reverse('productos:lote-detail', kwargs={'pk': self.object.pk})

    def form_valid(self, form):
        messages.success(self.request, 'Lote actualizado exitosamente.')
        return super().form_valid(form)


# ============================================================================
# IMPORTACIÓN MASIVA DESDE EXCEL
# ============================================================================

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils.exceptions import InvalidFileException
from django.http import HttpResponse
from django.db import transaction
from datetime import datetime
import io
import logging

logger = logging.getLogger(__name__)


class ItemImportarPlantillaView(PerfilRequeridoMixin, View):
    """Descargar plantilla Excel para importación masiva."""

    def get(self, request):
        # Crear libro de Excel
        wb = Workbook()

        # Hoja 1: Plantilla de datos
        ws = wb.active
        ws.title = "Plantilla Items"

        # Definir encabezados
        headers_obligatorios = [
            'serie', 'nombre', 'area', 'tipo_item', 'precio', 'fecha_adquisicion'
        ]
        headers_opcionales = [
            'codigo_utp', 'descripcion', 'ambiente_codigo', 'estado', 'garantia_hasta',
            'observaciones', 'lote_codigo', 'es_leasing', 'leasing_empresa',
            'leasing_contrato', 'leasing_vencimiento'
        ]
        headers_sistemas = [
            'marca', 'modelo', 'procesador', 'generacion_procesador',
            'ram_total_gb', 'ram_configuracion', 'ram_tipo',
            'almacenamiento_gb', 'almacenamiento_tipo', 'sistema_operativo'
        ]

        headers = headers_obligatorios + headers_opcionales + headers_sistemas

        # Estilo para encabezados obligatorios
        fill_obligatorio = PatternFill(start_color="C8102E", end_color="C8102E", fill_type="solid")
        font_obligatorio = Font(bold=True, color="FFFFFF")

        # Estilo para encabezados opcionales
        fill_opcional = PatternFill(start_color="4A4A4A", end_color="4A4A4A", fill_type="solid")
        font_opcional = Font(bold=True, color="FFFFFF")

        # Estilo para encabezados de sistemas
        fill_sistemas = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
        font_sistemas = Font(bold=True, color="FFFFFF")

        # Escribir encabezados con estilos
        for idx, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=idx, value=header)
            cell.alignment = Alignment(horizontal='center', vertical='center')

            if header in headers_obligatorios:
                cell.fill = fill_obligatorio
                cell.font = font_obligatorio
            elif header in headers_sistemas:
                cell.fill = fill_sistemas
                cell.font = font_sistemas
            else:
                cell.fill = fill_opcional
                cell.font = font_opcional

        # Ajustar ancho de columnas
        for idx in range(1, len(headers) + 1):
            ws.column_dimensions[ws.cell(row=1, column=idx).column_letter].width = 18

        # Agregar filas de ejemplo
        ejemplo1 = [
            'SN123456789',  # serie
            'Laptop Dell Latitude 5430',  # nombre
            'sistemas',  # area
            'Laptop',  # tipo_item
            '3500.00',  # precio
            '2026-01-10',  # fecha_adquisicion
            'UTP296375',  # codigo_utp (opcional, puede ser PENDIENTE o vacío)
            'Laptop corporativa i7',  # descripcion
            '',  # ambiente_codigo
            'nuevo',  # estado
            '2028-01-10',  # garantia_hasta
            '',  # observaciones
            '',  # lote_codigo
            'NO',  # es_leasing
            '',  # leasing_empresa
            '',  # leasing_contrato
            '',  # leasing_vencimiento
            'Dell',  # marca
            'Latitude 5430',  # modelo
            'Intel Core i7-1365U',  # procesador
            '13th Gen',  # generacion_procesador
            '16',  # ram_total_gb
            '2x8GB',  # ram_configuracion
            'DDR4',  # ram_tipo
            '512',  # almacenamiento_gb
            'NVMe',  # almacenamiento_tipo
            'Windows 11 Pro'  # sistema_operativo
        ]

        ejemplo2 = [
            'SN987654321',  # serie
            'Silla ergonómica',  # nombre
            'operaciones',  # area
            'Silla',  # tipo_item
            '450.00',  # precio
            '2026-01-10',  # fecha_adquisicion
            'PENDIENTE',  # codigo_utp (aún sin etiqueta de logística)
            'Silla de oficina con soporte lumbar',  # descripcion
            '',  # ambiente_codigo
            'nuevo',  # estado
            '',  # garantia_hasta
            '',  # observaciones
            '',  # lote_codigo
            'NO',  # es_leasing
        ]
        # Rellenar con vacíos para completar las columnas
        ejemplo2.extend([''] * (len(headers) - len(ejemplo2)))

        ws.append(ejemplo1)
        ws.append(ejemplo2)

        # Hoja 2: Instrucciones
        ws_instrucciones = wb.create_sheet("Instrucciones")
        instrucciones = [
            ["INSTRUCCIONES PARA IMPORTACIÓN MASIVA DE ÍTEMS", ""],
            ["", ""],
            ["1. COLUMNAS OBLIGATORIAS (Rojo):", ""],
            ["   - serie: Número de serie único del fabricante", ""],
            ["   - nombre: Nombre descriptivo del ítem", ""],
            ["   - area: sistemas, operaciones o laboratorio", ""],
            ["   - tipo_item: Debe existir en el sistema para el área", ""],
            ["   - precio: Precio en formato numérico (ej: 1500.00)", ""],
            ["   - fecha_adquisicion: Formato YYYY-MM-DD (ej: 2026-01-10)", ""],
            ["", ""],
            ["2. COLUMNAS OPCIONALES (Gris):", ""],
            ["   - codigo_utp: Código de etiqueta física (ej: UTP296375)", ""],
            ["     Dejar vacío o PENDIENTE si aún no tiene etiqueta de logística", ""],
            ["     Formato: UTP seguido de números", ""],
            ["   - descripcion: Descripción adicional", ""],
            ["   - ambiente_codigo: Código del ambiente (ej: CLN-SP-A-P1-LC-001)", ""],
            ["   - estado: nuevo, instalado, dañado u obsoleto (default: nuevo)", ""],
            ["   - garantia_hasta: Fecha en formato YYYY-MM-DD", ""],
            ["   - lote_codigo: Código del lote existente (ej: LOT-2026-0001)", ""],
            ["   - es_leasing: SI o NO", ""],
            ["", ""],
            ["3. COLUMNAS PARA SISTEMAS (Azul):", ""],
            ["   - Solo usar si area = sistemas", ""],
            ["   - Especificaciones técnicas del equipo", ""],
            ["", ""],
            ["4. NOTAS IMPORTANTES:", ""],
            ["   - El código interno se genera automáticamente (ej: SIS-2026-0001)", ""],
            ["   - El código UTP es la etiqueta física de logística (opcional)", ""],
            ["   - La serie debe ser única en el sistema", ""],
            ["   - Máximo 1000 ítems por archivo", ""],
            ["   - Formato de archivo: .xlsx", ""],
            ["", ""],
            ["5. VALIDACIONES:", ""],
            ["   - Serie única y no vacía", ""],
            ["   - Área válida y coincide con tu perfil", ""],
            ["   - Tipo de ítem existe para el área", ""],
            ["   - Precio numérico positivo", ""],
            ["   - Fechas en formato correcto", ""],
        ]

        for row_data in instrucciones:
            ws_instrucciones.append(row_data)

        # Ajustar ancho de columnas en instrucciones
        ws_instrucciones.column_dimensions['A'].width = 60
        ws_instrucciones.column_dimensions['B'].width = 20

        # Estilo para el título
        ws_instrucciones['A1'].font = Font(bold=True, size=14, color="C8102E")

        # Preparar respuesta HTTP
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = 'attachment; filename=plantilla_items_inventario.xlsx'

        wb.save(response)
        return response


class ItemImportarView(RateLimitMixin, PerfilRequeridoMixin, TemplateView):
    """Vista para subir archivo Excel y mostrar preview con validaciones."""
    ratelimit_key = 'import'
    ratelimit_method = 'POST'

    template_name = 'productos/item_importar.html'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['areas'] = Area.objects.filter(activo=True)
        context['lotes'] = Lote.objects.filter(activo=True)
        return context

    def post(self, request, *args, **kwargs):
        archivo = request.FILES.get('archivo_excel')
        crear_lote = request.POST.get('crear_lote') == 'on'
        lote_descripcion = request.POST.get('lote_descripcion', '')
        lote_existente_id = request.POST.get('lote_existente')

        if not archivo:
            messages.error(request, 'Debe seleccionar un archivo Excel.')
            return redirect('productos:item-importar')

        # CRÍTICO: Validar tamaño de archivo antes de procesarlo
        MAX_FILE_SIZE = 10 * 1024 * 1024  # 10MB
        if archivo.size > MAX_FILE_SIZE:
            messages.error(request, f'El archivo es demasiado grande. Tamaño máximo: 10MB')
            logger.warning(f'Usuario {request.user.username} intentó subir archivo de {archivo.size} bytes')
            return redirect('productos:item-importar')

        if not archivo.name.endswith('.xlsx'):
            messages.error(request, 'El archivo debe ser formato .xlsx')
            return redirect('productos:item-importar')

        try:
            # Cargar archivo Excel
            wb = load_workbook(archivo, data_only=True)
            ws = wb.active

            # Leer encabezados
            headers = [cell.value for cell in ws[1]]

            # Validar que existan las columnas obligatorias
            required_headers = ['serie', 'nombre', 'area', 'tipo_item', 'precio', 'fecha_adquisicion']
            for header in required_headers:
                if header not in headers:
                    messages.error(request, f'Falta la columna obligatoria: {header}')
                    return redirect('productos:item-importar')

            # Procesar filas
            items_preview = []
            errores_globales = []
            series_en_archivo = set()  # CRÍTICO: Detectar series duplicadas dentro del Excel

            # Obtener perfil del usuario
            perfil = request.user.perfil
            area_usuario = perfil.area if perfil.rol != 'admin' else None

            for idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
                if not any(row):  # Fila vacía
                    continue

                if len(items_preview) >= 1000:
                    errores_globales.append('Se alcanzó el límite de 1000 ítems. Las filas restantes no se procesaron.')
                    break

                # Crear diccionario de datos
                item_data = dict(zip(headers, row))

                # Función auxiliar para obtener valor como string
                def get_str(value):
                    if value is None:
                        return ''
                    return str(value).strip()

                # Validaciones
                errores = []
                advertencias = []

                # Validar serie
                serie = get_str(item_data.get('serie'))
                if not serie:
                    errores.append('Serie vacía')
                elif serie in series_en_archivo:
                    # CRÍTICO: Detectar duplicados dentro del mismo archivo
                    errores.append(f'Serie {serie} duplicada dentro del archivo')
                elif Item.objects.filter(serie=serie).exists():
                    errores.append(f'Serie {serie} ya existe en el sistema')
                else:
                    # Agregar a lista de series procesadas
                    series_en_archivo.add(serie)

                # Validar código UTP (opcional)
                codigo_utp = get_str(item_data.get('codigo_utp', 'PENDIENTE')).upper()
                if not codigo_utp:
                    codigo_utp = 'PENDIENTE'
                    advertencias.append('Código UTP pendiente - se asignará etiqueta de logística posteriormente')
                elif codigo_utp != 'PENDIENTE':
                    # Validar formato UTP + números
                    import re
                    if not re.match(r'^UTP\d+$', codigo_utp):
                        errores.append(f'Código UTP "{codigo_utp}" inválido - debe ser UTP seguido de números (ej: UTP296375)')
                    elif Item.objects.filter(codigo_utp=codigo_utp).exists():
                        errores.append(f'Código UTP {codigo_utp} ya existe en el sistema')
                else:
                    # Es PENDIENTE
                    advertencias.append('Código UTP pendiente - se asignará etiqueta de logística posteriormente')

                # Validar área
                area_codigo = get_str(item_data.get('area')).lower()
                if area_codigo not in ['sistemas', 'operaciones', 'laboratorio']:
                    errores.append(f'Área inválida: {area_codigo}')
                elif area_usuario and area_usuario.codigo != area_codigo:
                    errores.append(f'No tienes permiso para crear ítems en el área {area_codigo}')

                # Validar tipo_item
                tipo_item_nombre = get_str(item_data.get('tipo_item'))
                tipo_item = None
                if tipo_item_nombre and area_codigo in ['sistemas', 'operaciones', 'laboratorio']:
                    try:
                        area_obj = Area.objects.get(codigo=area_codigo)
                        tipo_item = TipoItem.objects.get(nombre__iexact=tipo_item_nombre, area=area_obj)
                    except TipoItem.DoesNotExist:
                        errores.append(f'Tipo de ítem "{tipo_item_nombre}" no existe para el área {area_codigo}')
                    except Area.DoesNotExist:
                        pass

                # Validar precio
                precio = item_data.get('precio', '')
                try:
                    precio_decimal = float(str(precio).replace(',', '')) if precio else 0
                    if precio_decimal <= 0:
                        errores.append('Precio debe ser mayor a 0')
                except (ValueError, TypeError):
                    errores.append(f'Precio inválido: {precio}')

                # Validar fecha_adquisicion
                fecha_adq = item_data.get('fecha_adquisicion', '')
                fecha_adq_obj = None
                if fecha_adq:
                    try:
                        if isinstance(fecha_adq, datetime):
                            fecha_adq_obj = fecha_adq.date()
                        else:
                            fecha_adq_str = str(fecha_adq).strip()
                            # Intentar parsear varios formatos
                            for fmt in ['%Y-%m-%d', '%d/%m/%Y', '%d-%m-%Y']:
                                try:
                                    fecha_adq_obj = datetime.strptime(fecha_adq_str, fmt).date()
                                    break
                                except ValueError:
                                    continue
                            if not fecha_adq_obj:
                                errores.append(f'Fecha de adquisición inválida: {fecha_adq}')
                    except:
                        errores.append(f'Fecha de adquisición inválida: {fecha_adq}')
                else:
                    errores.append('Fecha de adquisición vacía')

                # Validar ambiente (opcional)
                ambiente_codigo = get_str(item_data.get('ambiente_codigo'))
                ambiente = None
                if ambiente_codigo:
                    try:
                        ambiente = Ambiente.objects.get(codigo=ambiente_codigo, activo=True)
                    except Ambiente.DoesNotExist:
                        errores.append(f'Ambiente {ambiente_codigo} no existe')
                else:
                    advertencias.append('Sin ubicación asignada')

                # Validar lote (opcional)
                lote_codigo = get_str(item_data.get('lote_codigo'))
                lote = None
                if lote_codigo:
                    try:
                        lote = Lote.objects.get(codigo_interno=lote_codigo, activo=True)
                    except Lote.DoesNotExist:
                        errores.append(f'Lote {lote_codigo} no existe')

                # Validar estado
                estado = get_str(item_data.get('estado', 'nuevo')).lower()
                if estado not in ['nuevo', 'instalado', 'dañado', 'obsoleto']:
                    advertencias.append(f'Estado "{estado}" inválido, se usará "nuevo"')
                    estado = 'nuevo'

                # Validar garantía
                garantia_hasta = item_data.get('garantia_hasta', '')
                garantia_obj = None
                if garantia_hasta:
                    try:
                        if isinstance(garantia_hasta, datetime):
                            garantia_obj = garantia_hasta.date()
                        else:
                            garantia_str = str(garantia_hasta).strip()
                            for fmt in ['%Y-%m-%d', '%d/%m/%Y', '%d-%m-%Y']:
                                try:
                                    garantia_obj = datetime.strptime(garantia_str, fmt).date()
                                    break
                                except ValueError:
                                    continue
                            if not garantia_obj:
                                advertencias.append(f'Fecha de garantía inválida: {garantia_hasta}')
                    except:
                        advertencias.append(f'Fecha de garantía inválida: {garantia_hasta}')
                else:
                    advertencias.append('Sin fecha de garantía')

                # Determinar estado de la fila
                if errores:
                    estado_fila = 'error'
                elif advertencias:
                    estado_fila = 'advertencia'
                else:
                    estado_fila = 'ok'

                items_preview.append({
                    'fila': idx,
                    'data': item_data,
                    'errores': errores,
                    'advertencias': advertencias,
                    'estado': estado_fila
                })

            # Contar estados
            total = len(items_preview)
            con_errores = sum(1 for item in items_preview if item['estado'] == 'error')
            con_advertencias = sum(1 for item in items_preview if item['estado'] == 'advertencia')
            validos = sum(1 for item in items_preview if item['estado'] == 'ok')

            # Guardar en sesión para confirmar después
            request.session['items_preview'] = items_preview
            request.session['crear_lote'] = crear_lote
            request.session['lote_descripcion'] = lote_descripcion
            request.session['lote_existente_id'] = lote_existente_id

            context = self.get_context_data()
            context['items_preview'] = items_preview
            context['total'] = total
            context['con_errores'] = con_errores
            context['con_advertencias'] = con_advertencias
            context['validos'] = validos
            context['puede_importar'] = con_errores == 0 and total > 0
            context['errores_globales'] = errores_globales
            context['crear_lote'] = crear_lote
            context['lote_descripcion'] = lote_descripcion

            return self.render_to_response(context)

        except InvalidFileException:
            messages.error(request, 'El archivo Excel está corrupto o no es válido.')
            logger.error(f'Archivo Excel inválido subido por {request.user.username}')
            return redirect('productos:item-importar')
        except MemoryError:
            messages.error(request, 'El archivo es demasiado grande para procesar.')
            logger.error(f'MemoryError al procesar archivo de {request.user.username}')
            return redirect('productos:item-importar')
        except Exception as e:
            # Log detallado para debugging, mensaje genérico para usuario
            logger.exception(f'Error inesperado en importación de {request.user.username}: {e}')
            messages.error(request, 'Ocurrió un error al procesar el archivo. Por favor, contacte al administrador.')
            return redirect('productos:item-importar')


class ItemImportarConfirmarView(PerfilRequeridoMixin, View):
    """Confirmar y ejecutar la importación masiva."""

    def post(self, request):
        # Función auxiliar para obtener valor como string
        def get_str(value):
            if value is None:
                return ''
            return str(value).strip()

        items_preview = request.session.get('items_preview', [])
        crear_lote = request.session.get('crear_lote', False)
        lote_descripcion = request.session.get('lote_descripcion', '')
        lote_existente_id = request.session.get('lote_existente_id')

        if not items_preview:
            messages.error(request, 'No hay datos para importar. Por favor sube el archivo nuevamente.')
            return redirect('productos:item-importar')

        # Validar que no haya errores
        items_validos = [item for item in items_preview if item['estado'] != 'error']

        if not items_validos:
            messages.error(request, 'No hay ítems válidos para importar.')
            return redirect('productos:item-importar')

        try:
            with transaction.atomic():
                # Crear lote si es necesario
                lote = None
                if crear_lote and lote_descripcion:
                    lote = Lote.objects.create(
                        descripcion=lote_descripcion,
                        fecha_adquisicion=timezone.now().date(),
                        creado_por=request.user
                    )
                elif lote_existente_id:
                    lote = Lote.objects.get(pk=lote_existente_id)

                items_creados = []

                for item_info in items_validos:
                    data = item_info['data']

                    # Obtener área
                    area = Area.objects.get(codigo=get_str(data.get('area')).lower())

                    # Obtener tipo_item
                    tipo_item = TipoItem.objects.get(
                        nombre__iexact=get_str(data.get('tipo_item')),
                        area=area
                    )

                    # Obtener código UTP del Excel (opcional, default PENDIENTE)
                    codigo_utp = get_str(data.get('codigo_utp', 'PENDIENTE')).upper()
                    if not codigo_utp:
                        codigo_utp = 'PENDIENTE'

                    # Parsear fecha de adquisición
                    fecha_adq = data.get('fecha_adquisicion', '')
                    if isinstance(fecha_adq, datetime):
                        fecha_adq_obj = fecha_adq.date()
                    else:
                        fecha_adq_str = str(fecha_adq).strip()
                        for fmt in ['%Y-%m-%d', '%d/%m/%Y', '%d-%m-%Y']:
                            try:
                                fecha_adq_obj = datetime.strptime(fecha_adq_str, fmt).date()
                                break
                            except ValueError:
                                continue

                    # Parsear garantía
                    garantia_hasta = data.get('garantia_hasta', '')
                    garantia_obj = None
                    if garantia_hasta:
                        try:
                            if isinstance(garantia_hasta, datetime):
                                garantia_obj = garantia_hasta.date()
                            else:
                                garantia_str = str(garantia_hasta).strip()
                                for fmt in ['%Y-%m-%d', '%d/%m/%Y', '%d-%m-%Y']:
                                    try:
                                        garantia_obj = datetime.strptime(garantia_str, fmt).date()
                                        break
                                    except ValueError:
                                        continue
                        except:
                            pass

                    # Obtener ambiente
                    ambiente = None
                    ambiente_codigo = get_str(data.get('ambiente_codigo'))
                    if ambiente_codigo:
                        try:
                            ambiente = Ambiente.objects.get(codigo=ambiente_codigo)
                        except Ambiente.DoesNotExist:
                            pass

                    # Parsear precio
                    precio = float(str(data.get('precio', 0)).replace(',', ''))

                    # Estado
                    estado = get_str(data.get('estado', 'nuevo')).lower()
                    if estado not in ['nuevo', 'instalado', 'dañado', 'obsoleto']:
                        estado = 'nuevo'

                    # Leasing
                    es_leasing_str = get_str(data.get('es_leasing', 'NO')).upper()
                    es_leasing = es_leasing_str in ['SI', 'SÍ', 'YES', 'S', 'Y', '1', 'TRUE']

                    # Crear ítem
                    item = Item.objects.create(
                        codigo_utp=codigo_utp,
                        serie=get_str(data.get('serie')),
                        nombre=get_str(data.get('nombre')),
                        descripcion=get_str(data.get('descripcion')),
                        area=area,
                        tipo_item=tipo_item,
                        ambiente=ambiente,
                        estado=estado,
                        precio=precio,
                        fecha_adquisicion=fecha_adq_obj,
                        garantia_hasta=garantia_obj,
                        observaciones=get_str(data.get('observaciones')),
                        lote=lote,
                        es_leasing=es_leasing,
                        leasing_empresa=get_str(data.get('leasing_empresa')) if es_leasing else '',
                        leasing_contrato=get_str(data.get('leasing_contrato')) if es_leasing else '',
                        creado_por=request.user
                    )

                    # Si es área de sistemas, crear especificaciones
                    if area.codigo == 'sistemas':
                        specs_data = {
                            'marca': get_str(data.get('marca')),
                            'modelo': get_str(data.get('modelo')),
                            'procesador': get_str(data.get('procesador')),
                            'generacion_procesador': get_str(data.get('generacion_procesador')),
                            'ram_total_gb': data.get('ram_total_gb', None),
                            'ram_configuracion': get_str(data.get('ram_configuracion')),
                            'ram_tipo': get_str(data.get('ram_tipo')),
                            'almacenamiento_gb': data.get('almacenamiento_gb', None),
                            'almacenamiento_tipo': get_str(data.get('almacenamiento_tipo')),
                            'sistema_operativo': get_str(data.get('sistema_operativo')),
                        }

                        # Limpiar valores None y vacíos
                        specs_data = {k: v for k, v in specs_data.items() if v}

                        if specs_data:
                            EspecificacionesSistemas.objects.create(
                                item=item,
                                **specs_data
                            )

                    items_creados.append(item)

                # Limpiar sesión
                del request.session['items_preview']
                if 'crear_lote' in request.session:
                    del request.session['crear_lote']
                if 'lote_descripcion' in request.session:
                    del request.session['lote_descripcion']
                if 'lote_existente_id' in request.session:
                    del request.session['lote_existente_id']

                # Mensaje de éxito
                mensaje = f'Se importaron exitosamente {len(items_creados)} ítems.'
                if lote:
                    mensaje += f' Asociados al lote {lote.codigo_interno}.'

                messages.success(request, mensaje)

                # AUDITORÍA: Registrar importación exitosa
                logger.info(
                    f'IMPORT_SUCCESS: User={request.user.username}, '
                    f'Items={len(items_creados)}, Area={perfil.area.codigo if perfil.area else "todas"}, '
                    f'IP={request.META.get("REMOTE_ADDR")}, Lote={lote.codigo_interno if lote else "N/A"}'
                )

                # Redirigir a la lista de ítems
                return redirect('productos:item-list')

        except Exception as e:
            # Log detallado para debugging
            logger.exception(f'Error al confirmar importación de {request.user.username}: {e}')
            messages.error(request, 'Ocurrió un error al importar los ítems. Por favor, contacte al administrador.')
            return redirect('productos:item-importar')





# ==============================================================================
# VISTAS DE REPORTES Y EXPORTACIÓN
# ==============================================================================

from .utils.export_utils import ExcelExporter, PDFExporter, format_currency, format_date, format_boolean


class ReportesView(LoginRequiredMixin, TemplateView):
    """Vista principal para seleccionar y generar reportes"""
    template_name = 'productos/reportes.html'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        perfil = self.request.user.perfil

        if perfil.area:
            items = Item.objects.filter(area=perfil.area)
        else:
            items = Item.objects.all()

        context['total_items'] = items.count()
        context['valor_total'] = items.aggregate(Sum('precio'))['precio__sum'] or 0
        context['items_operativos'] = items.filter(estado='operativo').count()
        context['items_mantenimiento'] = items.filter(estado='en_mantenimiento').count()

        if perfil.area:
            context['areas'] = [perfil.area]
        else:
            context['areas'] = Area.objects.filter(activo=True)

        context['tipos_item'] = TipoItem.objects.filter(activo=True)
        return context


class ExportarInventarioExcelView(RateLimitMixin, LoginRequiredMixin, View):
    """Exporta el inventario completo a Excel"""
    ratelimit_key = 'export'

    def get(self, request, *args, **kwargs):
        perfil = request.user.perfil
        area_id = request.GET.get('area')
        tipo_id = request.GET.get('tipo')
        estado = request.GET.get('estado')

        if perfil.area:
            items = Item.objects.filter(area=perfil.area)
        else:
            items = Item.objects.all()

        if area_id:
            items = items.filter(area_id=area_id)
        if tipo_id:
            items = items.filter(tipo_item_id=tipo_id)
        if estado:
            items = items.filter(estado=estado)

        exporter = ExcelExporter(title="Inventario")
        titulo = "INVENTARIO COMPLETO"
        subtitulo = f"Total de ítems: {items.count()}"
        if perfil.area:
            subtitulo += f" | Área: {perfil.area.nombre}"

        exporter.add_title(titulo, subtitulo)

        headers = ['Código Interno', 'Código UTP', 'Serie', 'Nombre', 'Área', 'Tipo', 'Estado',
                   'Ubicación', 'Usuario Asignado', 'Precio', 'Fecha Adquisición', 'Garantía Hasta', 'Leasing']
        exporter.add_headers(headers)

        for idx, item in enumerate(items.select_related('area', 'tipo_item', 'ambiente', 'usuario_asignado')):
            ubicacion = item.ambiente.codigo_completo if item.ambiente else 'Sin asignar'
            usuario = item.usuario_asignado.get_full_name() if item.usuario_asignado else 'Sin asignar'
            row = [item.codigo_interno, item.codigo_utp, item.serie, item.nombre, item.area.nombre,
                   item.tipo_item.nombre, item.get_estado_display(), ubicacion, usuario,
                   format_currency(item.precio), format_date(item.fecha_adquisicion),
                   format_date(item.garantia_hasta), format_boolean(item.es_leasing)]
            exporter.add_row(row, alternate=(idx % 2 == 0))

        valor_total = items.aggregate(Sum('precio'))['precio__sum'] or 0
        summary = {'Total de ítems': items.count(), 'Valor total': format_currency(valor_total),
                   'Ítems operativos': items.filter(estado='operativo').count(),
                   'Ítems en mantenimiento': items.filter(estado='en_mantenimiento').count(),
                   'Ítems con código UTP pendiente': items.filter(codigo_utp='PENDIENTE').count()}
        exporter.add_summary(summary)

        fecha = timezone.now().strftime('%Y%m%d_%H%M%S')
        return exporter.get_response(f"inventario_{fecha}.xlsx")


class ExportarReportePorAreaExcelView(RateLimitMixin, LoginRequiredMixin, View):
    """Exporta reporte de ítems agrupados por área"""
    ratelimit_key = 'export'

    def get(self, request, *args, **kwargs):
        perfil = request.user.perfil
        if perfil.area:
            areas = Area.objects.filter(id=perfil.area.id, activo=True)
        else:
            areas = Area.objects.filter(activo=True)

        exporter = ExcelExporter(title="Reporte por Área")
        exporter.add_title("REPORTE DE INVENTARIO POR ÁREA", "Distribución de ítems y valores")

        headers = ['Área', 'Cantidad de Ítems', 'Valor Total', '% del Total', 'Operativos', 'En Mantenimiento', 'Dañados']
        exporter.add_headers(headers)

        total_items = Item.objects.count()
        total_valor = Item.objects.aggregate(Sum('precio'))['precio__sum'] or 0

        for idx, area in enumerate(areas):
            items_area = Item.objects.filter(area=area)
            cantidad = items_area.count()
            valor = items_area.aggregate(Sum('precio'))['precio__sum'] or 0
            porcentaje = (valor / total_valor * 100) if total_valor > 0 else 0
            row = [area.nombre, cantidad, format_currency(valor), f"{porcentaje:.2f}%",
                   items_area.filter(estado='operativo').count(),
                   items_area.filter(estado='en_mantenimiento').count(),
                   items_area.filter(estado='danado').count()]
            exporter.add_row(row, alternate=(idx % 2 == 0))

        summary = {'Total general de ítems': total_items, 'Valor total general': format_currency(total_valor)}
        exporter.add_summary(summary)

        fecha = timezone.now().strftime('%Y%m%d_%H%M%S')
        return exporter.get_response(f"reporte_por_area_{fecha}.xlsx")


class ExportarGarantiasVencenExcelView(RateLimitMixin, LoginRequiredMixin, View):
    """Exporta reporte de garantías próximas a vencer"""
    ratelimit_key = 'export'

    def get(self, request, *args, **kwargs):
        perfil = request.user.perfil
        dias = int(request.GET.get('dias', 30))
        hoy = date.today()
        fecha_limite = hoy + timedelta(days=dias)

        if perfil.area:
            items = Item.objects.filter(area=perfil.area, garantia_hasta__gte=hoy, garantia_hasta__lte=fecha_limite)
        else:
            items = Item.objects.filter(garantia_hasta__gte=hoy, garantia_hasta__lte=fecha_limite)

        exporter = ExcelExporter(title="Garantías")
        exporter.add_title(f"GARANTÍAS QUE VENCEN EN {dias} DÍAS",
                          f"Del {hoy.strftime('%d/%m/%Y')} al {fecha_limite.strftime('%d/%m/%Y')}")

        headers = ['Código Interno', 'Serie', 'Nombre', 'Área', 'Fecha Adquisición',
                   'Garantía Hasta', 'Días Restantes', 'Precio', 'Proveedor/Lote']
        exporter.add_headers(headers)

        for idx, item in enumerate(items.select_related('area', 'lote')):
            dias_restantes = (item.garantia_hasta - hoy).days
            proveedor = item.lote.contrato.proveedor.nombre if (item.lote and item.lote.contrato) else 'N/A'
            row = [item.codigo_interno, item.serie, item.nombre, item.area.nombre,
                   format_date(item.fecha_adquisicion), format_date(item.garantia_hasta),
                   f"{dias_restantes} días", format_currency(item.precio), proveedor]
            exporter.add_row(row, alternate=(idx % 2 == 0))

        valor_total = items.aggregate(Sum('precio'))['precio__sum'] or 0
        summary = {'Total de ítems': items.count(), 'Valor total en riesgo': format_currency(valor_total)}
        exporter.add_summary(summary)

        fecha = timezone.now().strftime('%Y%m%d_%H%M%S')
        return exporter.get_response(f"garantias_vencen_{dias}dias_{fecha}.xlsx")


class ExportarInventarioPDFView(RateLimitMixin, LoginRequiredMixin, View):
    """Exporta el inventario completo a PDF"""
    ratelimit_key = 'export'

    def get(self, request, *args, **kwargs):
        perfil = request.user.perfil
        area_id = request.GET.get('area')
        tipo_id = request.GET.get('tipo')
        estado = request.GET.get('estado')

        if perfil.area:
            items = Item.objects.filter(area=perfil.area)
        else:
            items = Item.objects.all()

        if area_id:
            items = items.filter(area_id=area_id)
        if tipo_id:
            items = items.filter(tipo_item_id=tipo_id)
        if estado:
            items = items.filter(estado=estado)

        exporter = PDFExporter(title="Inventario", orientation="landscape")
        titulo = "INVENTARIO COMPLETO"
        subtitulo = f"Total de ítems: {items.count()}"
        if perfil.area:
            subtitulo += f" | Área: {perfil.area.nombre}"

        exporter.add_title(titulo, subtitulo)

        headers = ['Código', 'Serie', 'Nombre', 'Área', 'Tipo', 'Estado', 'Precio']
        data = []
        for item in items.select_related('area', 'tipo_item')[:100]:
            data.append([item.codigo_interno, item.serie[:15], item.nombre[:25], item.area.codigo,
                        item.tipo_item.nombre[:15], item.get_estado_display()[:10], format_currency(item.precio)])

        exporter.add_table(headers, data)

        valor_total = items.aggregate(Sum('precio'))['precio__sum'] or 0
        summary = {'Total de ítems': items.count(), 'Valor total': format_currency(valor_total),
                   'Ítems operativos': items.filter(estado='operativo').count(),
                   'Ítems en mantenimiento': items.filter(estado='en_mantenimiento').count()}
        exporter.add_summary_section(summary)

        fecha = timezone.now().strftime('%Y%m%d_%H%M%S')
        return exporter.get_response(f"inventario_{fecha}.pdf")


class ExportarReportePorAreaPDFView(RateLimitMixin, LoginRequiredMixin, View):
    """Exporta reporte de ítems agrupados por área a PDF"""
    ratelimit_key = 'export'

    def get(self, request, *args, **kwargs):
        perfil = request.user.perfil
        if perfil.area:
            areas = Area.objects.filter(id=perfil.area.id, activo=True)
        else:
            areas = Area.objects.filter(activo=True)

        exporter = PDFExporter(title="Reporte por Área")
        exporter.add_title("REPORTE DE INVENTARIO POR ÁREA", "Distribución de ítems y valores")

        total_items = Item.objects.count()
        total_valor = Item.objects.aggregate(Sum('precio'))['precio__sum'] or 0

        headers = ['Área', 'Cantidad', 'Valor Total', '% Total', 'Operativos', 'Mant.', 'Dañados']
        data = []
        for area in areas:
            items_area = Item.objects.filter(area=area)
            cantidad = items_area.count()
            valor = items_area.aggregate(Sum('precio'))['precio__sum'] or 0
            porcentaje = (valor / total_valor * 100) if total_valor > 0 else 0
            data.append([area.nombre[:20], str(cantidad), format_currency(valor), f"{porcentaje:.1f}%",
                        str(items_area.filter(estado='operativo').count()),
                        str(items_area.filter(estado='en_mantenimiento').count()),
                        str(items_area.filter(estado='danado').count())])

        exporter.add_table(headers, data)

        summary = {'Total general de ítems': total_items, 'Valor total general': format_currency(total_valor)}
        exporter.add_summary_section(summary)

        fecha = timezone.now().strftime('%Y%m%d_%H%M%S')
        return exporter.get_response(f"reporte_por_area_{fecha}.pdf")


# ==============================================================================
# VISTAS DE MANTENIMIENTO
# ==============================================================================

class MantenimientoListView(PerfilRequeridoMixin, ListView):
    """Lista de mantenimientos"""
    model = Mantenimiento
    template_name = 'productos/mantenimiento_list.html'
    context_object_name = 'mantenimientos'
    paginate_by = 20

    def get_queryset(self):
        queryset = Mantenimiento.objects.select_related('item', 'responsable', 'creado_por')
        perfil = getattr(self.request.user, 'perfil', None)

        # Filtrar por área si no es admin
        if perfil and perfil.rol != 'admin' and perfil.area:
            queryset = queryset.filter(item__area=perfil.area)

        # Filtros
        estado = self.request.GET.get('estado')
        tipo = self.request.GET.get('tipo')

        if estado:
            queryset = queryset.filter(estado=estado)
        if tipo:
            queryset = queryset.filter(tipo=tipo)

        return queryset.order_by('-fecha_programada')

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        queryset = self.get_queryset()

        # Estadísticas
        context['total_mantenimientos'] = queryset.count()
        context['pendientes'] = queryset.filter(estado='pendiente').count()
        context['en_proceso'] = queryset.filter(estado='en_proceso').count()
        context['completados'] = queryset.filter(estado='completado').count()

        return context


class MantenimientoDetailView(PerfilRequeridoMixin, DetailView):
    """Detalle de un mantenimiento"""
    model = Mantenimiento
    template_name = 'productos/mantenimiento_detail.html'
    context_object_name = 'mantenimiento'

    def get_queryset(self):
        queryset = super().get_queryset()
        perfil = getattr(self.request.user, 'perfil', None)

        if perfil and perfil.rol != 'admin' and perfil.area:
            queryset = queryset.filter(item__area=perfil.area)

        return queryset


class MantenimientoCreateView(PerfilRequeridoMixin, CreateView):
    """Crear un nuevo mantenimiento"""
    model = Mantenimiento
    form_class = MantenimientoForm
    template_name = 'productos/mantenimiento_form.html'
    success_url = reverse_lazy('productos:mantenimiento-list')

    def get_form_kwargs(self):
        kwargs = super().get_form_kwargs()
        kwargs['user'] = self.request.user
        return kwargs

    def form_valid(self, form):
        mantenimiento = form.save(commit=False)
        mantenimiento.creado_por = self.request.user
        mantenimiento.responsable = self.request.user
        mantenimiento.save()

        messages.success(self.request, f'Mantenimiento programado correctamente para {mantenimiento.item.codigo_interno}.')
        return redirect('productos:mantenimiento-detail', pk=mantenimiento.pk)


class MantenimientoUpdateView(PerfilRequeridoMixin, UpdateView):
    """Editar un mantenimiento"""
    model = Mantenimiento
    form_class = MantenimientoForm
    template_name = 'productos/mantenimiento_form.html'

    def get_queryset(self):
        queryset = super().get_queryset()
        perfil = getattr(self.request.user, 'perfil', None)

        if perfil and perfil.rol != 'admin' and perfil.area:
            queryset = queryset.filter(item__area=perfil.area)

        return queryset

    def get_form_kwargs(self):
        kwargs = super().get_form_kwargs()
        kwargs['user'] = self.request.user
        return kwargs

    def form_valid(self, form):
        mantenimiento = form.save()
        messages.success(self.request, f'Mantenimiento actualizado correctamente.')
        return redirect('productos:mantenimiento-detail', pk=mantenimiento.pk)


class MantenimientoIniciarView(PerfilRequeridoMixin, View):
    """Iniciar un mantenimiento"""

    def post(self, request, pk):
        mantenimiento = get_object_or_404(Mantenimiento, pk=pk)
        perfil = request.user.perfil

        # Verificar permisos
        if perfil.rol != 'admin' and perfil.area and mantenimiento.item.area != perfil.area:
            messages.error(request, 'No tienes permisos para iniciar este mantenimiento.')
            return redirect('productos:mantenimiento-detail', pk=pk)

        if mantenimiento.estado != 'pendiente':
            messages.warning(request, 'Este mantenimiento ya fue iniciado.')
            return redirect('productos:mantenimiento-detail', pk=pk)

        # Cambiar estado del ítem a en_mantenimiento
        mantenimiento.item.estado = 'en_mantenimiento'
        mantenimiento.item.save()

        # Iniciar mantenimiento
        mantenimiento.iniciar(usuario=request.user)

        messages.success(request, f'Mantenimiento iniciado. El ítem {mantenimiento.item.codigo_interno} está ahora en mantenimiento.')
        return redirect('productos:mantenimiento-detail', pk=pk)


class MantenimientoFinalizarView(PerfilRequeridoMixin, View):
    """Finalizar un mantenimiento"""
    template_name = 'productos/mantenimiento_finalizar.html'

    def get(self, request, pk):
        mantenimiento = get_object_or_404(Mantenimiento, pk=pk)
        perfil = request.user.perfil

        # Verificar permisos
        if perfil.rol != 'admin' and perfil.area and mantenimiento.item.area != perfil.area:
            messages.error(request, 'No tienes permisos para finalizar este mantenimiento.')
            return redirect('productos:mantenimiento-detail', pk=pk)

        if mantenimiento.estado not in ['pendiente', 'en_proceso']:
            messages.warning(request, 'Este mantenimiento ya fue finalizado o cancelado.')
            return redirect('productos:mantenimiento-detail', pk=pk)

        form = MantenimientoFinalizarForm()
        return render(request, self.template_name, {
            'mantenimiento': mantenimiento,
            'form': form
        })

    def post(self, request, pk):
        mantenimiento = get_object_or_404(Mantenimiento, pk=pk)
        perfil = request.user.perfil

        # Verificar permisos
        if perfil.rol != 'admin' and perfil.area and mantenimiento.item.area != perfil.area:
            messages.error(request, 'No tienes permisos para finalizar este mantenimiento.')
            return redirect('productos:mantenimiento-detail', pk=pk)

        form = MantenimientoFinalizarForm(request.POST)
        if form.is_valid():
            resultado = form.cleaned_data['resultado']
            trabajo_realizado = form.cleaned_data['trabajo_realizado']
            costo = form.cleaned_data.get('costo')

            mantenimiento.finalizar(resultado, trabajo_realizado, costo)

            messages.success(request, f'Mantenimiento finalizado. Estado del ítem actualizado a {mantenimiento.item.get_estado_display()}.')
            return redirect('productos:mantenimiento-detail', pk=pk)

        return render(request, self.template_name, {
            'mantenimiento': mantenimiento,
            'form': form
        })


class MantenimientoCancelarView(PerfilRequeridoMixin, View):
    """Cancelar un mantenimiento"""

    def post(self, request, pk):
        mantenimiento = get_object_or_404(Mantenimiento, pk=pk)
        perfil = request.user.perfil

        # Verificar permisos
        if perfil.rol not in ['admin', 'supervisor']:
            messages.error(request, 'Solo administradores y supervisores pueden cancelar mantenimientos.')
            return redirect('productos:mantenimiento-detail', pk=pk)

        if perfil.rol != 'admin' and perfil.area and mantenimiento.item.area != perfil.area:
            messages.error(request, 'No tienes permisos para cancelar este mantenimiento.')
            return redirect('productos:mantenimiento-detail', pk=pk)

        if mantenimiento.estado in ['completado', 'cancelado']:
            messages.warning(request, 'Este mantenimiento ya fue completado o cancelado.')
            return redirect('productos:mantenimiento-detail', pk=pk)

        motivo = request.POST.get('motivo', '')
        mantenimiento.cancelar(motivo)

        # Si el ítem estaba en mantenimiento, regresarlo a operativo
        if mantenimiento.item.estado == 'en_mantenimiento':
            mantenimiento.item.estado = 'operativo'
            mantenimiento.item.save()

        messages.success(request, 'Mantenimiento cancelado.')
        return redirect('productos:mantenimiento-detail', pk=pk)


class MantenimientoDeleteView(AdminRequeridoMixin, DeleteView):
    """Eliminar un mantenimiento (solo admin)"""
    model = Mantenimiento
    template_name = 'productos/mantenimiento_confirm_delete.html'
    success_url = reverse_lazy('productos:mantenimiento-list')

    def delete(self, request, *args, **kwargs):
        mantenimiento = self.get_object()
        messages.success(request, f'Mantenimiento eliminado.')
        return super().delete(request, *args, **kwargs)


class MantenimientoLoteView(PerfilRequeridoMixin, View):
    """Crear mantenimientos en lote para múltiples ítems"""
    template_name = 'productos/mantenimiento_lote.html'

    def get(self, request):
        # Si vienen items pre-seleccionados por query params
        items_ids = request.GET.getlist('items')
        
        form = MantenimientoLoteForm(user=request.user)
        
        # Pre-seleccionar items si vienen en query params
        if items_ids:
            form.fields['items'].initial = items_ids
        
        # Obtener items disponibles para contexto
        perfil = request.user.perfil
        if perfil.area and perfil.rol != 'admin':
            items = Item.objects.filter(area=perfil.area)
        else:
            items = Item.objects.all()
        
        return render(request, self.template_name, {
            'form': form,
            'items': items,
            'items_preseleccionados': len(items_ids) if items_ids else 0
        })

    def post(self, request):
        form = MantenimientoLoteForm(request.POST, user=request.user)
        
        if form.is_valid():
            items = form.cleaned_data['items']
            tipo = form.cleaned_data['tipo']
            fecha_programada = form.cleaned_data['fecha_programada']
            descripcion = form.cleaned_data.get('descripcion_problema', '')
            tecnico = form.cleaned_data.get('tecnico_asignado', '')
            proveedor = form.cleaned_data.get('proveedor_servicio', '')
            costo = form.cleaned_data.get('costo_estimado')
            proximo = form.cleaned_data.get('proximo_mantenimiento')
            observaciones = form.cleaned_data.get('observaciones', '')
            
            # Crear mantenimientos para cada ítem seleccionado
            mantenimientos_creados = []
            for item in items:
                mantenimiento = Mantenimiento.objects.create(
                    item=item,
                    tipo=tipo,
                    fecha_programada=fecha_programada,
                    descripcion_problema=descripcion,
                    tecnico_asignado=tecnico,
                    proveedor_servicio=proveedor,
                    costo=costo,
                    proximo_mantenimiento=proximo,
                    observaciones=observaciones,
                    responsable=request.user,
                    creado_por=request.user
                )
                mantenimientos_creados.append(mantenimiento)
            
            messages.success(
                request,
                f'Se programaron {len(mantenimientos_creados)} mantenimientos correctamente.'
            )
            return redirect('productos:mantenimiento-list')
        
        # Si hay errores, volver a mostrar el formulario
        perfil = request.user.perfil
        if perfil.area and perfil.rol != 'admin':
            items = Item.objects.filter(area=perfil.area)
        else:
            items = Item.objects.all()
        
        return render(request, self.template_name, {
            'form': form,
            'items': items
        })


# ==============================================================================
# SISTEMA DE ACTAS DE ENTREGA/DEVOLUCIÓN
# ==============================================================================

from .models import Gerencia, Colaborador, SoftwareEstandar, ActaEntrega, ActaItem, ActaFoto, ActaSoftware
from .forms import (
    GerenciaForm, ColaboradorForm, SoftwareEstandarForm, ActaEntregaForm,
    ActaItemFormSet, ActaFotoFormSet, FirmaForm, SeleccionarSoftwareForm
)


class GerenciaListView(PerfilRequeridoMixin, ListView):
    """Lista de gerencias."""
    model = Gerencia
    template_name = 'productos/gerencia_list.html'
    context_object_name = 'gerencias'

    def get_queryset(self):
        return Gerencia.objects.all().order_by('nombre')


class GerenciaCreateView(PerfilRequeridoMixin, CreateView):
    """Crear nueva gerencia."""
    model = Gerencia
    form_class = GerenciaForm
    template_name = 'productos/gerencia_form.html'
    success_url = '/productos/gerencias/'

    def form_valid(self, form):
        messages.success(self.request, 'Gerencia creada correctamente.')
        return super().form_valid(form)


class GerenciaUpdateView(PerfilRequeridoMixin, UpdateView):
    """Editar gerencia."""
    model = Gerencia
    form_class = GerenciaForm
    template_name = 'productos/gerencia_form.html'
    success_url = '/productos/gerencias/'

    def form_valid(self, form):
        messages.success(self.request, 'Gerencia actualizada correctamente.')
        return super().form_valid(form)


class ColaboradorListView(PerfilRequeridoMixin, ListView):
    """Lista de colaboradores."""
    model = Colaborador
    template_name = 'productos/colaborador_list.html'
    context_object_name = 'colaboradores'
    paginate_by = 20

    def get_queryset(self):
        queryset = Colaborador.objects.select_related('gerencia', 'sede').all()

        # Filtros
        buscar = self.request.GET.get('buscar', '')
        gerencia = self.request.GET.get('gerencia', '')
        sede = self.request.GET.get('sede', '')
        activo = self.request.GET.get('activo', '')

        if buscar:
            queryset = queryset.filter(
                models.Q(nombre_completo__icontains=buscar) |
                models.Q(dni__icontains=buscar) |
                models.Q(correo__icontains=buscar)
            )

        if gerencia:
            queryset = queryset.filter(gerencia_id=gerencia)

        if sede:
            queryset = queryset.filter(sede_id=sede)

        if activo == '1':
            queryset = queryset.filter(activo=True)
        elif activo == '0':
            queryset = queryset.filter(activo=False)

        return queryset.order_by('nombre_completo')

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['gerencias'] = Gerencia.objects.filter(activo=True)
        context['sedes'] = Sede.objects.filter(activo=True)
        context['filtros'] = {
            'buscar': self.request.GET.get('buscar', ''),
            'gerencia': self.request.GET.get('gerencia', ''),
            'sede': self.request.GET.get('sede', ''),
            'activo': self.request.GET.get('activo', ''),
        }
        return context


class ColaboradorCreateView(PerfilRequeridoMixin, CreateView):
    """Crear nuevo colaborador."""
    model = Colaborador
    form_class = ColaboradorForm
    template_name = 'productos/colaborador_form.html'
    success_url = '/productos/colaboradores/'

    def form_valid(self, form):
        form.instance.creado_por = self.request.user
        messages.success(self.request, 'Colaborador creado correctamente.')
        return super().form_valid(form)


class ColaboradorUpdateView(PerfilRequeridoMixin, UpdateView):
    """Editar colaborador."""
    model = Colaborador
    form_class = ColaboradorForm
    template_name = 'productos/colaborador_form.html'
    success_url = '/productos/colaboradores/'

    def form_valid(self, form):
        messages.success(self.request, 'Colaborador actualizado correctamente.')
        return super().form_valid(form)


class ColaboradorDetailView(PerfilRequeridoMixin, DetailView):
    """Detalle de colaborador con sus ítems asignados."""
    model = Colaborador
    template_name = 'productos/colaborador_detail.html'
    context_object_name = 'colaborador'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['items_asignados'] = Item.objects.filter(
            colaborador_asignado=self.object
        ).select_related('tipo_item', 'area')
        context['actas'] = ActaEntrega.objects.filter(
            colaborador=self.object
        ).order_by('-fecha')[:10]
        return context


class BuscarColaboradorView(PerfilRequeridoMixin, View):
    """API para buscar colaborador por DNI (AJAX)."""

    def get(self, request):
        dni = request.GET.get('dni', '').strip()

        if not dni:
            return JsonResponse({'error': 'DNI requerido'}, status=400)

        try:
            colaborador = Colaborador.objects.select_related('gerencia', 'sede').get(dni=dni)
            return JsonResponse({
                'id': colaborador.id,
                'dni': colaborador.dni,
                'nombre_completo': colaborador.nombre_completo,
                'cargo': colaborador.cargo,
                'gerencia': colaborador.gerencia.nombre,
                'sede': str(colaborador.sede),
                'anexo': colaborador.anexo,
                'correo': colaborador.correo,
                'items_asignados': colaborador.cantidad_items_asignados,
            })
        except Colaborador.DoesNotExist:
            return JsonResponse({'error': 'Colaborador no encontrado'}, status=404)


class ActaListView(PerfilRequeridoMixin, CampusFilterMixin, ListView):
    """Lista de actas de entrega/devolución según permisos de campus."""
    model = ActaEntrega
    template_name = 'productos/acta_list.html'
    context_object_name = 'actas'
    paginate_by = 20

    def get_queryset(self):
        queryset = ActaEntrega.objects.select_related(
            'colaborador', 'creado_por'
        ).prefetch_related('items')

        # Filtrar por campus - solo actas que contengan items del campus del usuario
        perfil = getattr(self.request.user, 'perfil', None)
        if perfil and perfil.rol != 'admin':
            campus_ids = list(self.get_campus_permitidos().values_list('id', flat=True))
            # Obtener actas que tienen al menos un item en los campus permitidos
            actas_con_items_permitidos = ActaItem.objects.filter(
                item__ambiente__pabellon__sede__campus_id__in=campus_ids
            ).values_list('acta_id', flat=True).distinct()
            queryset = queryset.filter(id__in=actas_con_items_permitidos)

        # Filtros
        tipo = self.request.GET.get('tipo', '')
        buscar = self.request.GET.get('buscar', '')
        fecha_desde = self.request.GET.get('fecha_desde', '')
        fecha_hasta = self.request.GET.get('fecha_hasta', '')

        if tipo:
            queryset = queryset.filter(tipo=tipo)

        if buscar:
            queryset = queryset.filter(
                models.Q(numero_acta__icontains=buscar) |
                models.Q(colaborador__nombre_completo__icontains=buscar) |
                models.Q(colaborador__dni__icontains=buscar)
            )

        if fecha_desde:
            queryset = queryset.filter(fecha__date__gte=fecha_desde)

        if fecha_hasta:
            queryset = queryset.filter(fecha__date__lte=fecha_hasta)

        return queryset.order_by('-fecha')

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['filtros'] = {
            'tipo': self.request.GET.get('tipo', ''),
            'buscar': self.request.GET.get('buscar', ''),
            'fecha_desde': self.request.GET.get('fecha_desde', ''),
            'fecha_hasta': self.request.GET.get('fecha_hasta', ''),
        }
        return context


class ActaDetailView(PerfilRequeridoMixin, DetailView):
    """Detalle de un acta."""
    model = ActaEntrega
    template_name = 'productos/acta_detail.html'
    context_object_name = 'acta'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['acta_items'] = self.object.items.select_related(
            'item', 'item__tipo_item'
        ).all()
        context['acta_software'] = self.object.software.select_related('software').all()
        context['acta_fotos'] = self.object.fotos.all()
        return context


class ActaCreateView(PerfilRequeridoMixin, View):
    """Crear nueva acta de entrega/devolución - Wizard de múltiples pasos."""
    template_name = 'productos/acta_create.html'

    def get(self, request):
        # Paso inicial: seleccionar tipo y colaborador
        form = ActaEntregaForm(user=request.user)
        software_form = SeleccionarSoftwareForm()

        return render(request, self.template_name, {
            'form': form,
            'software_form': software_form,
            'paso': 1,
        })

    def post(self, request):
        paso = request.POST.get('paso', '1')

        if paso == '1':
            # Validar tipo y colaborador, mostrar selección de ítems
            form = ActaEntregaForm(request.POST, user=request.user)

            if form.is_valid():
                tipo = form.cleaned_data['tipo']
                colaborador = form.cleaned_data['colaborador']

                # Obtener ítems disponibles según el tipo
                if tipo == 'entrega':
                    # Para entrega: items sin asignar y en buen estado (nuevo/instalado)
                    items_disponibles = Item.objects.filter(
                        colaborador_asignado__isnull=True,
                        estado__in=['nuevo', 'instalado']
                    ).select_related('tipo_item', 'area')
                else:
                    # Para devolución: items asignados al colaborador
                    items_disponibles = Item.objects.filter(
                        colaborador_asignado=colaborador
                    ).select_related('tipo_item', 'area')

                # Filtrar por área si no es admin
                perfil = request.user.perfil
                if perfil.rol != 'admin' and perfil.area:
                    items_disponibles = items_disponibles.filter(area=perfil.area)

                # Guardar datos en sesión
                request.session['acta_tipo'] = tipo
                request.session['acta_colaborador_id'] = colaborador.id
                request.session['acta_ticket'] = form.cleaned_data.get('ticket', '')
                request.session['acta_observaciones'] = form.cleaned_data.get('observaciones', '')

                software_form = SeleccionarSoftwareForm()

                return render(request, self.template_name, {
                    'form': form,
                    'software_form': software_form,
                    'items_disponibles': items_disponibles,
                    'colaborador': colaborador,
                    'tipo': tipo,
                    'paso': 2,
                })

            software_form = SeleccionarSoftwareForm()
            return render(request, self.template_name, {
                'form': form,
                'software_form': software_form,
                'paso': 1,
            })

        elif paso == '2':
            # Validar ítems seleccionados, mostrar formulario de accesorios
            items_ids = request.POST.getlist('items')
            software_ids = request.POST.getlist('software')

            if not items_ids:
                messages.error(request, 'Debe seleccionar al menos un ítem.')
                return redirect('productos:acta-create')

            # Guardar en sesión
            request.session['acta_items_ids'] = items_ids
            request.session['acta_software_ids'] = software_ids

            items = Item.objects.filter(id__in=items_ids).select_related('tipo_item')
            colaborador_id = request.session.get('acta_colaborador_id')
            colaborador = Colaborador.objects.get(id=colaborador_id)

            return render(request, self.template_name, {
                'items': items,
                'colaborador': colaborador,
                'tipo': request.session.get('acta_tipo'),
                'paso': 3,
            })

        elif paso == '3':
            # Validar accesorios y firmas, crear el acta
            import base64
            from django.core.files.base import ContentFile

            firma_receptor_data = request.POST.get('firma_receptor')
            firma_emisor_data = request.POST.get('firma_emisor')

            if not firma_receptor_data or not firma_emisor_data:
                messages.error(request, 'Ambas firmas son obligatorias.')
                return redirect('productos:acta-create')

            # Recuperar datos de sesión
            tipo = request.session.get('acta_tipo')
            colaborador_id = request.session.get('acta_colaborador_id')
            ticket = request.session.get('acta_ticket', '')
            observaciones = request.session.get('acta_observaciones', '')
            items_ids = request.session.get('acta_items_ids', [])
            software_ids = request.session.get('acta_software_ids', [])

            colaborador = Colaborador.objects.get(id=colaborador_id)

            # Procesar firmas (base64 a imagen)
            def base64_to_image(data, filename):
                if ',' in data:
                    data = data.split(',')[1]
                image_data = base64.b64decode(data)
                return ContentFile(image_data, name=filename)

            firma_receptor_file = base64_to_image(
                firma_receptor_data,
                f'firma_receptor_{colaborador.dni}.png'
            )
            firma_emisor_file = base64_to_image(
                firma_emisor_data,
                f'firma_emisor_{request.user.username}.png'
            )

            # Crear el acta
            acta = ActaEntrega.objects.create(
                tipo=tipo,
                colaborador=colaborador,
                ticket=ticket,
                observaciones=observaciones,
                firma_receptor=firma_receptor_file,
                firma_emisor=firma_emisor_file,
                creado_por=request.user
            )

            # Crear ActaItems con accesorios
            items = Item.objects.filter(id__in=items_ids)
            for item in items:
                ActaItem.objects.create(
                    acta=acta,
                    item=item,
                    acc_cargador=request.POST.get(f'acc_cargador_{item.id}') == 'on',
                    acc_cable_seguridad=request.POST.get(f'acc_cable_seguridad_{item.id}') == 'on',
                    acc_bateria=request.POST.get(f'acc_bateria_{item.id}') == 'on',
                    acc_maletin=request.POST.get(f'acc_maletin_{item.id}') == 'on',
                    acc_cable_red=request.POST.get(f'acc_cable_red_{item.id}') == 'on',
                    acc_teclado_mouse=request.POST.get(f'acc_teclado_mouse_{item.id}') == 'on',
                )

                # Actualizar asignación del ítem
                if tipo == 'entrega':
                    item.colaborador_asignado = colaborador
                else:
                    item.colaborador_asignado = None
                item.save()

            # Crear ActaSoftware
            for software_id in software_ids:
                ActaSoftware.objects.create(
                    acta=acta,
                    software_id=software_id
                )

            # Procesar fotos si hay
            fotos = request.FILES.getlist('fotos')
            for foto in fotos:
                ActaFoto.objects.create(
                    acta=acta,
                    foto=foto
                )

            # Limpiar sesión
            for key in ['acta_tipo', 'acta_colaborador_id', 'acta_ticket',
                       'acta_observaciones', 'acta_items_ids', 'acta_software_ids']:
                request.session.pop(key, None)

            messages.success(
                request,
                f'Acta {acta.numero_acta} creada correctamente.'
            )

            return redirect('productos:acta-detail', pk=acta.pk)

        return redirect('productos:acta-create')


class ActaDescargarPDFView(PerfilRequeridoMixin, View):
    """Descargar PDF del acta."""

    def get(self, request, pk):
        acta = get_object_or_404(ActaEntrega, pk=pk)

        # Si ya tiene PDF generado, devolverlo
        if acta.pdf_archivo:
            response = HttpResponse(acta.pdf_archivo, content_type='application/pdf')
            response['Content-Disposition'] = f'attachment; filename="{acta.numero_acta}.pdf"'
            return response

        # Si no, generar el PDF
        from .utils.acta_pdf import generar_acta_pdf

        try:
            pdf_buffer = generar_acta_pdf(acta)
            pdf_bytes = pdf_buffer.getvalue()

            # Guardar el PDF en el modelo
            from django.core.files.base import ContentFile
            acta.pdf_archivo.save(
                f'{acta.numero_acta}.pdf',
                ContentFile(pdf_bytes),
                save=True
            )

            response = HttpResponse(pdf_bytes, content_type='application/pdf')
            response['Content-Disposition'] = f'attachment; filename="{acta.numero_acta}.pdf"'
            return response

        except Exception as e:
            messages.error(request, f'Error al generar PDF: {str(e)}')
            return redirect('productos:acta-detail', pk=pk)


class ActaEnviarCorreoView(PerfilRequeridoMixin, View):
    """Enviar acta por correo al colaborador."""

    def post(self, request, pk):
        acta = get_object_or_404(ActaEntrega, pk=pk)

        from .utils.acta_email import enviar_acta_por_correo
        from .utils.acta_pdf import generar_acta_pdf

        try:
            # Generar PDF si no existe
            if not acta.pdf_archivo:
                from django.core.files.base import ContentFile

                pdf_buffer = generar_acta_pdf(acta)
                pdf_bytes = pdf_buffer.getvalue()
                acta.pdf_archivo.save(
                    f'{acta.numero_acta}.pdf',
                    ContentFile(pdf_bytes),
                    save=True
                )

            # Leer el PDF para enviarlo
            acta.pdf_archivo.seek(0)
            pdf_bytes = acta.pdf_archivo.read()

            # Enviar correo
            enviar_acta_por_correo(acta, pdf_bytes)

            # Marcar como enviado
            acta.correo_enviado = True
            acta.fecha_envio_correo = timezone.now()
            acta.save()

            messages.success(
                request,
                f'Acta enviada por correo a {acta.colaborador.correo}'
            )

        except Exception as e:
            messages.error(request, f'Error al enviar correo: {str(e)}')

        return redirect('productos:acta-detail', pk=pk)


class SoftwareEstandarListView(PerfilRequeridoMixin, ListView):
    """Lista de software estándar."""
    model = SoftwareEstandar
    template_name = 'productos/software_list.html'
    context_object_name = 'software_list'


class SoftwareEstandarCreateView(PerfilRequeridoMixin, CreateView):
    """Crear nuevo software estándar."""
    model = SoftwareEstandar
    form_class = SoftwareEstandarForm
    template_name = 'productos/software_form.html'
    success_url = '/productos/software/'

    def form_valid(self, form):
        messages.success(self.request, 'Software agregado correctamente.')
        return super().form_valid(form)


class SoftwareEstandarUpdateView(PerfilRequeridoMixin, UpdateView):
    """Editar software estándar."""
    model = SoftwareEstandar
    form_class = SoftwareEstandarForm
    template_name = 'productos/software_form.html'
    success_url = '/productos/software/'

    def form_valid(self, form):
        messages.success(self.request, 'Software actualizado correctamente.')
        return super().form_valid(form)
