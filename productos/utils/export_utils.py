"""
Utilidades para exportación de datos a Excel y PDF
Sistema de Inventario UTP
"""

from datetime import datetime
from io import BytesIO
from django.http import HttpResponse
from django.utils import timezone

# Excel
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# PDF
from reportlab.lib import colors
from reportlab.lib.pagesizes import letter, A4, landscape
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, PageBreak
from reportlab.platypus import Image as RLImage
from reportlab.lib.enums import TA_CENTER, TA_LEFT, TA_RIGHT


class ExcelExporter:
    """Clase para exportar datos a Excel con estilos personalizados"""

    # Colores UTP
    COLOR_HEADER = 'C8102E'  # Rojo UTP
    COLOR_HEADER_TEXT = 'FFFFFF'  # Blanco
    COLOR_SUBHEADER = 'F0F0F0'  # Gris claro
    COLOR_ROW_ALTERNATE = 'F9F9F9'  # Gris muy claro

    def __init__(self, title="Reporte"):
        self.wb = Workbook()
        self.ws = self.wb.active
        self.ws.title = title[:31]  # Excel limita a 31 caracteres
        self.current_row = 1

    def add_title(self, title, subtitle=None):
        """Agrega título principal al reporte"""
        # Título
        self.ws.cell(row=self.current_row, column=1, value=title)
        title_cell = self.ws.cell(row=self.current_row, column=1)
        title_cell.font = Font(size=16, bold=True, color=self.COLOR_HEADER)
        self.current_row += 1

        # Subtítulo
        if subtitle:
            self.ws.cell(row=self.current_row, column=1, value=subtitle)
            subtitle_cell = self.ws.cell(row=self.current_row, column=1)
            subtitle_cell.font = Font(size=12, color='666666')
            self.current_row += 1

        # Fecha de generación
        fecha_generacion = f"Generado: {timezone.now().strftime('%d/%m/%Y %H:%M')}"
        self.ws.cell(row=self.current_row, column=1, value=fecha_generacion)
        fecha_cell = self.ws.cell(row=self.current_row, column=1)
        fecha_cell.font = Font(size=10, italic=True, color='999999')
        self.current_row += 2

    def add_headers(self, headers):
        """Agrega fila de encabezados con estilo"""
        for col_num, header in enumerate(headers, 1):
            cell = self.ws.cell(row=self.current_row, column=col_num, value=header)
            cell.font = Font(bold=True, color=self.COLOR_HEADER_TEXT)
            cell.fill = PatternFill(start_color=self.COLOR_HEADER, end_color=self.COLOR_HEADER, fill_type='solid')
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            cell.border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
        self.current_row += 1

    def add_row(self, values, alternate=False):
        """Agrega una fila de datos"""
        for col_num, value in enumerate(values, 1):
            cell = self.ws.cell(row=self.current_row, column=col_num, value=value)
            cell.alignment = Alignment(horizontal='left', vertical='center')
            cell.border = Border(
                left=Side(style='thin', color='CCCCCC'),
                right=Side(style='thin', color='CCCCCC'),
                top=Side(style='thin', color='CCCCCC'),
                bottom=Side(style='thin', color='CCCCCC')
            )
            if alternate:
                cell.fill = PatternFill(start_color=self.COLOR_ROW_ALTERNATE,
                                       end_color=self.COLOR_ROW_ALTERNATE,
                                       fill_type='solid')
        self.current_row += 1

    def add_summary(self, summary_data):
        """Agrega sección de resumen al final"""
        self.current_row += 1
        for label, value in summary_data.items():
            self.ws.cell(row=self.current_row, column=1, value=label)
            label_cell = self.ws.cell(row=self.current_row, column=1)
            label_cell.font = Font(bold=True)

            self.ws.cell(row=self.current_row, column=2, value=value)
            value_cell = self.ws.cell(row=self.current_row, column=2)
            value_cell.font = Font(color=self.COLOR_HEADER)

            self.current_row += 1

    def auto_adjust_columns(self):
        """Ajusta automáticamente el ancho de las columnas"""
        for column_cells in self.ws.columns:
            length = max(len(str(cell.value or '')) for cell in column_cells)
            self.ws.column_dimensions[get_column_letter(column_cells[0].column)].width = min(length + 2, 50)

    def get_response(self, filename="reporte.xlsx"):
        """Genera HttpResponse con el archivo Excel"""
        self.auto_adjust_columns()

        output = BytesIO()
        self.wb.save(output)
        output.seek(0)

        response = HttpResponse(
            output.read(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        return response


class PDFExporter:
    """Clase para exportar datos a PDF con estilos personalizados"""

    def __init__(self, title="Reporte", orientation="portrait"):
        self.title = title
        self.orientation = orientation
        self.elements = []
        self.styles = getSampleStyleSheet()

        # Crear estilos personalizados
        self.styles.add(ParagraphStyle(
            name='CustomTitle',
            parent=self.styles['Heading1'],
            fontSize=18,
            textColor=colors.HexColor('#C8102E'),
            spaceAfter=12,
            alignment=TA_CENTER
        ))

        self.styles.add(ParagraphStyle(
            name='CustomSubtitle',
            parent=self.styles['Normal'],
            fontSize=12,
            textColor=colors.grey,
            spaceAfter=6,
            alignment=TA_CENTER
        ))

        self.styles.add(ParagraphStyle(
            name='Footer',
            parent=self.styles['Normal'],
            fontSize=8,
            textColor=colors.grey,
            alignment=TA_CENTER
        ))

    def add_title(self, title, subtitle=None):
        """Agrega título al PDF"""
        self.elements.append(Paragraph(title, self.styles['CustomTitle']))
        if subtitle:
            self.elements.append(Paragraph(subtitle, self.styles['CustomSubtitle']))

        # Fecha de generación
        fecha = timezone.now().strftime('%d/%m/%Y %H:%M')
        self.elements.append(Paragraph(f"Generado: {fecha}", self.styles['Footer']))
        self.elements.append(Spacer(1, 0.3*inch))

    def add_table(self, headers, data, column_widths=None):
        """Agrega una tabla al PDF"""
        # Preparar datos de la tabla
        table_data = [headers] + data

        # Crear tabla
        if column_widths:
            table = Table(table_data, colWidths=column_widths)
        else:
            table = Table(table_data)

        # Estilo de la tabla
        table.setStyle(TableStyle([
            # Encabezado
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#C8102E')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 10),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),

            # Contenido
            ('BACKGROUND', (0, 1), (-1, -1), colors.white),
            ('TEXTCOLOR', (0, 1), (-1, -1), colors.black),
            ('ALIGN', (0, 1), (-1, -1), 'LEFT'),
            ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
            ('FONTSIZE', (0, 1), (-1, -1), 9),
            ('TOPPADDING', (0, 1), (-1, -1), 6),
            ('BOTTOMPADDING', (0, 1), (-1, -1), 6),

            # Bordes
            ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),

            # Filas alternadas
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#F9F9F9')])
        ]))

        self.elements.append(table)
        self.elements.append(Spacer(1, 0.2*inch))

    def add_summary_section(self, summary_data):
        """Agrega sección de resumen"""
        self.elements.append(Spacer(1, 0.2*inch))
        self.elements.append(Paragraph("<b>RESUMEN</b>", self.styles['Heading2']))
        self.elements.append(Spacer(1, 0.1*inch))

        summary_table_data = [[k, str(v)] for k, v in summary_data.items()]
        summary_table = Table(summary_table_data, colWidths=[3*inch, 2*inch])
        summary_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (0, -1), colors.HexColor('#F0F0F0')),
            ('TEXTCOLOR', (0, 0), (-1, -1), colors.black),
            ('ALIGN', (0, 0), (0, -1), 'RIGHT'),
            ('ALIGN', (1, 0), (1, -1), 'LEFT'),
            ('FONTNAME', (0, 0), (0, -1), 'Helvetica-Bold'),
            ('FONTNAME', (1, 0), (1, -1), 'Helvetica'),
            ('FONTSIZE', (0, 0), (-1, -1), 10),
            ('TOPPADDING', (0, 0), (-1, -1), 8),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 8),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
        ]))

        self.elements.append(summary_table)

    def add_page_break(self):
        """Agrega salto de página"""
        self.elements.append(PageBreak())

    def get_response(self, filename="reporte.pdf"):
        """Genera HttpResponse con el archivo PDF"""
        buffer = BytesIO()

        # Configurar página
        if self.orientation == "landscape":
            pagesize = landscape(A4)
        else:
            pagesize = A4

        # Crear documento
        doc = SimpleDocTemplate(
            buffer,
            pagesize=pagesize,
            rightMargin=0.5*inch,
            leftMargin=0.5*inch,
            topMargin=0.5*inch,
            bottomMargin=0.5*inch
        )

        # Generar PDF
        doc.build(self.elements)
        buffer.seek(0)

        response = HttpResponse(buffer.read(), content_type='application/pdf')
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        return response


# Funciones auxiliares para formateo
def format_currency(value):
    """Formatea valor como moneda"""
    if value is None:
        return "$0.00"
    return f"${value:,.2f}"


def format_date(date):
    """Formatea fecha"""
    if date is None:
        return "N/A"
    return date.strftime('%d/%m/%Y')


def format_boolean(value):
    """Formatea booleano"""
    return "Sí" if value else "No"


def generar_formato_traslado(items_data, origen_data, destino_data, fecha=None):
    """
    Genera el formato de traslado en Excel usando la plantilla oficial UTP.

    Args:
        items_data: Lista de diccionarios con {codigo_utp, descripcion, marca, modelo, serie, estado}
        origen_data: Dict con {sede, piso, ubicacion, usuario}
        destino_data: Dict con {sede, piso, ubicacion, usuario}
        fecha: Fecha del traslado (opcional, usa hoy si no se especifica)

    Returns:
        BytesIO con el archivo Excel
    """
    import os
    from copy import copy
    from openpyxl import load_workbook
    from django.conf import settings

    # Buscar la plantilla en diferentes ubicaciones
    posibles_rutas = [
        os.path.join(settings.BASE_DIR, 'static', 'templates', 'formato_traslado_plantilla.xlsx'),
        os.path.join(settings.STATIC_ROOT or '', 'templates', 'formato_traslado_plantilla.xlsx'),
        '/var/www/inventario/static/templates/formato_traslado_plantilla.xlsx',
    ]

    plantilla_path = None
    for ruta in posibles_rutas:
        if os.path.exists(ruta):
            plantilla_path = ruta
            break

    if not plantilla_path:
        # Si no se encuentra la plantilla, usar el método anterior simplificado
        return _generar_formato_traslado_simple(items_data, origen_data, destino_data, fecha)

    # Cargar la plantilla
    wb = load_workbook(plantilla_path)
    ws = wb.active

    # === LLENAR FECHA ===
    if fecha:
        fecha_str = fecha.strftime('%d/%m/%Y') if hasattr(fecha, 'strftime') else str(fecha)
    else:
        fecha_str = datetime.now().strftime('%d/%m/%Y')
    ws['F3'] = fecha_str

    # === LIMPIAR FILAS DE ITEMS EXISTENTES (filas 17-32) ===
    # Columnas: E=código, L=descripción, V=marca, AC=modelo, AJ=serie, AP=estado, AU=cantidad, AZ=observaciones
    for row in range(17, 33):
        ws[f'E{row}'] = None
        ws[f'L{row}'] = None
        ws[f'V{row}'] = None
        ws[f'AC{row}'] = None
        ws[f'AJ{row}'] = None
        ws[f'AP{row}'] = None
        ws[f'AU{row}'] = None
        ws[f'AZ{row}'] = None

    # === LLENAR ITEMS ===
    for i, item in enumerate(items_data[:16]):  # Máximo 16 items
        row = 17 + i
        ws[f'E{row}'] = item.get('codigo_utp', '')
        ws[f'L{row}'] = item.get('descripcion', '')
        ws[f'V{row}'] = item.get('marca', '')
        ws[f'AC{row}'] = item.get('modelo', '')
        ws[f'AJ{row}'] = item.get('serie', '')
        ws[f'AP{row}'] = item.get('estado', 'OPTIMO')
        ws[f'AU{row}'] = 1  # Cantidad siempre 1
        ws[f'AZ{row}'] = item.get('observaciones', '')

    # === LLENAR DATOS DE ORIGEN ===
    ws['M42'] = origen_data.get('sede', '')
    ws['M43'] = origen_data.get('piso', '')
    ws['M44'] = origen_data.get('ubicacion', '')
    ws['M45'] = origen_data.get('usuario', '')

    # === LLENAR DATOS DE DESTINO ===
    ws['AL42'] = destino_data.get('sede', '')
    ws['AL43'] = destino_data.get('piso', '')
    ws['AL44'] = destino_data.get('ubicacion', '')
    ws['AL45'] = destino_data.get('usuario', '')

    # Guardar en BytesIO
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    return buffer


def _generar_formato_traslado_simple(items_data, origen_data, destino_data, fecha=None):
    """
    Genera un formato de traslado simplificado cuando no hay plantilla disponible.
    """
    from openpyxl.styles import Border, Side, Alignment, Font, PatternFill

    wb = Workbook()
    ws = wb.active
    ws.title = "Formato Traslado"

    # Estilos
    titulo_font = Font(size=14, bold=True)
    header_font = Font(size=10, bold=True)
    normal_font = Font(size=9)
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    header_fill = PatternFill(start_color='D9D9D9', end_color='D9D9D9', fill_type='solid')

    # Título
    ws.merge_cells('A1:H1')
    ws['A1'] = 'FORMATO DE CONTROL DE ACTIVOS'
    ws['A1'].font = titulo_font
    ws['A1'].alignment = Alignment(horizontal='center')

    # Fecha
    ws['A3'] = 'FECHA:'
    ws['A3'].font = header_font
    if fecha:
        ws['B3'] = fecha.strftime('%d/%m/%Y') if hasattr(fecha, 'strftime') else str(fecha)
    else:
        ws['B3'] = datetime.now().strftime('%d/%m/%Y')

    # Encabezados de tabla
    headers = ['ITEM', 'CÓDIGO', 'DESCRIPCIÓN', 'MARCA', 'MODELO', 'SERIE', 'ESTADO', 'CANT.']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=5, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = thin_border
        cell.alignment = Alignment(horizontal='center')

    # Datos de items
    for i, item in enumerate(items_data[:16]):
        row = 6 + i
        ws.cell(row=row, column=1, value=i+1).border = thin_border
        ws.cell(row=row, column=2, value=item.get('codigo_utp', '')).border = thin_border
        ws.cell(row=row, column=3, value=item.get('descripcion', '')).border = thin_border
        ws.cell(row=row, column=4, value=item.get('marca', '')).border = thin_border
        ws.cell(row=row, column=5, value=item.get('modelo', '')).border = thin_border
        ws.cell(row=row, column=6, value=item.get('serie', '')).border = thin_border
        ws.cell(row=row, column=7, value=item.get('estado', 'OPTIMO')).border = thin_border
        ws.cell(row=row, column=8, value=1).border = thin_border

    # Origen y Destino
    row_base = 6 + len(items_data) + 2

    ws.cell(row=row_base, column=1, value='DATOS DE ORIGEN').font = header_font
    ws.cell(row=row_base, column=5, value='DATOS DE DESTINO').font = header_font

    campos = [
        ('SEDE:', 'sede'),
        ('PISO:', 'piso'),
        ('UBICACIÓN:', 'ubicacion'),
        ('USUARIO:', 'usuario'),
    ]

    for i, (label, key) in enumerate(campos):
        row = row_base + 1 + i
        ws.cell(row=row, column=1, value=label).font = header_font
        ws.cell(row=row, column=2, value=origen_data.get(key, ''))
        ws.cell(row=row, column=5, value=label).font = header_font
        ws.cell(row=row, column=6, value=destino_data.get(key, ''))

    # Ajustar anchos
    ws.column_dimensions['A'].width = 10
    ws.column_dimensions['B'].width = 15
    ws.column_dimensions['C'].width = 25
    ws.column_dimensions['D'].width = 12
    ws.column_dimensions['E'].width = 20
    ws.column_dimensions['F'].width = 15
    ws.column_dimensions['G'].width = 10
    ws.column_dimensions['H'].width = 8

    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    return buffer
