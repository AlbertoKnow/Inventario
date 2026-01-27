"""
Comando para actualizar especificaciones desde Excel para items existentes.
"""

from django.core.management.base import BaseCommand
from django.db import transaction
from openpyxl import load_workbook
from productos.models import (
    Item, MarcaEquipo, ModeloEquipo, ProcesadorEquipo, EspecificacionesSistemas
)
import re


class Command(BaseCommand):
    help = 'Actualiza especificaciones de items existentes desde Excel'

    def add_arguments(self, parser):
        parser.add_argument('archivo', type=str)
        parser.add_argument('--hoja', type=str, default='INVENTARIO GENERAL SEDE PARRA')

    def handle(self, *args, **options):
        archivo = options['archivo']
        hoja = options['hoja']

        wb = load_workbook(archivo, read_only=True)
        ws = wb[hoja]

        # Cache
        cache_marcas = {m.nombre.upper(): m for m in MarcaEquipo.objects.all()}
        cache_modelos = {}
        cache_procs = {p.nombre.upper(): p for p in ProcesadorEquipo.objects.all()}

        # Pre-cargar modelos
        for mod in ModeloEquipo.objects.select_related('marca').all():
            key = f'{mod.marca.nombre.upper()}_{mod.nombre}'
            cache_modelos[key] = mod

        stats = {'actualizados': 0, 'omitidos': 0, 'errores': 0, 'marcas': 0, 'modelos': 0, 'procs': 0}

        COL = {
            'so': 2, 'marca': 3, 'modelo': 4, 'inventario': 5,
            'procesador': 9, 'generacion': 10, 'ram': 11, 'hdd': 12, 'ssd': 13,
        }

        rows = list(ws.iter_rows(min_row=2, values_only=True))
        self.stdout.write(f'Total filas: {len(rows)}')

        with transaction.atomic():
            for idx, row in enumerate(rows, start=2):
                try:
                    codigo_utp = str(row[COL['inventario']] or '').strip()
                    if not codigo_utp or codigo_utp == 'None':
                        continue

                    if len(codigo_utp) > 20:
                        codigo_utp = codigo_utp[:20]

                    # Buscar item
                    try:
                        item = Item.objects.get(codigo_utp=codigo_utp)
                    except Item.DoesNotExist:
                        stats['omitidos'] += 1
                        continue

                    # Ya tiene especificaciones?
                    if hasattr(item, 'especificaciones_sistemas'):
                        stats['omitidos'] += 1
                        continue

                    so = str(row[COL['so']] or '').strip()
                    marca_str = str(row[COL['marca']] or '').strip().upper()
                    modelo_str = str(row[COL['modelo']] or '').strip()
                    procesador_str = str(row[COL['procesador']] or '').strip().upper()
                    generacion = str(row[COL['generacion']] or '').strip()
                    ram = str(row[COL['ram']] or '').strip()
                    hdd = str(row[COL['hdd']] or '').strip()
                    ssd = str(row[COL['ssd']] or '').strip()

                    # Marca
                    marca = None
                    if marca_str and marca_str not in ['NO APLICA', '-', 'None']:
                        if marca_str not in cache_marcas:
                            marca = MarcaEquipo.objects.create(nombre=marca_str)
                            cache_marcas[marca_str] = marca
                            stats['marcas'] += 1
                        marca = cache_marcas[marca_str]

                    # Modelo
                    modelo = None
                    if modelo_str and modelo_str not in ['NO APLICA', '-', 'None'] and marca:
                        modelo_key = f'{marca_str}_{modelo_str}'
                        if modelo_key not in cache_modelos:
                            modelo = ModeloEquipo.objects.create(nombre=modelo_str, marca=marca)
                            cache_modelos[modelo_key] = modelo
                            stats['modelos'] += 1
                        modelo = cache_modelos[modelo_key]

                    # Procesador
                    procesador = None
                    if procesador_str and procesador_str not in ['NO APLICA', '-', 'None']:
                        if procesador_str not in cache_procs:
                            procesador = ProcesadorEquipo.objects.create(nombre=procesador_str)
                            cache_procs[procesador_str] = procesador
                            stats['procs'] += 1
                        procesador = cache_procs[procesador_str]

                    # RAM
                    ram_total = None
                    if ram and ram != 'NO APLICA':
                        m = re.search(r'(\d+)', str(ram))
                        if m:
                            ram_total = int(m.group(1))

                    # Almacenamiento
                    almac_gb = None
                    almac_tipo = ''
                    if ssd and ssd not in ['NO APLICA', '-', 'None']:
                        m = re.search(r'(\d+)', str(ssd))
                        if m:
                            almac_gb = int(m.group(1))
                            almac_tipo = 'SSD'
                    elif hdd and hdd not in ['NO APLICA', '-', 'None']:
                        m = re.search(r'(\d+)', str(hdd))
                        if m:
                            almac_gb = int(m.group(1))
                            almac_tipo = 'HDD'

                    # Crear especificaciones
                    EspecificacionesSistemas.objects.create(
                        item=item,
                        marca_equipo=marca,
                        modelo_equipo=modelo,
                        procesador_equipo=procesador,
                        generacion_procesador=generacion if generacion != 'NO APLICA' else '',
                        ram_total_gb=ram_total,
                        almacenamiento_gb=almac_gb,
                        almacenamiento_tipo=almac_tipo,
                        sistema_operativo=so if so != 'NO APLICA' else '',
                    )
                    stats['actualizados'] += 1

                    if idx % 200 == 0:
                        self.stdout.write(f'  {idx}...')

                except Exception as e:
                    self.stderr.write(f'  ERROR fila {idx}: {e}')
                    stats['errores'] += 1

        wb.close()

        self.stdout.write('')
        self.stdout.write('=' * 40)
        for k, v in stats.items():
            self.stdout.write(f'{k}: {v}')
