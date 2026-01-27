"""
Comando para importar inventario desde archivo Excel.
Uso: python manage.py importar_inventario_excel /path/to/file.xlsx
"""

from django.core.management.base import BaseCommand
from django.db import transaction
from openpyxl import load_workbook
from productos.models import (
    Campus, Sede, Pabellon, Ambiente, Area, TipoItem, Item,
    MarcaEquipo, ModeloEquipo, ProcesadorEquipo, EspecificacionesSistemas
)
from django.contrib.auth.models import User
import re


class Command(BaseCommand):
    help = 'Importa inventario desde archivo Excel'

    def add_arguments(self, parser):
        parser.add_argument('archivo', type=str, help='Ruta al archivo Excel')
        parser.add_argument('--dry-run', action='store_true', help='Simular sin guardar')
        parser.add_argument('--hoja', type=str, default='INVENTARIO GENERAL SEDE PARRA', help='Nombre de la hoja')

    def handle(self, *args, **options):
        archivo = options['archivo']
        dry_run = options['dry_run']
        hoja = options['hoja']

        self.stdout.write(f'Leyendo archivo: {archivo}')
        self.stdout.write(f'Hoja: {hoja}')
        self.stdout.write(f'Modo: {"SIMULACIÓN" if dry_run else "REAL"}')
        self.stdout.write('')

        try:
            wb = load_workbook(archivo, read_only=True)
            ws = wb[hoja]
        except Exception as e:
            self.stderr.write(f'Error al abrir archivo: {e}')
            return

        # Obtener área de Sistemas
        try:
            area_sistemas = Area.objects.get(codigo='sistemas')
        except Area.DoesNotExist:
            self.stderr.write('ERROR: No existe el área "sistemas"')
            return

        # Obtener sede Parra 1
        try:
            sede_parra = Sede.objects.get(nombre__icontains='Parra 1')
        except Sede.DoesNotExist:
            self.stderr.write('ERROR: No existe la sede "Parra 1"')
            return

        self.stdout.write(f'Área: {area_sistemas.nombre}')
        self.stdout.write(f'Sede: {sede_parra.nombre}')
        self.stdout.write('')

        # Mapeo de columnas (índice 0-based)
        COL = {
            'equipo': 0,        # A - Nombre del equipo
            'tipo': 1,          # B - Tipo de Equipo
            'so': 2,            # C - Sistema Operativo
            'marca': 3,         # D - Marca
            'modelo': 4,        # E - Modelo
            'inventario': 5,    # F - Código UTP
            'serie': 6,         # G - Serie
            'procesador': 9,    # J - Procesador
            'generacion': 10,   # K - Generación
            'ram': 11,          # L - RAM
            'hdd': 12,          # M - HDD
            'ssd': 13,          # N - SSD
            'usuario': 14,      # O - Usuario
            'cod_ambiente': 19, # T - Código de Ambiente
            'ambiente': 20,     # U - Nombre de ambiente
            'pabellon': 21,     # V - Pabellón
            'piso': 23,         # X - Nro de piso
            'estado': 25,       # Z - Estado
        }

        # Mapeo de tipos de item
        TIPO_MAP = {
            'MONITOR': 'Monitor',
            'DESKTOP': 'DESKTOP',
            'MINI': 'MINI',
            'PROYECTOR': 'PROYECTOR',
            'LAPTOP': 'LAPTOP',
            'ALL IN ONE': 'ALL IN ONE',
            'MULTIFUNCIONAL': 'MULTIFUNCIONAL',
            'MINIX': 'MINI',
            'HUB': 'HUB',
            'TABLET': 'TABLET',
            'HUELLERO': 'HUELLERO',
            'TOUCH': 'TOUCH',
            'OPSCAN': 'OPSCAN',
        }

        # Mapeo de estados
        ESTADO_MAP = {
            'INSTALADO': 'activo',
            'BACKUP': 'almacenado',
            'DAÑADO': 'dañado',
            'DANADO': 'dañado',
        }

        # Contadores
        stats = {
            'pabellones_creados': 0,
            'ambientes_creados': 0,
            'tipos_creados': 0,
            'marcas_creadas': 0,
            'modelos_creados': 0,
            'procesadores_creados': 0,
            'items_creados': 0,
            'items_omitidos': 0,
            'errores': 0,
        }

        # Cache para evitar queries repetidas
        cache_pabellones = {p.nombre: p for p in Pabellon.objects.filter(sede=sede_parra)}
        cache_ambientes = {}
        cache_tipos = {t.nombre.upper(): t for t in TipoItem.objects.all()}
        cache_marcas = {}
        cache_modelos = {}
        cache_procesadores = {}

        # Leer todas las filas
        rows = list(ws.iter_rows(min_row=2, values_only=True))
        total_rows = len(rows)
        self.stdout.write(f'Total filas a procesar: {total_rows}')
        self.stdout.write('')

        with transaction.atomic():
            for idx, row in enumerate(rows, start=2):
                try:
                    # Extraer datos
                    equipo = str(row[COL['equipo']] or '').strip()
                    tipo_str = str(row[COL['tipo']] or '').strip().upper()
                    so = str(row[COL['so']] or '').strip()
                    marca_str = str(row[COL['marca']] or '').strip().upper()
                    modelo_str = str(row[COL['modelo']] or '').strip()
                    codigo_utp = str(row[COL['inventario']] or '').strip()
                    serie = str(row[COL['serie']] or '').strip()
                    procesador_str = str(row[COL['procesador']] or '').strip().upper()
                    generacion = str(row[COL['generacion']] or '').strip()
                    ram = str(row[COL['ram']] or '').strip()
                    hdd = str(row[COL['hdd']] or '').strip()
                    ssd = str(row[COL['ssd']] or '').strip()
                    usuario = str(row[COL['usuario']] or '').strip()
                    cod_ambiente = str(row[COL['cod_ambiente']] or '').strip()
                    nombre_ambiente = str(row[COL['ambiente']] or '').strip()
                    pabellon_str = str(row[COL['pabellon']] or '').strip().upper()
                    piso = str(row[COL['piso']] or '').strip()
                    estado_str = str(row[COL['estado']] or '').strip().upper()

                    # Validar código UTP
                    if not codigo_utp or codigo_utp == 'None':
                        self.stdout.write(f'  Fila {idx}: Sin código UTP, omitiendo')
                        stats['items_omitidos'] += 1
                        continue

                    # Verificar si ya existe
                    if Item.objects.filter(codigo_utp=codigo_utp).exists():
                        self.stdout.write(f'  Fila {idx}: {codigo_utp} ya existe, omitiendo')
                        stats['items_omitidos'] += 1
                        continue

                    # Limpiar pabellón (algunos tienen fórmulas rotas)
                    if pabellon_str.startswith('=') or pabellon_str == '-' or len(pabellon_str) > 2:
                        pabellon_str = 'X'  # Pabellón desconocido

                    # Obtener o crear pabellón
                    if pabellon_str not in cache_pabellones:
                        if not dry_run:
                            pab, created = Pabellon.objects.get_or_create(
                                nombre=pabellon_str,
                                sede=sede_parra,
                                defaults={'pisos': 5, 'activo': True}
                            )
                            cache_pabellones[pabellon_str] = pab
                            if created:
                                stats['pabellones_creados'] += 1
                                self.stdout.write(f'  Pabellón creado: {pabellon_str}')
                        else:
                            cache_pabellones[pabellon_str] = None
                            stats['pabellones_creados'] += 1

                    pabellon = cache_pabellones.get(pabellon_str)

                    # Obtener o crear ambiente
                    ambiente_key = f'{pabellon_str}_{nombre_ambiente}'
                    if ambiente_key not in cache_ambientes:
                        if not dry_run and pabellon:
                            # Determinar piso numérico
                            try:
                                piso_num = int(piso) if piso and piso.isdigit() else 1
                            except:
                                piso_num = 1

                            amb, created = Ambiente.objects.get_or_create(
                                nombre=nombre_ambiente,
                                pabellon=pabellon,
                                defaults={
                                    'piso': piso_num,
                                    'codigo': cod_ambiente if cod_ambiente else nombre_ambiente,
                                    'activo': True
                                }
                            )
                            cache_ambientes[ambiente_key] = amb
                            if created:
                                stats['ambientes_creados'] += 1
                        else:
                            cache_ambientes[ambiente_key] = None
                            if ambiente_key not in cache_ambientes:
                                stats['ambientes_creados'] += 1

                    ambiente = cache_ambientes.get(ambiente_key)

                    # Obtener o crear tipo de item
                    tipo_nombre = TIPO_MAP.get(tipo_str, tipo_str)
                    if tipo_nombre.upper() not in cache_tipos:
                        if not dry_run:
                            tipo_obj, created = TipoItem.objects.get_or_create(
                                nombre=tipo_nombre,
                                defaults={'descripcion': f'Tipo: {tipo_nombre}'}
                            )
                            cache_tipos[tipo_nombre.upper()] = tipo_obj
                            if created:
                                stats['tipos_creados'] += 1
                                self.stdout.write(f'  Tipo creado: {tipo_nombre}')
                        else:
                            cache_tipos[tipo_nombre.upper()] = None
                            stats['tipos_creados'] += 1

                    tipo_item = cache_tipos.get(tipo_nombre.upper())

                    # Obtener o crear marca
                    marca = None
                    if marca_str and marca_str != 'NO APLICA':
                        if marca_str not in cache_marcas:
                            if not dry_run:
                                marca_obj, created = MarcaEquipo.objects.get_or_create(
                                    nombre=marca_str
                                )
                                cache_marcas[marca_str] = marca_obj
                                if created:
                                    stats['marcas_creadas'] += 1
                            else:
                                cache_marcas[marca_str] = None
                                stats['marcas_creadas'] += 1
                        marca = cache_marcas.get(marca_str)

                    # Obtener o crear modelo
                    modelo = None
                    if modelo_str and modelo_str != 'NO APLICA':
                        modelo_key = f'{marca_str}_{modelo_str}'
                        if modelo_key not in cache_modelos:
                            if not dry_run and marca:
                                modelo_obj, created = ModeloEquipo.objects.get_or_create(
                                    nombre=modelo_str,
                                    marca=marca
                                )
                                cache_modelos[modelo_key] = modelo_obj
                                if created:
                                    stats['modelos_creados'] += 1
                            else:
                                cache_modelos[modelo_key] = None
                                stats['modelos_creados'] += 1
                        modelo = cache_modelos.get(modelo_key)

                    # Obtener o crear procesador
                    procesador = None
                    if procesador_str and procesador_str not in ['NO APLICA', '-', 'None']:
                        if procesador_str not in cache_procesadores:
                            if not dry_run:
                                proc_obj, created = ProcesadorEquipo.objects.get_or_create(
                                    nombre=procesador_str
                                )
                                cache_procesadores[procesador_str] = proc_obj
                                if created:
                                    stats['procesadores_creados'] += 1
                            else:
                                cache_procesadores[procesador_str] = None
                                stats['procesadores_creados'] += 1
                        procesador = cache_procesadores.get(procesador_str)

                    # Mapear estado
                    estado = ESTADO_MAP.get(estado_str, 'activo')

                    # Crear item
                    if not dry_run:
                        item = Item.objects.create(
                            codigo_utp=codigo_utp,
                            serie=serie if serie and serie != 'None' else '',
                            nombre=equipo,
                            area=area_sistemas,
                            tipo_item=tipo_item,
                            ambiente=ambiente,
                            estado=estado,
                        )

                        # Crear especificaciones si tiene datos técnicos
                        if marca or modelo or procesador or (ram and ram != 'NO APLICA'):
                            # Parsear RAM
                            ram_gb = None
                            if ram and ram != 'NO APLICA':
                                match = re.search(r'(\d+)', str(ram))
                                if match:
                                    ram_gb = int(match.group(1))

                            # Parsear almacenamiento
                            hdd_val = None
                            ssd_val = None
                            if hdd and hdd != 'NO APLICA' and hdd != '-':
                                match = re.search(r'(\d+)', str(hdd))
                                if match:
                                    hdd_val = match.group(1) + ' GB'
                            if ssd and ssd != 'NO APLICA' and ssd != '-':
                                match = re.search(r'(\d+)', str(ssd))
                                if match:
                                    ssd_val = match.group(1) + ' GB'

                            EspecificacionesSistemas.objects.create(
                                item=item,
                                marca_equipo=marca,
                                modelo_equipo=modelo,
                                procesador_equipo=procesador,
                                generacion_procesador=generacion if generacion != 'NO APLICA' else '',
                                ram_gb=ram_gb,
                                almacenamiento_principal=ssd_val or hdd_val or '',
                                almacenamiento_secundario=hdd_val if ssd_val else '',
                                sistema_operativo=so if so != 'NO APLICA' else '',
                            )

                        stats['items_creados'] += 1

                    else:
                        stats['items_creados'] += 1

                    # Progreso cada 100 filas
                    if idx % 100 == 0:
                        self.stdout.write(f'  Procesadas {idx}/{total_rows} filas...')

                except Exception as e:
                    self.stderr.write(f'  ERROR en fila {idx}: {e}')
                    stats['errores'] += 1

            if dry_run:
                self.stdout.write('')
                self.stdout.write('=== MODO SIMULACIÓN - NO SE GUARDARON CAMBIOS ===')
                raise Exception('Dry run - rollback')

        wb.close()

        # Resumen
        self.stdout.write('')
        self.stdout.write('=' * 50)
        self.stdout.write('RESUMEN DE IMPORTACIÓN')
        self.stdout.write('=' * 50)
        self.stdout.write(f'Pabellones creados: {stats["pabellones_creados"]}')
        self.stdout.write(f'Ambientes creados: {stats["ambientes_creados"]}')
        self.stdout.write(f'Tipos creados: {stats["tipos_creados"]}')
        self.stdout.write(f'Marcas creadas: {stats["marcas_creadas"]}')
        self.stdout.write(f'Modelos creados: {stats["modelos_creados"]}')
        self.stdout.write(f'Procesadores creados: {stats["procesadores_creados"]}')
        self.stdout.write(f'Items creados: {stats["items_creados"]}')
        self.stdout.write(f'Items omitidos: {stats["items_omitidos"]}')
        self.stdout.write(f'Errores: {stats["errores"]}')
        self.stdout.write('=' * 50)
